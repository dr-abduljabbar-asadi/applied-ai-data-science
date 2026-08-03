"""
utils.py

Utility functions extracted verbatim from code_step_by_step.ipynb.

This module exists purely to keep the notebook clean — every function
defined inside the notebook was moved here in its original order, with
no changes to names, parameters, return values, docstrings, or internal
logic. The notebook now imports these functions instead of defining them.
"""

# ── Standard library ─────────────────────────────────────────────
import json
import math
import os
from typing import Union

# ── Core data & numerics ─────────────────────────────────────────
import numpy as np
import pandas as pd

# ── Plotting ─────────────────────────────────────────────────────
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import matplotlib.colors as mcolors
import matplotlib.gridspec as gridspec
import matplotlib.ticker as ticker
from matplotlib.colorbar import ColorbarBase
from matplotlib.colors import Normalize
from matplotlib.lines import Line2D
from matplotlib.patches import Patch
import seaborn as sns
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# ── Scientific / statistics ──────────────────────────────────────
from scipy import stats
from scipy.stats import gaussian_kde, probplot, randint
from scipy.cluster import hierarchy

# ── Machine learning (scikit-learn) ──────────────────────────────
from sklearn.cluster import KMeans
from sklearn.ensemble import RandomForestRegressor
from sklearn.feature_selection import VarianceThreshold
from sklearn.impute import SimpleImputer
from sklearn.manifold import TSNE
from sklearn.metrics import (
    explained_variance_score,
    max_error,
    mean_absolute_error,
    mean_absolute_percentage_error,
    mean_squared_error,
    r2_score,
)
from sklearn.model_selection import (
    KFold,
    RandomizedSearchCV,
    cross_val_score,
    cross_validate,
    train_test_split,
)
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import (
    KBinsDiscretizer,
    LabelEncoder,
    StandardScaler,
)

# ── Network analysis ─────────────────────────────────────────────
import networkx as nx
from pyvis.network import Network

# ── Pathing & I/O ────────────────────────────────────────────────
from pathlib import Path
from typing import Optional, Tuple

# ── Persistence ──────────────────────────────────────────────────
import joblib

# ── Warnings ─────────────────────────────────────────────────────
import warnings
warnings.filterwarnings("ignore")

# ── Optional SHAP (used only by shap_analysis) ───────────────────
try:
    import shap
    HAS_SHAP = True
except ImportError:
    HAS_SHAP = False


def load_co2_data(file_path):
    """
    Load and preprocess CO2 data from a CSV file.

    Parameters
    ----------
    file_path : str
        Path to the CSV file containing the CO2 data.

    Returns
    -------
    data_table : pandas.DataFrame
        Processed DataFrame with:
        - First column renamed to 'Dateiname'
        - Leading/trailing spaces removed from 'Dateiname'
        - 'Dateiname' set as index
        - Missing values handled ("" , " ", "0" treated as NaN)

    Notes
    -----
    - Assumes the file is comma-separated.
    - Uses UTF-8 encoding.
    - Designed for datasets with sparse columns.
    """

    # Load data
    data_table = pd.read_csv(
        file_path,
        sep=",",
        encoding="utf-8",
        low_memory=False,
        na_values=["", " ", "0"]
    )

    # Rename first column
    data_table.rename(columns={data_table.columns[0]: "Dateiname"}, inplace=True)

    # Clean and set index
    data_table["Dateiname"] = data_table["Dateiname"].str.strip()
    data_table.set_index("Dateiname", inplace=True)

    print(f"✅ Loaded: {data_table.shape[0]} rows × {data_table.shape[1]} columns")

    return data_table


def explore_dataframe(data_table):
    """
    Generate a comprehensive overview of a pandas DataFrame.

    Parameters
    ----------
    data_table : pandas.DataFrame
        The DataFrame to analyze.

    Returns
    -------
    summary : dict
        Dictionary containing key DataFrame insights:
        - 'shape' : tuple
        - 'dtypes' : pandas.Series
        - 'dtypes_count' : pandas.DataFrame
        - 'describe_numeric' : pandas.DataFrame
        - 'describe_object' : pandas.DataFrame
        - 'missing_values' : pandas.DataFrame
        - 'duplicate_rows' : int

    Notes
    -----
    - Prints a structured overview of the dataset.
    - Includes both numeric and categorical summaries.
    - Useful for quick exploratory data analysis (EDA).
    """

    print("\n📊 ===== DATAFRAME OVERVIEW =====\n")

    # Shape
    print(f"Shape: {data_table.shape[0]} rows × {data_table.shape[1]} columns\n")

    # Column names
    print(f"Columns ({len(data_table.columns)}):\n")
    for i, col in enumerate(data_table.columns):
        print(f"  {i:3d}  {col}")
    
    print("\n🧾 ===== INFO =====")
    data_table.info()

    # Data types
    print("\n🔎 ===== DATA TYPES COUNT =====")
    dtypes_count = data_table.dtypes.value_counts().rename("count").to_frame()
    print(dtypes_count)

    # Missing values
    print("\n❗ ===== MISSING VALUES =====")
    missing = data_table.isna().sum().to_frame(name="missing_count")
    missing["missing_%"] = (missing["missing_count"] / len(data_table)) * 100
    missing = missing.sort_values(by="missing_count", ascending=False)
    print(missing.head(20))  # show top 20

    # Duplicates
    duplicates = data_table.duplicated().sum()
    print(f"\n🔁 Duplicate rows: {duplicates}")

    # Describe numeric
    print("\n📈 ===== NUMERIC SUMMARY =====")
    describe_numeric = data_table.describe()
    print(describe_numeric)

    # Describe categorical
    print("\n📝 ===== CATEGORICAL SUMMARY =====")
    describe_object = data_table.describe(include=["object", "category"])
    print(describe_object)

    # Collect results
    summary = {
        "shape": data_table.shape,
        "dtypes": data_table.dtypes,
        "dtypes_count": dtypes_count,
        "describe_numeric": describe_numeric,
        "describe_object": describe_object,
        "missing_values": missing,
        "duplicate_rows": duplicates
    }

    return summary


def detect_column_types(
    df: pd.DataFrame,
    numeric_threshold: float = 1.0,
    verbose: bool = True,
    convert: bool = True,
    excel_path: "str | None" = "column_type_report.xlsx",
) -> pd.DataFrame:
    """
    Detect and optionally convert each column to its correct dtype.

    Rule
    ----
    A column is classified as NUMERIC if the fraction of non-null
    values that can be successfully converted to float meets
    ``numeric_threshold``. Default is 1.0 (every non-null value
    must be numeric).

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame.
    numeric_threshold : float, optional
        Min. fraction of non-null values that must be numeric for a
        column to be classified as NUMERIC. Default 1.0.
    verbose : bool, optional
        Print the report to stdout. Default True.
    convert : bool, optional
        If True, return a DataFrame with detected types applied.
        If False, return only the report. Default True.
    excel_path : str or None, optional
        If a path is given, save a styled Excel workbook with one
        sheet per category (Summary, NUMERIC, STRING, EMPTY, Full).
        Pass ``None`` to skip. Default ``"column_type_report.xlsx"``.

    Returns
    -------
    pandas.DataFrame
        Either the converted DataFrame (``convert=True``) or the
        report DataFrame (``convert=False``).

    Examples
    --------
    >>> df_clean = detect_column_types(df)
    >>> df_clean = detect_column_types(df, numeric_threshold=0.9)
    >>> report  = detect_column_types(df, convert=False, excel_path=None)
    """

    # ── Analyse each column ───────────────────────────────────────
    report = []
    for col in df.columns:
        series  = df[col]
        n_total = len(series)
        n_null  = int(series.isna().sum())
        n_valid = n_total - n_null
        null_pct = round((n_null / n_total) * 100, 2) if n_total else 0.0

        if n_valid == 0:
            report.append({
                "column"        : col,
                "detected_type" : "EMPTY",
                "numeric_ratio" : 0.0,
                "n_total"       : n_total,
                "n_null"        : n_null,
                "null_pct"      : 100.0,
                "n_numeric"     : 0,
                "n_string"      : 0,
                "sample_values" : [],
            })
            continue

        n_numeric = 0
        n_string  = 0
        samples   = []
        for val in series.dropna().values:
            str_val = str(val).strip()
            try:
                float(str_val.replace(",", "."))
                n_numeric += 1
            except (ValueError, TypeError):
                n_string += 1
                if len(samples) < 3:
                    samples.append(str_val)

        numeric_ratio = n_numeric / n_valid
        detected_type = "NUMERIC" if numeric_ratio >= numeric_threshold else "STRING"

        report.append({
            "column"        : col,
            "detected_type" : detected_type,
            "numeric_ratio" : round(numeric_ratio, 4),
            "n_total"       : n_total,
            "n_null"        : n_null,
            "null_pct"      : null_pct,
            "n_numeric"     : n_numeric,
            "n_string"      : n_string,
            "sample_values" : samples,
        })

    report_df = pd.DataFrame(report)

    # ── Verbose stdout report ─────────────────────────────────────
    if verbose:
        n_num = (report_df["detected_type"] == "NUMERIC").sum()
        n_str = (report_df["detected_type"] == "STRING").sum()
        n_emp = (report_df["detected_type"] == "EMPTY").sum()
        sep = "─" * 65
        print(f"\n{sep}\n  Column Type Detection Report\n{sep}")
        print(f"  Total columns    : {len(df.columns)}")
        print(f"  → NUMERIC        : {n_num}")
        print(f"  → STRING         : {n_str}")
        print(f"  → EMPTY          : {n_emp}")
        print(f"  Threshold        : {numeric_threshold * 100:.0f}% numeric values required")
        print(f"{sep}\n")

    # ── Save styled Excel (one sheet per category) ────────────────
    if excel_path:
        from pathlib import Path
        out_path = Path(excel_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)

        report_excel = report_df.copy()
        report_excel["sample_values"] = report_excel["sample_values"].apply(
            lambda lst: ", ".join(map(str, lst)) if lst else ""
        )

        summary_df = pd.DataFrame({
            "metric": [
                "Total columns", "NUMERIC columns", "STRING columns",
                "EMPTY columns", "Numeric threshold",
                "Avg. null %", "Max. null %",
            ],
            "value": [
                len(df.columns),
                int((report_df["detected_type"] == "NUMERIC").sum()),
                int((report_df["detected_type"] == "STRING").sum()),
                int((report_df["detected_type"] == "EMPTY").sum()),
                f"{numeric_threshold * 100:.0f}%",
                f"{report_df['null_pct'].mean():.2f}%",
                f"{report_df['null_pct'].max():.2f}%",
            ],
        })

        sheets = {
            "Summary": summary_df,
            "NUMERIC": report_excel[report_excel["detected_type"] == "NUMERIC"]
                       .sort_values("null_pct"),
            "STRING":  report_excel[report_excel["detected_type"] == "STRING"]
                       .sort_values("null_pct"),
            "EMPTY":   report_excel[report_excel["detected_type"] == "EMPTY"],
            "Full":    report_excel.sort_values(["detected_type", "null_pct"]),
        }

        sheet_titles = {
            "Summary": "Column Type Detection — Summary",
            "NUMERIC": "Numeric Columns Report",
            "STRING":  "String / Categorical Columns Report",
            "EMPTY":   "Empty (All-Null) Columns Report",
            "Full":    "Full Column Type Detection Report",
        }

        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            TITLE_FONT   = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            TITLE_FILL   = PatternFill("solid", fgColor="1F4E78")
            HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            HEADER_FILL  = PatternFill("solid", fgColor="305496")
            ALT_ROW_FILL = PatternFill("solid", fgColor="F2F2F2")
            CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
            LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
            THIN         = Side(style="thin", color="BFBFBF")
            BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                for sheet_name, sheet_df in sheets.items():
                    title = sheet_titles.get(sheet_name, sheet_name)
                    df_w = sheet_df if not sheet_df.empty else pd.DataFrame(columns=sheet_df.columns)

                    df_w.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
                    ws = writer.sheets[sheet_name]
                    n_cols = max(len(df_w.columns), 1)

                    # Title row
                    ws.cell(row=1, column=1, value=title)
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
                    tc = ws.cell(row=1, column=1)
                    tc.font = TITLE_FONT
                    tc.fill = TITLE_FILL
                    tc.alignment = CENTER
                    ws.row_dimensions[1].height = 26

                    # Header row (row 3)
                    for c in range(1, n_cols + 1):
                        cell = ws.cell(row=3, column=c)
                        cell.font = HEADER_FONT
                        cell.fill = HEADER_FILL
                        cell.alignment = CENTER
                        cell.border = BORDER
                    ws.row_dimensions[3].height = 22

                    # Data rows
                    for r in range(4, 4 + len(df_w)):
                        for c in range(1, n_cols + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.border = BORDER
                            cell.alignment = LEFT if c == 1 else CENTER
                            if (r - 4) % 2 == 1:
                                cell.fill = ALT_ROW_FILL

                    # Auto-fit column widths
                    for c_idx, col_name in enumerate(df_w.columns, start=1):
                        if df_w.empty:
                            max_len = len(str(col_name))
                        else:
                            max_len = max(
                                df_w[col_name].astype(str).map(len).max(),
                                len(str(col_name)),
                            )
                        ws.column_dimensions[get_column_letter(c_idx)].width = min(
                            max(max_len + 2, 12), 60
                        )

                    ws.freeze_panes = "A4"
                    ws.sheet_view.showGridLines = False

            print(f"  📄 Excel report saved → {out_path.resolve()}")
        except ModuleNotFoundError:
            print("  ⚠️  'openpyxl' not installed — run: pip install openpyxl")

    # ── Convert if requested ──────────────────────────────────────
    if convert:
        df_out = df.copy()
        for row in report:
            if row["detected_type"] == "NUMERIC":
                col = row["column"]
                df_out[col] = pd.to_numeric(
                    df_out[col].astype(str).str.replace(",", "."),
                    errors="coerce",
                )
        print(f"  Converted DataFrame shape : {df_out.shape}")
        print(f"  Numeric columns now       : {df_out.select_dtypes(include='number').shape[1]}")
        print(f"  String  columns now       : {df_out.select_dtypes(include='object').shape[1]}")
        return df_out

    return report_df


def plotly_histogram_searchable(data_table, output_file="histograms_search.html"):
    numeric_cols = (
        data_table
        .select_dtypes(include="number")
        .dropna(axis=1, how="all")
        .columns.tolist()
    )
    print(f"Building histograms for {len(numeric_cols)} columns...")

    all_data = {}
    for col in numeric_cols:
        series = pd.to_numeric(data_table[col], errors="coerce").dropna()
        if len(series) < 2:
            continue
        counts, bin_edges = np.histogram(series, bins=40)
        bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
        bin_width   = bin_edges[1] - bin_edges[0]
        try:
            kde   = gaussian_kde(series, bw_method="scott")
            x_kde = np.linspace(series.min(), series.max(), 300)
            y_kde = kde(x_kde) * len(series) * bin_width
        except Exception:
            x_kde, y_kde = [], []
        all_data[col] = {
            "hist_x": bin_centers.tolist(),
            "hist_y": counts.tolist(),
            "kde_x" : list(x_kde),
            "kde_y" : list(y_kde),
            "mean"  : round(float(series.mean()),   3),
            "median": round(float(series.median()), 3),
            "std"   : round(float(series.std()),    3),
            "skew"  : round(float(series.skew()),   3),
            "n"     : int(len(series)),
            "min"   : round(float(series.min()),    3),
            "max"   : round(float(series.max()),    3),
        }

    first_col = list(all_data.keys())[0]
    data_json = json.dumps(all_data)
    cols_json = json.dumps(list(all_data.keys()))

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Histogram Explorer</title>
<script src="https://cdn.jsdelivr.net/npm/plotly.js-dist@2.27.0/plotly.min.js"></script>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
  background: #f4f6fb; padding: 28px 32px; color: #333;
}}
h2 {{ font-size: 22px; font-weight: 600; color: #1a1a2e; margin-bottom: 4px; }}
.subtitle {{ font-size: 13px; color: #999; margin-bottom: 22px; }}
#chart {{
  background: white; border-radius: 12px;
  box-shadow: 0 1px 6px rgba(0,0,0,0.07); padding: 8px;
}}
.search-section {{
  margin-top: 22px; background: white; border-radius: 12px;
  padding: 20px 22px 22px; box-shadow: 0 1px 6px rgba(0,0,0,0.07);
}}
.search-label {{
  font-size: 13px; font-weight: 600; color: #555;
  margin-bottom: 10px; display: block; letter-spacing: 0.2px;
}}
.search-wrapper {{ position: relative; width: 100%; }}
#col-search {{
  width: 100%; padding: 11px 16px; font-size: 14px;
  border: 1.5px solid #e0e0e0; border-radius: 8px;
  outline: none; background: #fafafa;
  transition: border-color 0.2s, box-shadow 0.2s; color: #333;
}}
#col-search:focus {{
  border-color: #4c8eda; background: white;
  box-shadow: 0 0 0 3px rgba(76,142,218,0.12);
}}
#col-search::placeholder {{ color: #bbb; }}
#dropdown-list {{
  position: absolute; top: calc(100% + 6px); left: 0; right: 0;
  background: white; border: 1.5px solid #e0e0e0; border-radius: 10px;
  max-height: 280px; overflow-y: auto; z-index: 9999; display: none;
  box-shadow: 0 8px 24px rgba(0,0,0,0.13);
}}
.dropdown-item {{
  padding: 10px 16px; font-size: 13px; cursor: pointer;
  border-bottom: 1px solid #f5f5f5; transition: background 0.12s;
  white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
}}
.dropdown-item:last-child {{ border-bottom: none; }}
.dropdown-item:hover {{ background: #eef4fd; color: #4c8eda; }}
.dropdown-item.selected {{ background: #e8f0fb; color: #4c8eda; font-weight: 600; }}
.dropdown-item.active {{ background: #ddeeff; }}
.dropdown-item mark {{
  background: #fff176; border-radius: 2px;
  padding: 0 1px; font-weight: 700; color: inherit;
}}
.no-results {{ padding: 14px 16px; font-size: 13px; color: #bbb; font-style: italic; }}
.stats-bar {{ display: flex; flex-wrap: wrap; gap: 10px; margin-top: 18px; }}
.stat-card {{
  flex: 1; min-width: 80px; background: #f4f7fb;
  border-radius: 8px; padding: 9px 14px; text-align: center;
}}
.stat-label {{ font-size: 10px; color: #aaa; text-transform: uppercase; letter-spacing: 0.6px; }}
.stat-value {{ font-size: 17px; font-weight: 700; color: #1a1a2e; margin-top: 2px; }}
.badge {{
  display: inline-block; background: #eef4fd; color: #4c8eda;
  border-radius: 20px; padding: 2px 10px; font-size: 12px;
  font-weight: 600; margin-left: 8px; vertical-align: middle;
}}
</style>
</head>
<body>
<h2>Histogram Explorer <span class="badge">{len(all_data)} features</span></h2>
<p class="subtitle">Search or type any column name in the box below the chart</p>
<div id="chart"></div>
<div class="search-section">
  <label class="search-label">Search or select a feature:</label>
  <div class="search-wrapper">
    <input id="col-search" type="text"
      placeholder="Type to search  e.g. SDCOM_WIDTH, KgCO2EQ ..."
      autocomplete="off"/>
    <div id="dropdown-list"></div>
  </div>
  <div class="stats-bar" id="stats-bar"></div>
</div>
<script>
const ALL_DATA = {data_json};
const ALL_COLS = {cols_json};
let activeCol  = "{first_col}";

function buildChart(col) {{
  const d = ALL_DATA[col];
  if (!d) return;
  activeCol = col;
  const maxY = Math.max(...d.hist_y);
  Plotly.react("chart", [
    {{
      type:"bar", x:d.hist_x, y:d.hist_y, name:"Count",
      marker:{{ color:"rgba(76,142,218,0.7)", line:{{color:"rgba(76,142,218,1)",width:0.8}} }},
      hovertemplate:"Value: %{{x:.3f}}<br>Count: %{{y}}<extra></extra>"
    }},
    {{
      type:"scatter", x:d.kde_x, y:d.kde_y, mode:"lines", name:"KDE",
      line:{{color:"#e05c5c",width:2.5}},
      hovertemplate:"KDE: %{{y:.1f}}<extra></extra>"
    }},
    {{
      type:"scatter", x:[d.mean,d.mean], y:[0,maxY*0.95], mode:"lines",
      name:`Mean ${{d.mean}}`, line:{{color:"#f0a500",width:2,dash:"dash"}},
      hovertemplate:`Mean: ${{d.mean}}<extra></extra>`
    }},
    {{
      type:"scatter", x:[d.median,d.median], y:[0,maxY*0.85], mode:"lines",
      name:`Median ${{d.median}}`, line:{{color:"#6ab187",width:2,dash:"dot"}},
      hovertemplate:`Median: ${{d.median}}<extra></extra>`
    }}
  ], {{
    title:{{text:`Distribution of <b>${{col}}</b>`,font:{{size:16}}}},
    xaxis:{{title:col, gridcolor:"#eee", showgrid:true}},
    yaxis:{{title:"Count", gridcolor:"#eee", showgrid:true}},
    plot_bgcolor:"white", paper_bgcolor:"white",
    height:460, bargap:0.05,
    legend:{{orientation:"h",y:-0.18}},
    margin:{{t:60,r:30,l:60,b:60}}
  }}, {{responsive:true}});

  const stats=[["n",d.n.toLocaleString()],["mean",d.mean],["median",d.median],
               ["std",d.std],["skew",d.skew],["min",d.min],["max",d.max]];
  document.getElementById("stats-bar").innerHTML = stats.map(([l,v])=>
    `<div class="stat-card">
       <div class="stat-label">${{l}}</div>
       <div class="stat-value">${{v}}</div>
     </div>`).join("");
  document.getElementById("col-search").value = col;
}}

const searchInput  = document.getElementById("col-search");
const dropdownList = document.getElementById("dropdown-list");
let filteredCols = [...ALL_COLS];
let highlighted  = -1;

function hl(text, q) {{
  if (!q) return text;
  return text.replace(new RegExp(`(${{q.replace(/[.*+?^${{}}()|[\\]\\\\]/g,"\\\\$&")}})`, "gi"),
    "<mark>$1</mark>");
}}

function renderDropdown(q) {{
  q = q.toLowerCase().trim();
  filteredCols = ALL_COLS.filter(c => c.toLowerCase().includes(q));
  highlighted  = -1;
  dropdownList.innerHTML = filteredCols.length === 0
    ? `<div class="no-results">No columns match "${{q}}"</div>`
    : filteredCols.map((col,i) =>
        `<div class="dropdown-item ${{col===activeCol?"selected":""}}" data-col="${{col}}" data-idx="${{i}}">
           ${{hl(col,q)}}
         </div>`).join("");
  dropdownList.querySelectorAll(".dropdown-item").forEach(el => {{
    el.addEventListener("mousedown", e => {{ e.preventDefault(); selectCol(el.dataset.col); }});
  }});
  dropdownList.style.display = "block";
}}

function selectCol(col) {{
  buildChart(col);
  dropdownList.style.display = "none";
}}

function setHL(idx) {{
  const items = dropdownList.querySelectorAll(".dropdown-item");
  items.forEach(el => el.classList.remove("active"));
  if (idx >= 0 && idx < items.length) {{
    items[idx].classList.add("active");
    items[idx].scrollIntoView({{block:"nearest"}});
    highlighted = idx;
  }}
}}

searchInput.addEventListener("focus",  () => renderDropdown(searchInput.value));
searchInput.addEventListener("input",  () => renderDropdown(searchInput.value));
searchInput.addEventListener("keydown", e => {{
  const items = dropdownList.querySelectorAll(".dropdown-item");
  if      (e.key==="ArrowDown") {{ e.preventDefault(); setHL(Math.min(highlighted+1, items.length-1)); }}
  else if (e.key==="ArrowUp")   {{ e.preventDefault(); setHL(Math.max(highlighted-1, 0)); }}
  else if (e.key==="Enter")     {{
    e.preventDefault();
    if (highlighted>=0 && filteredCols[highlighted]) selectCol(filteredCols[highlighted]);
    else if (filteredCols.length>0) selectCol(filteredCols[0]);
  }}
  else if (e.key==="Escape")    {{ dropdownList.style.display="none"; searchInput.blur(); }}
}});
document.addEventListener("click", e => {{
  if (!e.target.closest(".search-wrapper")) dropdownList.style.display = "none";
}});

buildChart("{first_col}");
</script>
</body>
</html>"""

    with open(output_file, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"✅ Saved → {output_file}")
    return output_file


def categorical_waterfall_searchable(data_table, output_file="categorical_waterfall.html", top_n=20):
    """
    Interactive searchable TRUE waterfall chart for categorical columns.
    Each bar starts where the previous one ended — cumulative style.
    NaN shown explicitly in red. Searchable dropdown below the chart.
    """

    cat_cols = data_table.select_dtypes(include=["object", "category", "str"]).columns.tolist()
    for col in data_table.select_dtypes(include="number").columns:
        if 2 <= data_table[col].nunique() <= 20:
            cat_cols.append(col)
    cat_cols = list(dict.fromkeys(cat_cols))
    print(f"Building waterfall charts for {len(cat_cols)} categorical columns...")

    all_data = {}
    for col in cat_cols:
        series    = data_table[col].copy()
        n_total   = len(series)
        n_missing = int(series.isna().sum())
        vc        = series.dropna().astype(str).value_counts()

        if len(vc) > top_n:
            top   = vc.iloc[:top_n]
            other = vc.iloc[top_n:].sum()
            vc    = pd.concat([top, pd.Series({"Other (grouped)": int(other)})])

        labels = list(vc.index)
        counts = [int(v) for v in vc.values]

        if n_missing > 0:
            labels.append("NaN / Missing")
            counts.append(n_missing)

        # ── True waterfall: compute base (start) for each bar ─────
        bases    = []
        running  = 0
        for c in counts:
            bases.append(running)
            running += c

        # ── Add Total bar at the end — starts at 0, full height ───
        labels.append("▌ Total")
        counts.append(n_total)
        bases.append(0)          # total bar always starts at 0

        pcts = [round(c / n_total * 100, 1) for c in counts]
        pcts[-1] = 100.0         # total is always 100%

        # colors
        colors = []
        for lbl in labels:
            if lbl == "NaN / Missing":
                colors.append("#e05c5c")
            elif lbl == "Other (grouped)":
                colors.append("#aaaaaa")
            elif lbl == "▌ Total":
                colors.append("#2d6a4f")  # dark green for total
            else:
                colors.append(None)  # blue gradient assigned in JS

        all_data[col] = {
            "labels"     : labels,
            "counts"     : counts,
            "bases"      : bases,
            "pcts"       : pcts,
            "colors"     : colors,
            "n_total"    : n_total,
            "n_missing"  : n_missing,
            "n_unique"   : int(series.nunique()),
            "pct_missing": round(n_missing / n_total * 100, 1),
            "top_value"  : labels[0] if labels else "—",
            "top_pct"    : pcts[0]   if pcts   else 0,
        }

    if not all_data:
        print("No categorical columns found.")
        return

    first_col = list(all_data.keys())[0]
    data_json = json.dumps(all_data)
    cols_json = json.dumps(list(all_data.keys()))

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Categorical Waterfall Explorer</title>
<script src="https://cdn.jsdelivr.net/npm/plotly.js-dist@2.27.0/plotly.min.js"></script>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
  background: #f4f6fb; padding: 28px 32px; color: #333;
}}
h2 {{ font-size: 22px; font-weight: 600; color: #1a1a2e; margin-bottom: 4px; }}
.subtitle {{ font-size: 13px; color: #999; margin-bottom: 22px; }}
#chart {{
  background: white; border-radius: 12px;
  box-shadow: 0 1px 6px rgba(0,0,0,0.07); padding: 8px;
}}
.search-section {{
  margin-top: 22px; background: white; border-radius: 12px;
  padding: 20px 22px 22px; box-shadow: 0 1px 6px rgba(0,0,0,0.07);
}}
.search-label {{ font-size: 13px; font-weight: 600; color: #555; margin-bottom: 10px; display: block; }}
.search-wrapper {{ position: relative; width: 100%; }}
#col-search {{
  width: 100%; padding: 11px 16px; font-size: 14px;
  border: 1.5px solid #e0e0e0; border-radius: 8px;
  outline: none; background: #fafafa;
  transition: border-color 0.2s, box-shadow 0.2s; color: #333;
}}
#col-search:focus {{
  border-color: #4c8eda; background: white;
  box-shadow: 0 0 0 3px rgba(76,142,218,0.12);
}}
#col-search::placeholder {{ color: #bbb; }}
#dropdown-list {{
  position: absolute; top: calc(100% + 6px); left: 0; right: 0;
  background: white; border: 1.5px solid #e0e0e0; border-radius: 10px;
  max-height: 280px; overflow-y: auto; z-index: 9999; display: none;
  box-shadow: 0 8px 24px rgba(0,0,0,0.13);
}}
.dropdown-item {{
  padding: 10px 16px; font-size: 13px; cursor: pointer;
  border-bottom: 1px solid #f5f5f5; transition: background 0.12s;
  white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
}}
.dropdown-item:last-child {{ border-bottom: none; }}
.dropdown-item:hover    {{ background: #eef4fd; color: #4c8eda; }}
.dropdown-item.selected {{ background: #e8f0fb; color: #4c8eda; font-weight: 600; }}
.dropdown-item.active   {{ background: #ddeeff; }}
.dropdown-item mark {{ background: #fff176; border-radius: 2px; padding: 0 1px; font-weight: 700; color: inherit; }}
.no-results {{ padding: 14px 16px; font-size: 13px; color: #bbb; font-style: italic; }}
.stats-bar {{ display: flex; flex-wrap: wrap; gap: 10px; margin-top: 18px; }}
.stat-card {{ flex: 1; min-width: 90px; border-radius: 8px; padding: 10px 14px; text-align: center; }}
.stat-card.normal  {{ background: #f4f7fb; }}
.stat-card.warning {{ background: #fff4f4; }}
.stat-label {{ font-size: 10px; color: #aaa; text-transform: uppercase; letter-spacing: 0.6px; }}
.stat-value {{ font-size: 16px; font-weight: 700; color: #1a1a2e; margin-top: 3px; }}
.stat-card.warning .stat-value {{ color: #e05c5c; }}
.legend-row {{
  display: flex; gap: 18px; flex-wrap: wrap; margin-top: 14px;
  font-size: 12px; color: #666; align-items: center;
}}
.legend-pill {{ display: flex; align-items: center; gap: 5px; }}
.pill-dot {{ width: 10px; height: 10px; border-radius: 2px; flex-shrink: 0; }}
.badge {{
  display: inline-block; background: #eef4fd; color: #4c8eda;
  border-radius: 20px; padding: 2px 10px; font-size: 12px;
  font-weight: 600; margin-left: 8px; vertical-align: middle;
}}
</style>
</head>
<body>

<h2>Categorical Waterfall Explorer <span class="badge">{len(all_data)} columns</span></h2>
<p class="subtitle">True waterfall chart · each bar starts where the previous ends · NaN shown in red</p>

<div id="chart"></div>

<div class="search-section">
  <label class="search-label">Search or select a categorical feature:</label>
  <div class="search-wrapper">
    <input id="col-search" type="text"
      placeholder="Type to search  e.g.  SI_TORTYP,  SI_DEKOR ..."
      autocomplete="off"/>
    <div id="dropdown-list"></div>
  </div>

  <div class="legend-row">
    <div class="legend-pill">
      <div class="pill-dot" style="background:linear-gradient(to right,#1a3a5c,#89b4fa)"></div>
      Category values (dark → light = highest → lowest count)
    </div>
    <div class="legend-pill"><div class="pill-dot" style="background:#aaaaaa"></div>Other (grouped)</div>
    <div class="legend-pill"><div class="pill-dot" style="background:#e05c5c"></div>NaN / Missing</div>
    <div class="legend-pill"><div class="pill-dot" style="background:#2d6a4f"></div>Total (all rows)</div>
  </div>

  <div class="stats-bar" id="stats-bar"></div>
</div>

<script>
const ALL_DATA = {data_json};
const ALL_COLS = {cols_json};
let activeCol  = "{first_col}";

// blue gradient from dark to light based on rank
function blueGradient(i, total) {{
  const t = total <= 1 ? 0 : i / (total - 1);
  const r = Math.round(26  + t * (137 - 26));
  const g = Math.round(58  + t * (180 - 58));
  const b = Math.round(92  + t * (250 - 92));
  return `rgb(${{r}},${{g}},${{b}})`;
}}

function buildChart(col) {{
  const d = ALL_DATA[col];
  if (!d) return;
  activeCol = col;

  const n      = d.labels.length;
  const colors = d.labels.map((lbl, i) => {{
    if (lbl === "NaN / Missing")   return "#e05c5c";
    if (lbl === "Other (grouped)") return "#aaaaaa";
    if (lbl === "▌ Total")        return "#2d6a4f";
    const specialCount = d.labels.filter(l =>
      l === "NaN / Missing" || l === "Other (grouped)" || l === "▌ Total"
    ).length;
    const normalTotal = n - specialCount;
    return blueGradient(i, normalTotal);
  }});

  const traces = [];

  // ── Invisible base bars (for waterfall offset) ────────────────
  traces.push({{
    type: "bar",
    x: d.labels,
    y: d.bases,
    marker: {{ color: "rgba(0,0,0,0)" }},
    hoverinfo: "none",
    showlegend: false,
    name: "base"
  }});

  // ── Visible value bars ────────────────────────────────────────
  traces.push({{
    type: "bar",
    x: d.labels,
    y: d.counts,
    marker: {{
      color: colors,
      line: {{ color: "white", width: 1 }}
    }},
    text: d.pcts.map((p, i) =>
      `${{d.counts[i].toLocaleString()}}<br>${{p}}%`
    ),
    textposition: "outside",
    textfont: {{ size: 11, color: "#444" }},
    cliponaxis: false,
    hovertemplate: d.labels.map((lbl, i) => {{
      if (lbl === "▌ Total") {{
        return `<b>Total</b><br>Count : ${{d.counts[i].toLocaleString()}}<br>Share : 100%<br><i>Sum of all rows</i><extra></extra>`;
      }}
      return `<b>${{lbl}}</b><br>` +
        `Count : ${{d.counts[i].toLocaleString()}}<br>` +
        `Share : ${{d.pcts[i]}}%<br>` +
        `Running total: ${{(d.bases[i] + d.counts[i]).toLocaleString()}}` +
        (lbl === "NaN / Missing" ? "<br><i>Missing values</i>" : "") +
        "<extra></extra>";
    }}),
    name: col,
    showlegend: false
  }});

  // ── Connector lines between bars (skip connector TO Total) ─────
  const connectorShapes = [];
  for (let i = 0; i < d.labels.length - 2; i++) {{
    const yTop = d.bases[i] + d.counts[i];
    connectorShapes.push({{
      type: "line",
      x0: i + 0.45,
      x1: i + 0.55,
      y0: yTop,
      y1: yTop,
      line: {{ color: "#cccccc", width: 1, dash: "dot" }}
    }});
  }}

  // ── Vertical separator before Total bar ──────────────────────
  connectorShapes.push({{
    type: "line",
    x0: d.labels.length - 1.5,
    x1: d.labels.length - 1.5,
    y0: 0,
    y1: d.n_total,
    line: {{ color: "#cccccc", width: 1, dash: "dash" }}
  }});

  const totalY = d.n_total;

  Plotly.react("chart", traces, {{
    title: {{
      text: `Waterfall distribution — <b>${{col}}</b><br>` +
            `<sup>n=${{d.n_total.toLocaleString()}}  ·  ` +
            `${{d.n_unique}} unique values  ·  ` +
            `${{d.pct_missing}}% missing</sup>`,
      font: {{ size: 15 }}
    }},
    barmode: "stack",
    xaxis: {{
      title: "Category value",
      tickangle: -40,
      tickfont: {{ size: 11 }},
      showgrid: false,
      categoryorder: "array",
      categoryarray: d.labels
    }},
    yaxis: {{
      title: "Cumulative count",
      gridcolor: "#eeeeee",
      showgrid: true
    }},
    shapes: connectorShapes,
    annotations: [{{
      x: "▌ Total",
      y: totalY,
      text: `n = ${{totalY.toLocaleString()}}`,
      showarrow: false,
      yanchor: "bottom",
      font: {{ size: 11, color: "#2d6a4f" }}
    }}],
    plot_bgcolor: "white",
    paper_bgcolor: "white",
    height: 520,
    margin: {{ t: 80, r: 30, l: 80, b: 140 }},
    showlegend: false
  }}, {{responsive: true}});

  // stats
  const hasMissing = d.n_missing > 0;
  const stats = [
    ["total rows",    d.n_total.toLocaleString(), "normal"],
    ["unique values", d.n_unique,                 "normal"],
    ["top value",     d.top_value,                "normal"],
    ["top %",         d.top_pct + "%",            "normal"],
    ["missing",       d.n_missing.toLocaleString(), hasMissing ? "warning" : "normal"],
    ["missing %",     d.pct_missing + "%",          hasMissing ? "warning" : "normal"],
  ];
  document.getElementById("stats-bar").innerHTML = stats.map(([lbl, val, cls]) =>
    `<div class="stat-card ${{cls}}">
       <div class="stat-label">${{lbl}}</div>
       <div class="stat-value">${{val}}</div>
     </div>`
  ).join("");

  document.getElementById("col-search").value = col;
}}

// ── Searchable dropdown ─────────────────────────────────────────
const searchInput  = document.getElementById("col-search");
const dropdownList = document.getElementById("dropdown-list");
let filteredCols = [...ALL_COLS];
let highlighted  = -1;

function hl(text, q) {{
  if (!q) return text;
  return text.replace(
    new RegExp(`(${{q.replace(/[.*+?^${{}}()|[\\]\\\\]/g,"\\\\$&")}})`, "gi"),
    "<mark>$1</mark>"
  );
}}

function renderDropdown(q) {{
  q = q.toLowerCase().trim();
  filteredCols = ALL_COLS.filter(c => c.toLowerCase().includes(q));
  highlighted  = -1;
  dropdownList.innerHTML = filteredCols.length === 0
    ? `<div class="no-results">No columns match "${{q}}"</div>`
    : filteredCols.map((col, i) => {{
        const d   = ALL_DATA[col] || {{}};
        const nan  = d.n_missing > 0
          ? ` <span style="color:#e05c5c;font-size:11px">· ${{d.pct_missing}}% NaN</span>` : "";
        const uniq = d.n_unique !== undefined
          ? ` <span style="color:#aaa;font-size:11px">(${{d.n_unique}} values)</span>` : "";
        return `<div class="dropdown-item ${{col===activeCol?"selected":""}}"
                     data-col="${{col}}" data-idx="${{i}}">
                  ${{hl(col,q)}}${{uniq}}${{nan}}
                </div>`;
      }}).join("");

  dropdownList.querySelectorAll(".dropdown-item").forEach(el => {{
    el.addEventListener("mousedown", e => {{ e.preventDefault(); selectCol(el.dataset.col); }});
  }});
  dropdownList.style.display = "block";
}}

function selectCol(col) {{ buildChart(col); dropdownList.style.display = "none"; }}

function setHL(idx) {{
  const items = dropdownList.querySelectorAll(".dropdown-item");
  items.forEach(el => el.classList.remove("active"));
  if (idx >= 0 && idx < items.length) {{
    items[idx].classList.add("active");
    items[idx].scrollIntoView({{block:"nearest"}});
    highlighted = idx;
  }}
}}

searchInput.addEventListener("focus", () => renderDropdown(searchInput.value));
searchInput.addEventListener("input", () => renderDropdown(searchInput.value));
searchInput.addEventListener("keydown", e => {{
  const items = dropdownList.querySelectorAll(".dropdown-item");
  if      (e.key==="ArrowDown") {{ e.preventDefault(); setHL(Math.min(highlighted+1, items.length-1)); }}
  else if (e.key==="ArrowUp")   {{ e.preventDefault(); setHL(Math.max(highlighted-1, 0)); }}
  else if (e.key==="Enter") {{
    e.preventDefault();
    if (highlighted>=0 && filteredCols[highlighted]) selectCol(filteredCols[highlighted]);
    else if (filteredCols.length>0) selectCol(filteredCols[0]);
  }}
  else if (e.key==="Escape") {{ dropdownList.style.display="none"; searchInput.blur(); }}
}});
document.addEventListener("click", e => {{
  if (!e.target.closest(".search-wrapper")) dropdownList.style.display = "none";
}});

buildChart("{first_col}");
</script>
</body>
</html>"""

    with open(output_file, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"✅ Saved → {output_file}  |  open in your browser")
    return output_file


def categorical_stacked_bar_horizontal(data_table, output_file="categorical_stacked_h.html", top_n=10):
    """
    Horizontal 100% stacked bar chart — ALL categorical columns on Y axis.
    Scrollable vertically. No colorbar — legend only.
    Controls: row height, % labels, NaN only, sort by NaN %.
    """

    cat_cols = data_table.select_dtypes(
        include=["object", "category", "str"]
    ).columns.tolist()

    for col in data_table.select_dtypes(include="number").columns:
        if 2 <= data_table[col].nunique() <= 15:
            cat_cols.append(col)

    cat_cols = list(dict.fromkeys(cat_cols))
    if not cat_cols:
        print("No categorical columns found.")
        return

    print(f"Building horizontal stacked bar for {len(cat_cols)} columns...")

    n_total = len(data_table)
    all_values_ordered = []
    col_data = {}

    for col in cat_cols:
        series = data_table[col].copy().astype(str)
        series[data_table[col].isna()] = "NaN / None"
        vc = series.value_counts()

        if len(vc) > top_n:
            top   = vc.iloc[:top_n]
            other = vc.iloc[top_n:].sum()
            vc    = pd.concat([top, pd.Series({"Other (grouped)": int(other)})])

        col_data[col] = {
            "labels": list(vc.index),
            "counts": [int(v) for v in vc.values],
            "pcts"  : [round(v / n_total * 100, 1) for v in vc.values],
        }

        for lbl in vc.index:
            if lbl not in all_values_ordered:
                all_values_ordered.append(lbl)

    for s in ["Other (grouped)", "NaN / None"]:
        if s in all_values_ordered:
            all_values_ordered.remove(s)
            all_values_ordered.append(s)

    base_colors = [
        "#4c8eda","#f0a500","#6ab187","#9b59b6","#e67e22",
        "#1abc9c","#e74c3c","#2980b9","#8e44ad","#27ae60",
        "#d35400","#16a085","#c0392b","#2c3e50","#f39c12",
        "#3498db","#e91e63","#009688","#ff5722","#795548",
        "#607d8b","#ff9800","#4caf50","#03a9f4","#9c27b0",
        "#673ab7","#cddc39","#00bcd4","#8bc34a","#ff4081",
    ]

    color_map = {}
    idx = 0
    for val in all_values_ordered:
        if val == "NaN / None":
            color_map[val] = "#e05c5c"
        elif val == "Other (grouped)":
            color_map[val] = "#b0bec5"
        else:
            color_map[val] = base_colors[idx % len(base_colors)]
            idx += 1

    traces_data = []
    for val in all_values_ordered:
        y_cols, x_pcts, x_cnts, hover = [], [], [], []
        for col in cat_cols:
            d = col_data[col]
            if val in d["labels"]:
                i   = d["labels"].index(val)
                cnt = d["counts"][i]
                pct = d["pcts"][i]
            else:
                cnt, pct = 0, 0.0
            y_cols.append(col)
            x_pcts.append(pct)
            x_cnts.append(cnt)
            hover.append(
                f"<b>{col}</b><br>"
                f"Value : <b>{val}</b><br>"
                f"Count : {cnt:,}<br>"
                f"Share : {pct}%"
                + ("<br><i>Missing / null</i>" if val == "NaN / None" else "")
                + "<extra></extra>"
            )
        traces_data.append({
            "name"  : val,
            "y"     : y_cols,
            "x"     : x_pcts,
            "counts": x_cnts,
            "color" : color_map[val],
            "hover" : hover,
        })

    traces_json = json.dumps(traces_data)
    n_cols      = len(cat_cols)
    bar_px      = 36
    visible_bars = min(n_cols, 20)
    inner_h      = n_cols * bar_px
    outer_h      = visible_bars * bar_px + 120

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Categorical Stacked Bar — Horizontal</title>
<script src="https://cdn.jsdelivr.net/npm/plotly.js-dist@2.27.0/plotly.min.js"></script>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
  background: #f4f6fb; padding: 24px 28px; color: #333;
}}
h2 {{ font-size: 21px; font-weight: 600; color: #1a1a2e; margin-bottom: 4px; }}
.subtitle {{ font-size: 13px; color: #999; margin-bottom: 18px; }}
.badge {{
  display: inline-block; background: #eef4fd; color: #4c8eda;
  border-radius: 20px; padding: 2px 10px; font-size: 12px;
  font-weight: 600; margin-left: 8px; vertical-align: middle;
}}
.controls {{
  display: flex; gap: 10px; flex-wrap: wrap;
  align-items: center; margin-bottom: 14px;
}}
.ctrl-group {{
  display: flex; align-items: center; gap: 8px;
  background: white; border-radius: 8px; padding: 8px 14px;
  box-shadow: 0 1px 4px rgba(0,0,0,0.07); font-size: 13px;
}}
.ctrl-group label {{ color: #555; font-weight: 500; white-space: nowrap; }}
input[type=range]    {{ width: 110px; cursor: pointer; }}
input[type=checkbox] {{ cursor: pointer; width: 15px; height: 15px; }}
.chart-outer {{
  background: white; border-radius: 12px;
  box-shadow: 0 1px 6px rgba(0,0,0,0.07); overflow: hidden;
}}
.scroll-hint {{
  text-align: center; font-size: 12px; color: #aaa; padding: 6px 0 2px;
  display: {'"block"' if n_cols > 20 else '"none"'};
}}
.chart-scroll {{
  overflow-y: scroll; overflow-x: hidden;
  height: {outer_h}px;
}}
#chart {{ height: {inner_h + 100}px; min-height: {inner_h + 100}px; }}
.footer {{ margin-top: 12px; font-size: 12px; color: #aaa; text-align: right; }}
</style>
</head>
<body>

<h2>
  Categorical Distribution — 100% Horizontal
  <span class="badge">{n_cols} columns</span>
  <span class="badge" style="background:#fff4f4;color:#e05c5c">incl. NaN</span>
</h2>
<p class="subtitle">
  Y axis = feature columns · scroll down to see all ·
  X axis = % of rows · Red = NaN/None · Gray = Other
</p>

<div class="controls">
  <div class="ctrl-group">
    <label>Row height</label>
    <input type="range" id="bar-height" min="20" max="70" step="2" value="36"/>
    <span id="bar-height-val" style="min-width:40px">36 px</span>
  </div>
  <div class="ctrl-group">
    <label>% labels</label>
    <input type="checkbox" id="label-toggle" checked/>
  </div>
  <div class="ctrl-group">
    <label>NaN only</label>
    <input type="checkbox" id="nan-toggle"/>
  </div>
  <div class="ctrl-group">
    <label>Sort by NaN %</label>
    <input type="checkbox" id="sort-nan"/>
  </div>
</div>

<div class="chart-outer">
  <div class="scroll-hint">↕ Scroll to see all {n_cols} columns</div>
  <div class="chart-scroll" id="scroll-wrap">
    <div id="chart"></div>
  </div>
</div>

<div class="footer">
  Hover for details · Click legend to show/hide · Scroll to navigate columns
</div>

<script>
const TRACES_DATA = {traces_json};
const N_TOTAL     = {n_total};
const N_COLS      = {n_cols};

let showLabels = true;
let nanOnly    = false;
let sortNaN    = false;
let barPx      = 36;

function getNaNOrder() {{
  const t = TRACES_DATA.find(t => t.name === "NaN / None");
  if (!t) return null;
  const pairs = t.y.map((col, i) => ({{ col, pct: t.x[i] }}));
  pairs.sort((a, b) => b.pct - a.pct);
  return pairs.map(p => p.col);
}}

function getColOrder() {{
  const base = [...TRACES_DATA[0].y].reverse();
  if (!sortNaN) return base;
  const sorted = getNaNOrder();
  return sorted ? sorted.slice().reverse() : base;
}}

function buildTraces(colOrder) {{
  return TRACES_DATA
    .filter(t => !nanOnly || t.name === "NaN / None")
    .map(t => {{
      const yNew = [], xNew = [], hNew = [];
      colOrder.forEach(col => {{
        const i = t.y.indexOf(col);
        if (i >= 0) {{
          yNew.push(t.y[i]); xNew.push(t.x[i]); hNew.push(t.hover[i]);
        }} else {{
          yNew.push(col); xNew.push(0);
          hNew.push(`<b>${{col}}</b><br>Value: <b>${{t.name}}</b><br>Count: 0<br>Share: 0%<extra></extra>`);
        }}
      }});
      return {{
        type         : "bar",
        orientation  : "h",
        name         : t.name,
        y            : yNew,
        x            : xNew,
        marker       : {{ color: t.color, line: {{ color: "white", width: 0.5 }} }},
        text         : showLabels ? xNew.map(v => v >= 5 ? v.toFixed(1) + "%" : "") : [],
        textposition : "inside",
        textfont     : {{ size: 10, color: "white" }},
        insidetextanchor: "middle",
        hovertemplate: hNew,
        showlegend   : true,
      }};
    }});
}}

function render() {{
  const colOrder = getColOrder();
  const totalH   = colOrder.length * barPx + 100;

  document.getElementById("chart").style.height    = totalH + "px";
  document.getElementById("chart").style.minHeight = totalH + "px";
  document.getElementById("scroll-wrap").style.height =
    Math.min(totalH, 20 * barPx + 120) + "px";

  Plotly.react("chart", buildTraces(colOrder), {{
    barmode      : "stack",
    height       : totalH,
    plot_bgcolor : "white",
    paper_bgcolor: "white",
    legend: {{
      orientation : "h",
      x: 0, y: -0.05,
      font: {{ size: 11 }},
      bgcolor: "rgba(255,255,255,0.95)",
      bordercolor: "#eee", borderwidth: 1,
    }},
    xaxis: {{
      title     : {{ text: "% of total rows  (n={n_total:,})", font: {{ size: 12 }} }},
      range     : [0, 100],
      ticksuffix: "%",
      gridcolor : "#eeeeee",
      showgrid  : true,
      side      : "top",
      fixedrange: true,
    }},
    yaxis: {{
      tickfont  : {{ size: 11 }},
      showgrid  : false,
      automargin: true,
      fixedrange: true,
    }},
    margin: {{ t: 50, r: 20, l: 220, b: 80 }},
    hoverlabel: {{ bgcolor: "white", font: {{ size: 12 }}, bordercolor: "#ddd" }}
  }}, {{
    responsive: true,
    displayModeBar: true,
    modeBarButtonsToRemove: ["lasso2d","select2d","zoom2d","pan2d","zoomIn2d","zoomOut2d","autoScale2d"],
    toImageButtonOptions: {{
      format: "png", filename: "categorical_stacked_horizontal",
      scale: 2, height: totalH, width: 1400
    }}
  }});
}}

document.getElementById("bar-height").addEventListener("input", function() {{
  barPx = +this.value;
  document.getElementById("bar-height-val").textContent = barPx + " px";
  render();
}});
document.getElementById("label-toggle").addEventListener("change", function() {{
  showLabels = this.checked; render();
}});
document.getElementById("nan-toggle").addEventListener("change", function() {{
  nanOnly = this.checked; render();
}});
document.getElementById("sort-nan").addEventListener("change", function() {{
  sortNaN = this.checked; render();
}});

render();
</script>
</body>
</html>"""

    with open(output_file, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"✅ Saved → {output_file}  |  open in your browser")
    return output_file


def drop_empty_columns(data_table, inplace=False, preview_rows=3):
    """
    Remove fully empty columns and print detailed report.

    Parameters
    ----------
    data_table : pandas.DataFrame
        Input DataFrame
    inplace : bool, optional
        If True, modifies original DataFrame
    preview_rows : int, optional
        Number of preview rows to display

    Returns
    -------
    pandas.DataFrame
        Cleaned DataFrame
    """

    data_table_clean = data_table if inplace else data_table.copy()

    # Identify empty columns
    empty_cols = data_table_clean.columns[data_table_clean.isna().all()].tolist()

    before = data_table_clean.shape[1]

    # Drop them
    data_table_clean.drop(columns=empty_cols, inplace=True)

    after = data_table_clean.shape[1]

    # Print report
    print(f"🧹 Dropped {len(empty_cols)} fully-empty columns → {after} columns remain")

    if empty_cols:
        print("\n🗑️ Deleted columns:")
        for col in empty_cols:
            print(f"   - {col}")
    else:
        print("\n✅ No fully-empty columns found.")

    return data_table_clean


def drop_constant_columns(df, inplace=False, preview_rows=3):
    """
    Remove columns that contain only a single unique value.

    Parameters
    ----------
    df : pandas.DataFrame
        Input DataFrame
    inplace : bool, optional
        If True, modifies original DataFrame
    preview_rows : int, optional
        Number of preview rows

    Returns
    -------
    pandas.DataFrame
        Cleaned DataFrame without constant columns
    """

    df_clean = df if inplace else df.copy()

    # Find constant columns (1 unique value excluding NaN)
    constant_cols = [col for col in df_clean.columns if df_clean[col].nunique(dropna=True) <= 1]

    before = df_clean.shape[1]

    # Drop them
    df_clean.drop(columns=constant_cols, inplace=True)

    after = df_clean.shape[1]

    # Report
    print(f"🧹 Dropped {len(constant_cols)} constant columns → {after} columns remain")

    if constant_cols:
        print("\n🗑️ Constant columns removed:")
        for col in constant_cols:
            unique_val = df[col].dropna().unique()
            print(f"   - {col} (value: {unique_val[0] if len(unique_val) > 0 else 'NaN'})")
    else:
        print("\n✅ No constant columns found.")

    # Preview
    if preview_rows > 0:
        print("\n🔍 Preview:")
        print(df_clean.head(preview_rows))

    return df_clean


def tsne_visualization(
    df,
    target_col=None,
    drop_sparse_threshold=0.2,
    perplexity=30,
    random_state=42,
    figsize=(13, 8),
    colormap="RdYlGn_r",
    point_size=80,
    alpha=0.8,
    annotate=False,
    save_path=None,
):
    """
    Professional t-SNE visualization with continuous colorscale bar.

    Steps
    -----
    1.  Select numeric columns only
    2.  Drop target column from features (target is only used for coloring)
    3.  Drop sparse columns (below threshold)
    4.  Impute missing values with median
    5.  Standardize features (critical for t-SNE)
    6.  Run t-SNE (2D embedding)
    7.  Plot with colorscale bar if target is continuous,
        or discrete legend if target is categorical

    Parameters
    ----------
    df                   : pd.DataFrame  — input dataset
    target_col           : str or None   — column used for point coloring only
                           (excluded from t-SNE features entirely)
    drop_sparse_threshold: float         — min ratio of non-null values per column
                           e.g. 0.2 = drop columns with >80% missing
    perplexity           : int           — t-SNE perplexity (5–50, default 30)
    random_state         : int           — reproducibility seed
    figsize              : tuple         — (width, height) in inches
    colormap             : str           — matplotlib colormap name
                           continuous: "RdYlGn_r", "plasma", "viridis", "coolwarm"
                           categorical: auto-assigned from tab10/tab20
    point_size           : int           — scatter point size
    alpha                : float         — point transparency (0–1)
    annotate             : bool          — annotate points with df.index labels
    save_path            : str or None   — save figure to this path if provided

    Returns
    -------
    tsne_df : pd.DataFrame
        Columns: ["Dim1", "Dim2"] + target_col (if provided)
        Index matches original df index
    """

    print("─" * 55)
    print("  t-SNE Visualization Pipeline")
    print("─" * 55)

    # ── 1. Numeric columns only ───────────────────────────────────
    df_num = df.select_dtypes(include="number").copy()
    print(f"  Numeric columns       : {df_num.shape[1]}")

    # ── 2. Remove target from features ───────────────────────────
    if target_col and target_col in df_num.columns:
        df_num = df_num.drop(columns=[target_col])
        print(f"  Target removed        : {target_col} (used only for color)")

    # ── 3. Extract target values ──────────────────────────────────
    if target_col and target_col in df.columns:
        target_series = pd.to_numeric(df[target_col], errors="coerce")
        target_vals   = target_series.values
        is_continuous = target_series.nunique() > 10
    else:
        target_vals   = None
        is_continuous = False

    # ── 4. Drop sparse columns ────────────────────────────────────
    thresh  = int(drop_sparse_threshold * len(df_num))
    before  = df_num.shape[1]
    df_num  = df_num.dropna(axis=1, thresh=thresh)
    dropped = before - df_num.shape[1]
    print(f"  Sparse cols dropped   : {dropped}  →  {df_num.shape[1]} remain")

    # ── 5. Impute missing ─────────────────────────────────────────
    imputer  = SimpleImputer(strategy="median")
    X_imp    = imputer.fit_transform(df_num)

    # ── 6. Scale ──────────────────────────────────────────────────
    scaler   = StandardScaler()
    X_scaled = scaler.fit_transform(X_imp)

    print(f"  Final matrix          : {X_scaled.shape[0]} rows × {X_scaled.shape[1]} features")
    print(f"  Perplexity            : {perplexity}")
    print(f"  Colormap              : {colormap}")
    print(f"  Target column         : {target_col or 'None (no coloring)'}")

    # ── 7. Run t-SNE ──────────────────────────────────────────────
    print("\n  Running t-SNE  ⏳  (may take a moment...)")
    tsne = TSNE(
        n_components=2,
        perplexity=min(perplexity, X_scaled.shape[0] - 1),
        random_state=random_state,
        init="pca",
        learning_rate="auto",
        max_iter=1000,
    )
    X_tsne = tsne.fit_transform(X_scaled)
    print("  t-SNE done ✅")

    # ── 8. Build result DataFrame ─────────────────────────────────
    tsne_df = pd.DataFrame(
        X_tsne,
        columns=["Dim1", "Dim2"],
        index=df.index
    )
    if target_vals is not None:
        tsne_df[target_col] = target_vals

    # ── 9. Plot ───────────────────────────────────────────────────
    if target_vals is not None and is_continuous:
        # ── Continuous target → colorscale bar ───────────────────
        fig, (ax, cax) = plt.subplots(
            1, 2,
            figsize=figsize,
            gridspec_kw={"width_ratios": [20, 1]}
        )

        # normalize color values
        valid_mask = ~np.isnan(target_vals.astype(float))
        vals_clean = target_vals[valid_mask].astype(float)
        norm       = mcolors.Normalize(vmin=vals_clean.min(), vmax=vals_clean.max())
        cmap_obj   = cm.get_cmap(colormap)
        colors_arr = cmap_obj(norm(target_vals.astype(float)))

        # scatter — colored points
        sc = ax.scatter(
            X_tsne[:, 0], X_tsne[:, 1],
            c=target_vals.astype(float),
            cmap=colormap,
            norm=norm,
            s=point_size,
            alpha=alpha,
            edgecolors="white",
            linewidths=0.3,
        )

        # colorscale bar in dedicated axis
        cb = ColorbarBase(
            cax,
            cmap=cmap_obj,
            norm=norm,
            orientation="vertical"
        )
        cb.set_label(target_col, fontsize=12, labelpad=10)
        cb.ax.tick_params(labelsize=10)

        ax.set_title(
            f"t-SNE  —  colored by  {target_col}",
            fontsize=14, fontweight="bold", pad=14
        )

    elif target_vals is not None and not is_continuous:
        # ── Categorical target → discrete legend ──────────────────
        fig, ax = plt.subplots(figsize=figsize)

        unique_vals = pd.Series(target_vals).dropna().unique()
        palette     = cm.get_cmap("tab20", len(unique_vals))
        color_dict  = {v: palette(i) for i, v in enumerate(unique_vals)}

        for val in unique_vals:
            mask = pd.Series(target_vals) == val
            ax.scatter(
                X_tsne[mask, 0], X_tsne[mask, 1],
                label=str(val),
                color=color_dict[val],
                s=point_size,
                alpha=alpha,
                edgecolors="white",
                linewidths=0.3,
            )

        ax.legend(
            title=target_col,
            fontsize=9,
            title_fontsize=10,
            bbox_to_anchor=(1.01, 1),
            loc="upper left",
            framealpha=0.9
        )
        ax.set_title(
            f"t-SNE  —  colored by  {target_col}",
            fontsize=14, fontweight="bold", pad=14
        )

    else:
        # ── No target → single color ──────────────────────────────
        fig, ax = plt.subplots(figsize=figsize)
        ax.scatter(
            X_tsne[:, 0], X_tsne[:, 1],
            color="#4c8eda",
            s=point_size,
            alpha=alpha,
            edgecolors="white",
            linewidths=0.3,
        )
        ax.set_title("t-SNE Visualization", fontsize=14, fontweight="bold", pad=14)

    # ── Shared axis styling ───────────────────────────────────────
    ax.set_xlabel("t-SNE Dimension 1", fontsize=12)
    ax.set_ylabel("t-SNE Dimension 2", fontsize=12)
    ax.grid(True, alpha=0.25, linewidth=0.5)
    ax.set_facecolor("#fafafa")

    # ── Annotate index labels ─────────────────────────────────────
    if annotate:
        for i, label in enumerate(df.index):
            ax.annotate(
                str(label),
                (X_tsne[i, 0], X_tsne[i, 1]),
                fontsize=7,
                alpha=0.7,
                xytext=(4, 4),
                textcoords="offset points"
            )

    plt.tight_layout()

    if save_path:
        plt.savefig(save_path, dpi=150, bbox_inches="tight")
        print(f"\n  Saved → {save_path}")

    plt.show()

    print("─" * 55)
    print(f"  Output shape: {tsne_df.shape}")
    print("─" * 55)

    return tsne_df


def plot_interactive_clustered_corr(df, title="Clustered Correlation Matrix"):
    """
    Creates an interactive, clustered heatmap using Plotly.
    
    Parameters
    ----------
    df : pandas.DataFrame
        The correlation matrix or raw data.
    title : str
        The title for the plot.
        
    Returns
    -------
    plotly.graph_objects.Figure
        The interactive Plotly figure.
    """
    # 1. Safety Cleaning (Matches your Seaborn logic)
    # Drop rows/cols that are all NaN, fill others with 0
    corr = df.dropna(axis=0, how="all").dropna(axis=1, how="all").fillna(0)
    
    # 2. Hierarchical Clustering
    # We need to reorder the rows/cols so similar variables are adjacent
    if corr.shape[0] > 1:
        linkage = hierarchy.linkage(corr, method='ward')
        order = hierarchy.leaves_list(linkage)
        
        # Reorder the correlation matrix based on the clustering
        corr_clustered = corr.iloc[order, order]
    else:
        corr_clustered = corr

    # 3. Create the Plotly Heatmap
    fig = px.imshow(
        corr_clustered,
        text_auto=False,
        aspect="auto",
        color_continuous_scale="RdBu_r", # Red-Blue reversed (Red=Pos, Blue=Neg)
        zmin=-1,
        zmax=1,
        labels=dict(color="Correlation"),
        title=title
    )

    # 4. Styling for the Browser
    fig.update_layout(
        width=900,
        height=900,
        xaxis_nticks=len(corr_clustered.columns),
        yaxis_nticks=len(corr_clustered.index),
        xaxis_tickangle=-90,
        title_x=0.5, # Center title
        hovermode="closest"
    )
    
    # Adjust font size for labels to keep it readable
    fig.update_xaxes(tickfont=dict(size=8))
    fig.update_yaxes(tickfont=dict(size=8))

    return fig


def clean_numeric_features(
    df: pd.DataFrame,
    missing_threshold: float = 0.5,
    variance_threshold: float = 0.01,
    verbose: bool = True,
) -> pd.DataFrame:
    """
    Clean numeric features by removing columns with excessive missing values
    or near-zero variance. Non-numeric columns are preserved unchanged.

    Pipeline
    --------
    1. Isolate numeric columns
    2. Drop columns exceeding the missing value threshold
    3. Drop columns with variance below the variance threshold
    4. Reattach non-numeric columns
    5. Print a detailed report (if verbose=True)

    Parameters
    ----------
    df                 : pd.DataFrame
        Input DataFrame — may contain mixed column types.
    missing_threshold  : float, default 0.5
        Minimum fraction of non-null values required to keep a column.
        e.g. 0.5 → drop columns with more than 50% missing values.
    variance_threshold : float, default 0.01
        Minimum variance required to keep a column.
        e.g. 0.01 → drop columns where values barely change across rows.
        Uses sklearn VarianceThreshold internally.
    verbose            : bool, default True
        Print a full report of dropped columns and reasons.

    Returns
    -------
    df_clean : pd.DataFrame
        Cleaned DataFrame with low-quality numeric columns removed
        and all non-numeric columns preserved.

    Examples
    --------
    >>> df_clean = clean_numeric_features(df)
    >>> df_clean = clean_numeric_features(df, missing_threshold=0.7, variance_threshold=0.05)
    >>> df_clean = clean_numeric_features(df, verbose=False)
    """

    n_rows = len(df)

    # ── Step 0: separate numeric from non-numeric ─────────────────
    df_num     = df.select_dtypes(include="number").copy()
    df_non_num = df.select_dtypes(exclude="number").copy()

    initial_num_cols = df_num.columns.tolist()
    n_initial        = len(initial_num_cols)

    dropped_missing  = []    # (col, pct_missing)
    dropped_variance = []    # (col, variance)
    kept_cols        = []

    # ── Step 1: missing value filter ──────────────────────────────
    min_non_null = int(missing_threshold * n_rows)
    surviving_missing = df_num.dropna(
        axis=1, thresh=min_non_null
    ).columns.tolist()

    for col in initial_num_cols:
        if col not in surviving_missing:
            pct = df_num[col].isna().mean() * 100
            dropped_missing.append((col, pct))

    df_after_missing = df_num[surviving_missing]

    # ── Step 2: variance filter (sklearn VarianceThreshold) ───────
    df_filled  = df_after_missing.fillna(df_after_missing.median())
    selector   = VarianceThreshold(threshold=variance_threshold)
    selector.fit(df_filled)

    variances       = dict(zip(df_filled.columns, selector.variances_))
    support         = selector.get_support()
    cols_after_miss = df_after_missing.columns.tolist()

    for col, keep in zip(cols_after_miss, support):
        if keep:
            kept_cols.append(col)
        else:
            dropped_variance.append((col, variances[col]))

    df_clean_num = df_after_missing[kept_cols]

    # ── Step 3: reattach non-numeric columns ──────────────────────
    df_clean = pd.concat([df_clean_num, df_non_num], axis=1)

    # ── Step 4: verbose report ────────────────────────────────────
    if verbose:
        n_dropped_miss = len(dropped_missing)
        n_dropped_var  = len(dropped_variance)
        n_kept         = len(kept_cols)
        n_non_num      = len(df_non_num.columns)

        sep  = "─" * 65
        sep2 = "─" * 65

        print(f"\n{sep}")
        print(f"  Feature Cleaning Report")
        print(f"{sep}")
        print(f"  Input  : {n_rows:,} rows  ×  {len(df.columns)} columns total")
        print(f"           └─ {n_initial} numeric  +  {n_non_num} non-numeric")
        print(f"  Missing threshold  : >{(1 - missing_threshold)*100:.0f}% missing  → dropped")
        print(f"  Variance threshold : < {variance_threshold}  → dropped")
        print(f"{sep}")

        # ── Missing drops ─────────────────────────────────────────
        if dropped_missing:
            print(f"\n  ✗ DROPPED — Excessive missing values  ({n_dropped_miss} columns)")
            print(f"  {'Column':<45} {'Missing %':>10}")
            print(f"  {'─'*45} {'─'*10}")
            for col, pct in sorted(dropped_missing, key=lambda x: -x[1]):
                bar   = "█" * int(pct / 5)
                print(f"  {col:<45} {pct:>8.1f}%  {bar}")
        else:
            print(f"\n  ✓ No columns dropped for missing values")

        # ── Variance drops ────────────────────────────────────────
        if dropped_variance:
            print(f"\n  ✗ DROPPED — Low variance  ({n_dropped_var} columns)")
            print(f"  {'Column':<45} {'Variance':>12}")
            print(f"  {'─'*45} {'─'*12}")
            for col, var in sorted(dropped_variance, key=lambda x: x[1]):
                print(f"  {col:<45} {var:>12.6f}")
        else:
            print(f"\n  ✓ No columns dropped for low variance")

        # ── Summary ───────────────────────────────────────────────
        print(f"\n{sep2}")
        print(f"  {'SUMMARY':}")
        print(f"  {'─'*40}")
        print(f"  Numeric in          : {n_initial}")
        print(f"  Dropped (missing)   : {n_dropped_miss}")
        print(f"  Dropped (variance)  : {n_dropped_var}")
        print(f"  Numeric kept        : {n_kept}")
        print(f"  Non-numeric kept    : {n_non_num}  (untouched)")
        print(f"  {'─'*40}")
        print(f"  Output shape        : {df_clean.shape[0]:,} rows × {df_clean.shape[1]} columns")
        retention = n_kept / n_initial * 100 if n_initial > 0 else 0
        print(f"  Retention rate      : {retention:.1f}%")
        print(f"{sep2}\n")

    return df_clean


def plot_clustered_correlation(corr_matrix, title="Clustered Spearman Correlation Matrix", save_path=None):
    """
    Generate and save a cleaned hierarchical clustering heatmap.

    This function handles NaN values, performs hierarchical clustering to group 
    similar features, and applies aesthetic formatting to the resulting heatmap.

    Parameters
    ----------
    corr_matrix : pandas.DataFrame
        A square correlation matrix (e.g., from df.corr()).
    title : str, default "Clustered Spearman Correlation Matrix"
        The title to display on the heatmap.
    save_path : str, optional
        The file path to save the resulting image (e.g., 'plot.png'). 
        If None, the plot is not saved.

    Returns
    -------
    seaborn.matrix.ClusterGrid
        The Seaborn ClusterGrid object containing the plot.
    """
    # 1. Safety Cleaning: Drop entirely empty rows/cols and fill remaining NaNs
    # Clustermap math fails if NaNs are present
    corr_clean = corr_matrix.dropna(axis=0, how="all").dropna(axis=1, how="all")
    corr_clean = corr_clean.fillna(0)

    # 2. Generate Clustermap
    g = sns.clustermap(
        corr_clean,
        cmap="coolwarm",
        center=0,
        vmin=-1, 
        vmax=1,
        figsize=(18, 18),
        xticklabels=True,
        yticklabels=True,
        annot=False,
        linewidths=0,
        dendrogram_ratio=0.1,    # Smaller trees, more heatmap space
        cbar_pos=(0.02, 0.8, 0.03, 0.18)  # Position: (left, bottom, width, height)
    )

    # 3. Aesthetics & Title
    # Note: Titles must be set on g.ax_heatmap, not the figure itself
    g.ax_heatmap.set_title(title, fontsize=15, pad=20)

    # Refine label sizes for high-density matrices
    g.ax_heatmap.set_xticklabels(
        g.ax_heatmap.get_xticklabels(), 
        fontsize=6, 
        rotation=90
    )
    g.ax_heatmap.set_yticklabels(
        g.ax_heatmap.get_yticklabels(), 
        fontsize=6, 
        rotation=0
    )

    # 4. Save/Show logic
    if save_path:
        g.figure.savefig(save_path, dpi=150, bbox_inches="tight")
        print(f"Plot saved to {save_path}")

    plt.show()
    return g


def compute_correlation(
    df: pd.DataFrame,
    method: str = "spearman",
    threshold: float = 0.6,
    top_n: int = 20,
    verbose: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Compute a cleaned correlation matrix and extract strongly correlated
    feature pairs above a given threshold.

    Supports three correlation methods, each with different assumptions:

    pearson
        Measures linear relationship between two continuous variables.
        Assumes normality and is sensitive to outliers.

        Formula:
            r = sum((x_i - mean_x) * (y_i - mean_y))
                / sqrt(sum((x_i - mean_x)^2) * sum((y_i - mean_y)^2))

        Best for: normally distributed data, linear relationships.

    spearman
        Measures monotonic relationship using ranked values.
        Robust to outliers and does not assume normality.

        Formula:
            rho = 1 - (6 * sum(d_i^2)) / (n * (n^2 - 1))

        where d_i is the rank difference per observation i.
        Best for: skewed data, ordinal variables, non-linear monotonic trends.

    kendall
        Measures ordinal association using concordant/discordant pairs.
        More robust than Spearman for small samples or many ties.

        Formula:
            tau = (C - D) / sqrt((C + D + T_x) * (C + D + T_y))

        where C = concordant pairs, D = discordant pairs,
        T_x / T_y = ties in x / y respectively.
        Best for: small datasets, heavy ties, ordinal data.

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame. Non-numeric columns are coerced to numeric
        automatically. Columns that fail coercion become NaN.
    method : str, optional
        Correlation method to use. One of:
            "pearson"   — linear correlation (default in pandas)
            "spearman"  — rank-based, robust to outliers
            "kendall"   — concordance-based, robust to ties
        Default is "spearman".
    threshold : float, optional
        Minimum absolute correlation |r| to flag a pair as strong.
        Must be in range (0, 1). Default is 0.6.

        Interpretation:
            |r| >= 0.9  → very strong
            |r| >= 0.7  → strong
            |r| >= 0.5  → moderate
            |r| <  0.5  → weak

    top_n : int, optional
        Number of top pairs to show in the printed report. Default 20.
    verbose : bool, optional
        Print detailed report if True. Default is True.

    Returns
    -------
    corr_clean : pd.DataFrame
        Symmetric n_features × n_features correlation matrix.
        All-NaN rows/cols removed. Remaining NaNs filled with 0.

    pairs : pd.DataFrame
        Long-form table of feature pairs with |r| >= threshold.
        Columns: Feature A, Feature B, Correlation, Abs_r, Direction.
        Sorted by Abs_r descending. Self-correlations excluded.

    Raises
    ------
    ValueError
        If method is not one of "pearson", "spearman", "kendall".

    Examples
    --------
    >>> corr_clean, pairs = compute_correlation(df_clean)
    >>> corr_clean, pairs = compute_correlation(df_clean, method="pearson")
    >>> corr_clean, pairs = compute_correlation(df_clean, method="kendall", threshold=0.5)
    >>> corr_clean, pairs = compute_correlation(df_clean, verbose=False)
    """

    # ── Validate method ───────────────────────────────────────────
    valid_methods = {"pearson", "spearman", "kendall"}
    if method not in valid_methods:
        raise ValueError(
            f"Invalid method '{method}'. "
            f"Choose from: {sorted(valid_methods)}"
        )

    # ── Step 1: coerce all to numeric ────────────────────────────
    df_numeric = df.apply(pd.to_numeric, errors="coerce")

    # ── Step 2: compute correlation matrix ───────────────────────
    corr = df_numeric.corr(method=method)

    # ── Step 3: clean the matrix ──────────────────────────────────
    corr_clean = (
        corr
        .dropna(axis=0, how="all")
        .dropna(axis=1, how="all")
        .fillna(0.0)
    )

    # ── Step 4: extract strong pairs (upper triangle only) ────────
    upper_mask = np.triu(np.ones(corr_clean.shape, dtype=bool), k=1)

    strong_mask = (
        (corr_clean.abs() >= threshold) &
        (corr_clean != 1.0) &
        upper_mask
    )

    pairs = (
        corr_clean
        .where(strong_mask)
        .stack()
        .reset_index()
        .rename(columns={
            "level_0": "Feature A",
            "level_1": "Feature B",
            0        : "Correlation"
        })
        .assign(
            Abs_r     = lambda x: x["Correlation"].abs().round(4),
            Direction = lambda x: x["Correlation"].apply(
                lambda v: "positive" if v > 0 else "negative"
            ),
            Correlation = lambda x: x["Correlation"].round(4),
        )
        .sort_values("Abs_r", ascending=False)
        .reset_index(drop=True)
    )

    # ── Step 5: verbose report ────────────────────────────────────
    if verbose:
        method_desc = {
            "pearson" : "Linear (Pearson r)",
            "spearman": "Rank-based (Spearman ρ)",
            "kendall" : "Concordance-based (Kendall τ)",
        }
        sep = "─" * 62
        print(f"\n{sep}")
        print(f"  Correlation Report")
        print(f"{sep}")
        print(f"  Method            : {method_desc[method]}")
        print(f"  Input features    : {df_numeric.shape[1]}")
        print(f"  Matrix shape      : {corr_clean.shape}")
        print(f"  Threshold |r| ≥   : {threshold}")
        print(f"  Strong pairs found: {len(pairs)}")
        print(f"    ↳ Positive      : {(pairs['Direction'] == 'positive').sum()}")
        print(f"    ↳ Negative      : {(pairs['Direction'] == 'negative').sum()}")
        print(f"{sep}")

        if len(pairs) > 0:
            print(f"\n  Top {min(top_n, len(pairs))} strongest pairs:\n")
            print(f"  {'Feature A':<30} {'Feature B':<30} {'r':>8}  Direction")
            print(f"  {'─'*30} {'─'*30} {'─'*8}  {'─'*9}")
            for _, row in pairs.head(top_n).iterrows():
                arrow = "↑" if row["Direction"] == "positive" else "↓"
                print(
                    f"  {row['Feature A']:<30} "
                    f"{row['Feature B']:<30} "
                    f"{row['Correlation']:>8.4f}  "
                    f"{arrow} {row['Direction']}"
                )
        else:
            print(f"  No pairs found above threshold {threshold}")

        print(f"\n{sep}\n")

    return corr_clean, pairs


def plot_column_correlations(
    corr_clean: pd.DataFrame,
    column: str,
    *,
    top_n: int = 25,
    min_abs_corr: float = 0.0,
    figsize: Tuple[float, float] = (10, 8),
    pos_color: str = "#3b82f6",
    neg_color: str = "#ef4444",
    fade_by_magnitude: bool = True,
    title: Optional[str] = None,
    show: bool = True,
    save_path: Optional[str] = None,
    dpi: int = 150,
) -> Tuple[pd.Series, dict]:
    """
    Visualise the correlations of one target column against all
    other features from a precomputed correlation matrix.

    Improvements over the previous version
    --------------------------------------
    * Optional ``min_abs_corr`` threshold to drop weak correlations.
    * Gradient colour intensity proportional to |r| (toggle via
      ``fade_by_magnitude``).
    * Returns both the full sorted Series and a summary dict.
    * ``show=False`` lets you build a figure without rendering it
      (useful for batch reporting or unit tests).
    * Auto-creates parent directories for ``save_path``.
    * Strict input validation with helpful error messages.

    Parameters
    ----------
    corr_clean : pd.DataFrame
        Symmetric correlation matrix with matching index and columns.
    column : str
        Target column name to compare against.
    top_n : int, default 25
        Number of features to display, ranked by absolute correlation.
    min_abs_corr : float, default 0.0
        Drop features whose ``|r|`` is below this value before ranking.
        Must be in [0, 1].
    figsize : (float, float), default (10, 8)
        Figure size in inches.
    pos_color, neg_color : str
        Base colours for positive / negative bars.
    fade_by_magnitude : bool, default True
        If True, bar opacity scales with ``|r|`` (stronger = solid).
    title : str, optional
        Custom plot title. Defaults to a sensible auto-generated one.
    show : bool, default True
        Call ``plt.show()`` after drawing.
    save_path : str, optional
        Path to save the figure (.png / .pdf / .svg). Parent dir
        is created automatically.
    dpi : int, default 150
        Resolution when saving.

    Returns
    -------
    col_corr : pd.Series
        All non-NaN correlations vs ``column``, sorted by value desc.
    summary : dict
        Keys: ``column``, ``n_features``, ``n_positive``, ``n_negative``,
        ``strongest_positive`` (name, r), ``strongest_negative`` (name, r),
        ``mean_abs_corr``, ``displayed`` (count after filtering).

    Raises
    ------
    TypeError
        If ``corr_clean`` is not a DataFrame.
    ValueError
        If ``corr_clean`` is not square / symmetric, ``column`` is
        missing, or ``min_abs_corr`` is out of [0, 1].

    Examples
    --------
    >>> col_corr, info = plot_column_correlations(corr_clean, "KgCO2EQ")
    >>> col_corr, info = plot_column_correlations(
    ...     corr_clean, "KgCO2EQ",
    ...     top_n=15, min_abs_corr=0.2,
    ...     save_path="plots/corr_KgCO2EQ.png",
    ... )
    """

    # ── Validate inputs ───────────────────────────────────────────
    if not isinstance(corr_clean, pd.DataFrame):
        raise TypeError("`corr_clean` must be a pandas DataFrame.")
    if corr_clean.shape[0] != corr_clean.shape[1]:
        raise ValueError(
            f"Correlation matrix must be square, got shape {corr_clean.shape}."
        )
    if column not in corr_clean.columns:
        preview = ", ".join(map(str, corr_clean.columns[:8]))
        raise ValueError(
            f"Column '{column}' not found in correlation matrix.\n"
            f"First columns available: [{preview}, ...]"
        )
    if not 0.0 <= min_abs_corr <= 1.0:
        raise ValueError(
            f"`min_abs_corr` must be in [0, 1], got {min_abs_corr}."
        )

    # ── Extract and rank correlations ─────────────────────────────
    col_corr = (
        corr_clean[column]
        .drop(labels=column, errors="ignore")
        .dropna()
        .sort_values(ascending=False)
    )

    if col_corr.empty:
        raise ValueError(f"No non-NaN correlations available for '{column}'.")

    # apply threshold + take top_n by |r|
    filtered = col_corr[col_corr.abs() >= min_abs_corr]
    top_corr = (
        filtered
        .reindex(filtered.abs().sort_values(ascending=False).index)
        .head(top_n)
        .sort_values(ascending=True)  # ascending for barh readability
    )

    # ── Summary dict (always built) ───────────────────────────────
    summary = {
        "column":             column,
        "n_features":         int(len(col_corr)),
        "n_positive":         int((col_corr > 0).sum()),
        "n_negative":         int((col_corr < 0).sum()),
        "strongest_positive": (col_corr.idxmax(), float(col_corr.max())),
        "strongest_negative": (col_corr.idxmin(), float(col_corr.min())),
        "mean_abs_corr":      float(col_corr.abs().mean()),
        "displayed":          int(len(top_corr)),
    }

    # ── Plot ──────────────────────────────────────────────────────
    if not top_corr.empty:
        fig, ax = plt.subplots(figsize=figsize)

        # color (+ optional alpha proportional to |r|)
        max_abs = max(top_corr.abs().max(), 1e-9)
        colors, alphas = [], []
        for v in top_corr:
            colors.append(pos_color if v >= 0 else neg_color)
            alphas.append(0.35 + 0.65 * (abs(v) / max_abs) if fade_by_magnitude else 1.0)

        bars = ax.barh(
            y=top_corr.index,
            width=top_corr.values,
            color=colors,
            edgecolor="white",
            linewidth=0.6,
        )
        for bar, a in zip(bars, alphas):
            bar.set_alpha(a)

        # zero reference line
        ax.axvline(0, color="#555555", linewidth=1.0, linestyle="--")

        # value labels
        for bar, val in zip(bars, top_corr):
            x = val + 0.01 if val >= 0 else val - 0.01
            ax.text(
                x, bar.get_y() + bar.get_height() / 2,
                f"{val:+.3f}",
                va="center",
                ha="left" if val >= 0 else "right",
                fontsize=8, color="#333",
            )

        # legend
        ax.legend(
            handles=[
                Patch(facecolor=pos_color, label="Positive correlation"),
                Patch(facecolor=neg_color, label="Negative correlation"),
            ],
            fontsize=9, loc="lower right", frameon=False,
        )

        # title + cosmetics
        ax.set_title(
            title or f"Top {len(top_corr)} correlations with '{column}'"
                    + (f"  (|r| ≥ {min_abs_corr:g})" if min_abs_corr > 0 else ""),
            fontsize=13, fontweight="bold", pad=12,
        )
        ax.set_xlabel("Correlation coefficient  r", fontsize=11)
        ax.set_ylabel("Feature", fontsize=11)
        # tighter, data-aware limits (still padded for labels)
        bound = max(0.2, top_corr.abs().max()) + 0.1
        ax.set_xlim(-bound, bound)
        ax.grid(True, axis="x", alpha=0.25, linewidth=0.5)
        ax.set_facecolor("#fafafa")
        for spine in ("top", "right"):
            ax.spines[spine].set_visible(False)

        plt.tight_layout()

        if save_path:
            out = Path(save_path)
            out.parent.mkdir(parents=True, exist_ok=True)
            plt.savefig(out, dpi=dpi, bbox_inches="tight")
            print(f"  💾 Saved → {out.resolve()}")

        if show:
            plt.show()
        else:
            plt.close(fig)
    else:
        print(f"⚠️  No features pass |r| ≥ {min_abs_corr} for '{column}'.")

    # ── Print summary ─────────────────────────────────────────────
    sep = "─" * 55
    print(sep)
    print(f"  Column              : {summary['column']}")
    print(f"  Features in matrix  : {summary['n_features']}")
    print(f"  Displayed (top_n)   : {summary['displayed']}")
    print(f"  Positive / Negative : {summary['n_positive']} / {summary['n_negative']}")
    print(f"  Mean |r|            : {summary['mean_abs_corr']:.3f}")
    print(f"  Strongest +         : {summary['strongest_positive'][0]}  "
          f"({summary['strongest_positive'][1]:+.3f})")
    print(f"  Strongest −         : {summary['strongest_negative'][0]}  "
          f"({summary['strongest_negative'][1]:+.3f})")
    print(sep)

    return col_corr, summary


def plot_distribution(
    df: pd.DataFrame,
    columns: Union[str, list],
    bins: int = 30,
    figsize: tuple = (14, 10),
    save: bool = False,
    save_dir: str = ".",
) -> dict:
    """
    Plot a 4-panel distribution analysis for one or more numeric columns.

    For each column the following panels are produced:

    Panel 1 — Histogram + KDE
        Combines a frequency histogram with a kernel density estimate
        curve. Vertical lines mark the mean and median, making it easy
        to spot skew (large mean–median gap = right-skewed distribution).

    Panel 2 — ECDF (Empirical Cumulative Distribution Function)
        Shows what fraction of observations fall below each value.
        The horizontal dashed line at 0.5 marks the median.
        An S-curve shape indicates a near-normal distribution;
        a steep early rise indicates most values are clustered low.

    Panel 3 — Boxplot + Strip Plot
        The box spans Q1–Q3 (interquartile range). The line inside
        is the median. Whiskers extend to 1.5×IQR. Points beyond
        the whiskers are outliers (shown in red). The strip plot
        overlays all individual observations so the actual density
        of points is visible — not hidden behind the box summary.

    Panel 4 — Log-scale Histogram
        Applies log1p transformation: log(1 + x). This compresses
        right-skewed distributions into a more symmetric shape,
        making it easier to see the underlying structure when most
        values are small but a few are very large.

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame containing the columns to analyse.
    columns : str or list of str
        One column name or a list of column names to plot.
        Each column produces its own independent 4-panel figure.
        Non-numeric values are coerced to NaN and dropped silently.
    bins : int, optional
        Number of bins for histogram panels. Default is 30.
    figsize : tuple, optional
        Width and height of each figure in inches. Default is (14, 10).
    save : bool, optional
        If True, saves each figure as a PNG file. Default is False.
        Filename pattern: ``{save_dir}/{column}_distribution.png``
    save_dir : str, optional
        Directory to save figures when ``save=True``. Default is ".".

    Returns
    -------
    results : dict
        Dictionary mapping each column name to its summary statistics.
        Each value is a dict with keys:
            n, mean, median, std, skewness, kurtosis, min, max, q25, q75

    Raises
    ------
    ValueError
        If none of the requested columns exist in ``df``.

    Examples
    --------
    >>> # single column
    >>> stats = plot_distribution(df, columns="KgCO2EQ")

    >>> # multiple columns — one figure per column
    >>> stats = plot_distribution(df, columns=["KgCO2EQ", "SDCOM_WIDTH", "SI_BW"])

    >>> # all numeric columns at once
    >>> num_cols = df.select_dtypes(include="number").columns.tolist()
    >>> stats = plot_distribution(df, columns=num_cols, save=True)

    >>> # access stats after plotting
    >>> stats["KgCO2EQ"]["skewness"]
    """

    # ── Normalise input to list ───────────────────────────────────
    if isinstance(columns, str):
        columns = [columns]

    # ── Validate columns ──────────────────────────────────────────
    missing = [c for c in columns if c not in df.columns]
    valid   = [c for c in columns if c in df.columns]

    if missing:
        print(f"⚠️  Columns not found and skipped: {missing}")
    if not valid:
        raise ValueError("None of the requested columns exist in df.")

    results = {}

    for col in valid:

        # ── Prepare data ──────────────────────────────────────────
        data = pd.to_numeric(df[col], errors="coerce").dropna()

        if len(data) < 2:
            print(f"⚠️  '{col}' has fewer than 2 valid values — skipped.")
            continue

        skewness = data.skew()
        kurt     = data.kurt()
        mean_val = data.mean()
        med_val  = data.median()
        std_val  = data.std()
        q25      = data.quantile(0.25)
        q75      = data.quantile(0.75)

        # ── Figure ────────────────────────────────────────────────
        fig, axes = plt.subplots(2, 2, figsize=figsize)
        fig.suptitle(
            f"Distribution Analysis — {col}",
            fontsize=15, fontweight="bold", y=1.01
        )

        # ── Panel 1: Histogram + KDE ──────────────────────────────
        ax = axes[0, 0]
        sns.histplot(
            data, bins=bins, kde=True,
            color="#4c8eda",
            line_kws={"linewidth": 2},
            ax=ax
        )
        ax.axvline(
            mean_val, color="#e05c5c",
            linestyle="--", linewidth=1.5,
            label=f"Mean   {mean_val:.2f}"
        )
        ax.axvline(
            med_val, color="#f0a500",
            linestyle="--", linewidth=1.5,
            label=f"Median {med_val:.2f}"
        )
        ax.set_title("Histogram + KDE")
        ax.set_xlabel(col)
        ax.legend(fontsize=9)
        ax.set_facecolor("#fafafa")

        # ── Panel 2: ECDF ─────────────────────────────────────────
        ax = axes[0, 1]
        sns.ecdfplot(data, color="#4c8eda", linewidth=2, ax=ax)
        ax.axhline(
            0.5, color="#f0a500",
            linestyle="--", linewidth=1,
            alpha=0.7, label="50th percentile"
        )
        ax.set_title("Cumulative Distribution (ECDF)")
        ax.set_xlabel(col)
        ax.set_ylabel("Proportion")
        ax.legend(fontsize=9)
        ax.set_facecolor("#fafafa")

        # ── Panel 3: Boxplot + Strip ──────────────────────────────
        ax = axes[1, 0]
        sns.boxplot(
            x=data, color="#4c8eda", width=0.4,
            flierprops=dict(
                marker="o",
                markerfacecolor="#e05c5c",
                markersize=5, alpha=0.6
            ),
            ax=ax
        )
        sns.stripplot(
            x=data, color="#2c3e50",
            alpha=0.35, size=4,
            jitter=True, ax=ax
        )
        ax.set_title("Boxplot + Individual Points")
        ax.set_xlabel(col)
        ax.set_facecolor("#fafafa")

        # ── Panel 4: Log-scale Histogram ──────────────────────────
        ax = axes[1, 1]
        data_log = np.log1p(data[data > 0])

        if len(data_log) > 1:
            sns.histplot(
                data_log, bins=bins, kde=True,
                color="#6ab187",
                line_kws={"linewidth": 2},
                ax=ax
            )
            ax.set_title(
                f"Log Histogram  {'(skewed — log helps)' if abs(skewness) > 1 else '(low skew)'}"
            )
            ax.set_xlabel(f"log(1 + {col})")
        else:
            ax.text(
                0.5, 0.5, "Not applicable\n(no positive values)",
                ha="center", va="center",
                transform=ax.transAxes, color="#aaa"
            )
            ax.set_title("Log Histogram")

        ax.set_facecolor("#fafafa")

        plt.tight_layout()

        if save:
            import os
            os.makedirs(save_dir, exist_ok=True)
            path = os.path.join(save_dir, f"{col}_distribution.png")
            plt.savefig(path, dpi=150, bbox_inches="tight")
            print(f"  ✅ Saved → {path}")

        plt.show()

        # ── Summary stats ─────────────────────────────────────────
        stats = {
            "n"       : len(data),
            "mean"    : round(mean_val, 4),
            "median"  : round(med_val,  4),
            "std"     : round(std_val,  4),
            "skewness": round(skewness, 4),
            "kurtosis": round(kurt,     4),
            "min"     : round(data.min(), 4),
            "max"     : round(data.max(), 4),
            "q25"     : round(q25, 4),
            "q75"     : round(q75, 4),
        }

        results[col] = stats

        sep = "─" * 40
        print(f"\n{sep}")
        print(f"  Column     : {col}")
        print(f"  n          : {stats['n']:,}")
        print(f"  mean       : {stats['mean']:.4f}")
        print(f"  median     : {stats['median']:.4f}")
        print(f"  std        : {stats['std']:.4f}")
        print(f"  skewness   : {stats['skewness']:.4f}"
              + ("  ← right-skewed" if skewness > 1
                 else "  ← left-skewed" if skewness < -1
                 else "  ← approx. symmetric"))
        print(f"  kurtosis   : {stats['kurtosis']:.4f}"
              + ("  ← heavy tails" if kurt > 3
                 else "  ← light tails" if kurt < -1
                 else ""))
        print(f"  Q25 / Q75  : {stats['q25']:.4f} / {stats['q75']:.4f}")
        print(f"  min / max  : {stats['min']:.4f} / {stats['max']:.4f}")
        print(f"{sep}\n")

    return results


def plot_bivariate(
    df: pd.DataFrame,
    feature: str,
    target: str,
    bins: int = 50,
    kde_levels: int = 5,
    figsize: tuple = (10, 10),
    cmap: str = "mako",
    show_marginals: bool = True,
    save: bool = False,
    save_dir: str = ".",
) -> dict:
    """
    Plot a joint density scatter between one feature and one target column.
    Optionally includes marginal histograms on the X and Y axes.

    Parameters
    ----------
    df             : pd.DataFrame — input data
    feature        : str          — X axis column (predictor)
    target         : str          — Y axis column (outcome)
    bins           : int          — histogram + 2D density bins (default 50)
    kde_levels     : int          — KDE contour levels (default 5)
    figsize        : tuple        — figure size in inches (default 10×10)
    cmap           : str          — colormap for 2D histogram (default "mako")
    show_marginals : bool         — show top + right marginal histograms
                                   True  → 3-panel layout with marginals
                                   False → single joint plot only
    save           : bool         — save figure to disk (default False)
    save_dir       : str          — directory for saved file (default ".")

    Returns
    -------
    result : dict
        n, pearson_r, spearman_r, p_value, slope, intercept, r_squared

    Examples
    --------
    >>> # with marginal histograms (default)
    >>> result = plot_bivariate(df, feature="SDCOM_WIDTH", target="KgCO2EQ")

    >>> # without marginal histograms
    >>> result = plot_bivariate(df, feature="SDCOM_WIDTH", target="KgCO2EQ",
    ...                         show_marginals=False)
    """

    # ── Validate ──────────────────────────────────────────────────
    for col in [feature, target]:
        if col not in df.columns:
            raise ValueError(f"Column '{col}' not found in DataFrame.")

    # ── Prepare data ──────────────────────────────────────────────
    plot_df          = df[[feature, target]].copy()
    plot_df[feature] = pd.to_numeric(plot_df[feature], errors="coerce")
    plot_df[target]  = pd.to_numeric(plot_df[target],  errors="coerce")
    plot_df          = plot_df.dropna()

    if len(plot_df) < 5:
        raise ValueError(f"Only {len(plot_df)} valid rows — not enough to plot.")

    x = plot_df[feature].values
    y = plot_df[target].values

    # ── Statistics ────────────────────────────────────────────────
    pearson_r,  p_value        = stats.pearsonr(x, y)
    spearman_r, _              = stats.spearmanr(x, y)
    slope, intercept, _, _, _  = stats.linregress(x, y)
    r_squared                  = pearson_r ** 2

    print(f"{'─'*55}")
    print(f"  {feature}  ↔  {target}")
    print(f"{'─'*55}")
    print(f"  Valid rows     : {len(x):,}")
    print(f"  {feature:<22}: [{x.min():.2f} – {x.max():.2f}]")
    print(f"  {target:<22}: [{y.min():.2f} – {y.max():.2f}]")
    print(f"  Pearson r      : {pearson_r:.4f}  (p = {p_value:.4e})")
    print(f"  Spearman ρ     : {spearman_r:.4f}")
    print(f"  R²             : {r_squared:.4f}")
    print(f"  Slope          : {slope:.4f}")
    print(f"  Intercept      : {intercept:.4f}")
    print(f"  Marginal hists : {'ON' if show_marginals else 'OFF'}")
    print(f"{'─'*55}")

    # ── Build figure layout ───────────────────────────────────────
    sns.set_theme(style="dark")

    if show_marginals:
        fig = plt.figure(figsize=figsize)
        gs  = fig.add_gridspec(
            2, 2,
            width_ratios  = [4, 1],
            height_ratios = [1, 4],
            hspace        = 0.04,
            wspace        = 0.04,
        )
        ax_main  = fig.add_subplot(gs[1, 0])
        ax_top   = fig.add_subplot(gs[0, 0], sharex=ax_main)
        ax_right = fig.add_subplot(gs[1, 1], sharey=ax_main)

        # ── Top marginal — X histogram ────────────────────────────
        ax_top.hist(
            x, bins=bins // 2,
            color="#4c8eda", alpha=0.4,
            density=True, edgecolor="none"
        )
        kde_x   = stats.gaussian_kde(x)
        x_range = np.linspace(x.min(), x.max(), 300)
        ax_top.plot(x_range, kde_x(x_range), color="#1a1a2e", linewidth=1.8)
        ax_top.set_ylabel("Density", fontsize=9, color="#888")
        ax_top.tick_params(labelbottom=False, labelsize=8)
        ax_top.set_facecolor("#eef2f7")
        for spine in ax_top.spines.values():
            spine.set_visible(False)
        ax_top.grid(False)

        # ── Right marginal — Y histogram (rotated) ────────────────
        ax_right.hist(
            y, bins=bins // 2,
            color="#4c8eda", alpha=0.4,
            density=True, edgecolor="none",
            orientation="horizontal"
        )
        kde_y   = stats.gaussian_kde(y)
        y_range = np.linspace(y.min(), y.max(), 300)
        ax_right.plot(kde_y(y_range), y_range, color="#1a1a2e", linewidth=1.8)
        ax_right.set_xlabel("Density", fontsize=9, color="#888")
        ax_right.tick_params(labelleft=False, labelsize=8)
        ax_right.set_facecolor("#eef2f7")
        for spine in ax_right.spines.values():
            spine.set_visible(False)
        ax_right.grid(False)

    else:
        # simple single-panel figure
        fig, ax_main = plt.subplots(figsize=figsize)

    # ── Main panel — joint density plot ───────────────────────────
    sns.scatterplot(
        x=x, y=y,
        s=15, color=".15", alpha=0.6,
        ax=ax_main
    )
    sns.histplot(
        x=x, y=y,
        bins=bins, pthresh=0.1,
        cmap=cmap, ax=ax_main
    )
    sns.kdeplot(
        x=x, y=y,
        levels=kde_levels,
        color="w", linewidths=1,
        ax=ax_main
    )

    # regression line
    x_line = np.linspace(x.min(), x.max(), 300)
    ax_main.plot(
        x_line, slope * x_line + intercept,
        color="#f0a500", linewidth=1.8,
        linestyle="--",
        label=f"OLS  slope = {slope:.3f}"
    )

    # annotation box
    sig = "✓ significant" if p_value < 0.05 else "✗ not significant"
    annotation = (
        f"Pearson r  = {pearson_r:+.3f}\n"
        f"Spearman ρ = {spearman_r:+.3f}\n"
        f"R²         = {r_squared:.3f}\n"
        f"p-value    = {p_value:.2e}  {sig}"
    )
    ax_main.annotate(
        annotation,
        xy=(0.04, 0.97), xycoords="axes fraction",
        va="top", ha="left",
        fontsize=9, color="white", family="monospace",
        bbox=dict(boxstyle="round,pad=0.4", fc="#222", alpha=0.75)
    )

    ax_main.set_xlabel(feature, fontsize=12)
    ax_main.set_ylabel(target,  fontsize=12)
    ax_main.legend(fontsize=10)

    fig.suptitle(
        f"{feature}  ↔  {target}\n"
        f"Density + Contour + OLS"
        + (" + Marginal Distributions" if show_marginals else ""),
        fontsize=13, fontweight="bold", y=1.01
    )

    plt.tight_layout()

    if save:
        import os
        os.makedirs(save_dir, exist_ok=True)
        path = os.path.join(save_dir, f"{feature}_vs_{target}.png")
        plt.savefig(path, dpi=150, bbox_inches="tight")
        print(f"\n  ✅ Saved → {path}")

    plt.show()

    return {
        "n"         : int(len(x)),
        "pearson_r" : round(pearson_r,  4),
        "spearman_r": round(spearman_r, 4),
        "p_value"   : round(p_value,    6),
        "slope"     : round(slope,      4),
        "intercept" : round(intercept,  4),
        "r_squared" : round(r_squared,  4),
    }


def plot_categorical_distribution_all_or_multiple_features(
    df: pd.DataFrame,
    columns: list = None,
    top_n: int = 10,
    min_pct_label: float = 4.0,
    bar_gap: float = 0.3,
    height: int = 600,
    color_palette: list = None,
    include_nan: bool = True,
    output_file: str = "categorical_100pct.html",
    show: bool = True,
) -> go.Figure:
    """
    Create a 100% stacked vertical bar chart showing the value distribution
    of all categorical columns in a DataFrame.

    Each bar represents one categorical column. Each colored segment within
    a bar represents one unique value, scaled to percentage of total rows.
    Values beyond ``top_n`` are grouped into an "Other" segment.

    The chart uses a consistent global color mapping — the same value
    receives the same color across all columns, making it easy to track
    one value across multiple features (e.g. "DE" in language columns).

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame. Columns of dtype ``object`` and ``category``
        are used by default. Low-cardinality numeric columns can be
        included by passing them explicitly via ``columns``.
    columns : list of str or None, optional
        Explicit list of column names to include. If None, all object
        and category columns are used automatically. Default is None.
    top_n : int, optional
        Maximum number of distinct values to show per column.
        Values ranked beyond ``top_n`` by frequency are merged into
        a single "Other" segment. Default is 10.
    min_pct_label : float, optional
        Minimum percentage required to show a text label inside a bar
        segment. Segments below this threshold show no label to avoid
        overcrowding. Default is 4.0 (%).
    bar_gap : float, optional
        Gap between bars as a fraction of bar width. Range 0–1.
        Default is 0.3.
    height : int, optional
        Figure height in pixels. Default is 600.
    color_palette : list of str or None, optional
        List of hex color strings for value coloring. If None, uses
        ``plotly.express.colors.qualitative.Safe`` by default.
    include_nan : bool, optional
        If True, NaN values are counted and shown as a distinct
        "NaN / None" segment colored red. Default is True.
    output_file : str, optional
        File path for the HTML output. Default is
        "categorical_100pct.html".
    show : bool, optional
        If True, opens the figure in the browser. Default is True.

    Returns
    -------
    fig : plotly.graph_objects.Figure
        The fully constructed Plotly figure. Can be further customised
        or embedded in dashboards.

    Raises
    ------
    ValueError
        If no categorical columns are found or none of the requested
        ``columns`` exist in ``df``.

    Examples
    --------
    >>> fig = plot_categorical_distribution_all_or_multiple_features(df)

    >>> # specific columns only
    >>> fig = plot_categorical_distribution_all_or_multiple_features(
    ...     df,
    ...     columns = ["SI_TORTYP", "SI_DEKOR", "SI_PROFSERIE"]
    ... )

    >>> # more values per bar + include NaN
    >>> fig = plot_categorical_distribution_all_or_multiple_features(df, top_n=15, include_nan=True)

    >>> # silent — no browser, just save
    >>> fig = plot_categorical_distribution_all_or_multiple_features(df, show=False,
    ...                                     output_file="my_cats.html")
    """

    # ── Resolve columns ───────────────────────────────────────────
    if columns is not None:
        missing  = [c for c in columns if c not in df.columns]
        if missing:
            print(f"⚠️  Columns not found and skipped: {missing}")
        cat_cols = [c for c in columns if c in df.columns]
    else:
        cat_cols = df.select_dtypes(
            include=["object", "category", "str"]
        ).columns.tolist()

    if not cat_cols:
        raise ValueError("No categorical columns found to plot.")

    n_total = len(df)

    # ── Compute value counts per column ───────────────────────────
    col_counts = {}
    for col in cat_cols:
        series = df[col].copy()

        # handle NaN
        if include_nan:
            nan_count = int(series.isna().sum())
            series    = series.astype(str)
            series[df[col].isna()] = "NaN / None"
        else:
            series    = series.dropna().astype(str)
            nan_count = 0

        vc = series.value_counts(normalize=False)
        vc_pct = (vc / n_total * 100).round(1)

        # group tail into Other
        if len(vc_pct) > top_n:
            top   = vc_pct.iloc[:top_n]
            other = vc_pct.iloc[top_n:].sum().round(1)
            vc_pct = pd.concat([top, pd.Series({"Other": other})])

        col_counts[col] = vc_pct

    # ── Build global color map ────────────────────────────────────
    palette = color_palette or px.colors.qualitative.Safe

    all_values = sorted(set(
        val for counts in col_counts.values()
        for val in counts.index
        if val not in ("NaN / None", "Other")
    ))

    color_map = {v: palette[i % len(palette)] for i, v in enumerate(all_values)}
    color_map["Other"]      = "#b0bec5"    # gray
    color_map["NaN / None"] = "#e05c5c"    # red

    # ── Build figure ──────────────────────────────────────────────
    fig           = go.Figure()
    added_legend  = set()

    # collect all unique values in display order
    # (most frequent first, Other and NaN always last)
    display_order = list(all_values)
    if "Other"      in color_map: display_order.append("Other")
    if "NaN / None" in color_map: display_order.append("NaN / None")

    for val_str in display_order:
        x_cols, y_pcts, hovers = [], [], []

        for col in cat_cols:
            counts  = col_counts[col]
            pct     = counts.get(val_str, 0.0)
            x_cols.append(col)
            y_pcts.append(pct)
            hovers.append(
                f"<b>{col}</b><br>"
                f"Value : <b>{val_str}</b><br>"
                f"Share : <b>{pct}%</b>"
                + ("<br><i>Missing values</i>" if val_str == "NaN / None" else "")
                + "<extra></extra>"
            )

        if all(p == 0 for p in y_pcts):
            continue   # skip values that never appear

        show_legend = val_str not in added_legend
        added_legend.add(val_str)

        fig.add_trace(go.Bar(
            name          = val_str,
            x             = x_cols,
            y             = y_pcts,
            marker_color  = color_map.get(val_str, "#aaaaaa"),
            text          = [
                f"{p}%" if p >= min_pct_label else ""
                for p in y_pcts
            ],
            textposition  = "inside",
            textfont      = dict(size=9),
            showlegend    = show_legend,
            hovertemplate = hovers,
        ))

    # ── Layout ────────────────────────────────────────────────────
    fig.update_layout(
        barmode       = "stack",
        title         = dict(
            text = "100% Distribution of categorical features",
            font = dict(size=15)
        ),
        yaxis         = dict(
            title     = "Percentage of rows (%)",
            range     = [0, 100],
            ticksuffix= "%",
            gridcolor = "#eeeeee",
        ),
        xaxis         = dict(
            title     = "Categorical feature",
            tickangle = 45,
            tickfont  = dict(size=10),
        ),
        legend        = dict(
            title       = "Values",
            orientation = "v",
            x=1.01, y=1,
            font        = dict(size=9),
            bgcolor     = "white",
            bordercolor = "#eee",
            borderwidth = 1,
        ),
        plot_bgcolor  = "white",
        paper_bgcolor = "white",
        height        = height,
        width         = max(800, len(cat_cols) * 80),
        bargap        = bar_gap,
    )

    # ── Print report ──────────────────────────────────────────────
    print(f"{'─'*50}")
    print(f"  Categorical Distribution Report")
    print(f"{'─'*50}")
    print(f"  Columns plotted  : {len(cat_cols)}")
    print(f"  Total rows       : {n_total:,}")
    print(f"  Top N per col    : {top_n}")
    print(f"  Include NaN      : {include_nan}")
    print(f"  Unique values    : {len(all_values)}")
    print(f"  Output file      : {output_file}")
    print(f"{'─'*50}")

    # ── Save ──────────────────────────────────────────────────────
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(fig.to_html())
    print(f"✅ Saved → {output_file}")

    if show:
        fig.show(renderer="browser")

    return fig


def plot_correlation_network(
    df: pd.DataFrame,
    top_n: int = 30,
    threshold: float = 0.5,
    method: str = "spearman",
    output_file: str = "correlation_network.html",
    height: str = "800px",
    bgcolor: str = "#1a1a2e",
    gravity: float = -50,
    spring_length: int = 150,
    show_buttons: bool = True,
    prefix_to_remove: list = None,
) -> nx.Graph:
    """
    Build and visualize an interactive feature correlation network using
    NetworkX for graph analysis and Pyvis for physics-based rendering.

    Each node represents one numeric feature. An edge is drawn between
    two nodes only if their absolute correlation exceeds ``threshold``.
    The resulting graph reveals which features move together (clusters),
    which are isolated (unique signal), and which act as bridges between
    groups (high betweenness centrality).

    Graph construction
    ------------------
    1. Select the top ``top_n`` most variable numeric columns.
    2. Compute a pairwise correlation matrix using ``method``.
    3. For each feature pair (i, j) with i < j:
           if |r(i, j)| >= threshold → add edge with weight = |r|
    4. Detect communities using greedy modularity maximisation:
           Q = (1/2m) * sum[ A_ij - k_i*k_j/2m ] * delta(c_i, c_j)
       where A is the adjacency matrix, k_i is node degree, m is total
       edges, and delta assigns nodes in the same community.
    5. Compute betweenness centrality for all nodes:
           BC(v) = sum_(s≠v≠t) [ sigma(s,t|v) / sigma(s,t) ]
       where sigma(s,t) = shortest paths from s to t, and
       sigma(s,t|v) = those passing through v.

    Visual encoding
    ---------------
    Node size    : proportional to degree (more connections = larger)
    Node color   : community membership (same color = same cluster)
    Edge color   : red = positive correlation, blue = negative
    Edge width   : proportional to |r| (stronger = thicker)
    Hover (node) : name, degree, community, betweenness centrality
    Hover (edge) : exact r value

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame. Only numeric columns are used.
    top_n : int, optional
        Number of highest-variance features to include. Default is 30.
        Reducing this speeds up rendering and reduces clutter.
    threshold : float, optional
        Minimum absolute correlation |r| to draw an edge. Default 0.5.
        Lower → more edges (denser graph).
        Higher → fewer edges (sparser, cleaner graph).
    method : str, optional
        Correlation method passed to ``pd.DataFrame.corr()``.
        One of "spearman", "pearson", "kendall". Default is "spearman".
    output_file : str, optional
        Path for the saved HTML file. Default "correlation_network.html".
    height : str, optional
        Height of the Pyvis canvas as a CSS string. Default "800px".
    bgcolor : str, optional
        Background color of the network canvas. Default "#1a1a2e" (dark).
    gravity : float, optional
        ForceAtlas2 gravity parameter. More negative = nodes spread wider.
        Default is -50.
    spring_length : int, optional
        Rest length of edges in the physics simulation. Default is 150.
        Longer = nodes pushed farther apart.
    show_buttons : bool, optional
        If True, adds a live physics control panel to the HTML output
        so the user can tweak forces interactively. Default is True.
    prefix_to_remove : list of str or None, optional
        List of string prefixes to strip from node labels for readability.
        e.g. ["SI_", "SDCOM_"] → "SI_TGH" displayed as "TGH".
        Default is ["SI_", "SDCOM_"].

    Returns
    -------
    G : networkx.Graph
        The constructed graph object. Use for further graph analysis:
            G.number_of_nodes()
            G.number_of_edges()
            nx.betweenness_centrality(G)
            nx.clustering(G)

    Raises
    ------
    ValueError
        If fewer than 2 numeric columns are found in ``df``.

    Examples
    --------
    >>> G = plot_correlation_network(df)

    >>> # stricter threshold — only very strong correlations
    >>> G = plot_correlation_network(df, threshold=0.7, top_n=40)

    >>> # Pearson instead of Spearman
    >>> G = plot_correlation_network(df, method="pearson")

    >>> # custom output + light background
    >>> G = plot_correlation_network(
    ...     df,
    ...     output_file      = "my_network.html",
    ...     bgcolor          = "#ffffff",
    ...     prefix_to_remove = ["SI_", "SDCOM_", "HM_"],
    ... )

    >>> # analyse the returned graph
    >>> G = plot_correlation_network(df, threshold=0.6)
    >>> print(nx.density(G))
    >>> print(sorted(nx.degree(G), key=lambda x: x[1], reverse=True)[:5])
    """

    # ── Validate ──────────────────────────────────────────────────
    num_df = df.select_dtypes(include="number")
    if num_df.shape[1] < 2:
        raise ValueError("Need at least 2 numeric columns to build a network.")

    prefix_to_remove = prefix_to_remove or ["SI_", "SDCOM_"]

    # ── Step 1: top N most variable columns ───────────────────────
    top_cols = (
        num_df.var().dropna()
        .nlargest(top_n)
        .index.tolist()
    )
    print(f"{'─'*55}")
    print(f"  Correlation Network Report")
    print(f"{'─'*55}")
    print(f"  Method           : {method}")
    print(f"  Top N columns    : {len(top_cols)}")
    print(f"  Threshold |r| ≥  : {threshold}")

    # ── Step 2: correlation matrix ────────────────────────────────
    corr = (
        df[top_cols]
        .apply(pd.to_numeric, errors="coerce")
        .corr(method=method)
    )

    # ── Step 3: build NetworkX graph ──────────────────────────────
    G = nx.Graph()
    for col in top_cols:
        G.add_node(col)

    for i in range(len(top_cols)):
        for j in range(i + 1, len(top_cols)):
            r = corr.iloc[i, j]
            if pd.notna(r) and abs(r) >= threshold:
                G.add_edge(
                    top_cols[i],
                    top_cols[j],
                    weight      = round(abs(r), 3),
                    correlation = round(r, 3),
                    sign        = "positive" if r > 0 else "negative"
                )

    print(f"  Nodes            : {G.number_of_nodes()}")
    print(f"  Edges            : {G.number_of_edges()}")
    print(f"  Density          : {nx.density(G):.3f}")

    # ── Step 4: community detection ───────────────────────────────
    communities = list(nx.community.greedy_modularity_communities(G))
    print(f"  Communities      : {len(communities)}")

    community_colors = [
        "#00838f", "#f0a500", "#e05c5c", "#6ab187",
        "#9b59b6", "#e67e22", "#1abc9c", "#3498db",
        "#e74c3c", "#2ecc71", "#f39c12", "#8e44ad",
    ]
    node_community = {}
    for i, comm in enumerate(communities):
        for node in comm:
            node_community[node] = i

    # ── Step 5: centrality metrics ────────────────────────────────
    degree      = dict(G.degree())
    betweenness = nx.betweenness_centrality(G, weight="weight")

    # top 5 hubs
    top_hubs = sorted(degree.items(), key=lambda x: x[1], reverse=True)[:5]
    print(f"\n  Top 5 hubs (by degree):")
    for node, deg in top_hubs:
        print(f"    {node:<35} degree={deg}  BC={betweenness[node]:.3f}")
    print(f"{'─'*55}")

    # ── Step 6: build Pyvis network ───────────────────────────────
    net = Network(
        height       = height,
        width        = "100%",
        bgcolor      = bgcolor,
        font_color   = "white",
        notebook     = False,
        cdn_resources= "in_line",
    )

    net.force_atlas_2based(
        gravity        = gravity,
        central_gravity= 0.01,
        spring_length  = spring_length,
        spring_strength= 0.08,
        damping        = 0.4,
    )

    # ── Add nodes ─────────────────────────────────────────────────
    for node in G.nodes():
        comm_idx = node_community.get(node, 0)
        color    = community_colors[comm_idx % len(community_colors)]
        size     = 15 + degree[node] * 8

        # clean label
        label = node
        for prefix in prefix_to_remove:
            label = label.replace(prefix, "")

        net.add_node(
            node,
            label            = label,
            title            = (
                f"<b>{node}</b><br>"
                f"Degree       : {degree[node]}<br>"
                f"Community    : {comm_idx + 1}<br>"
                f"Betweenness  : {betweenness[node]:.3f}"
            ),
            size             = size,
            color            = color,
            borderWidth      = 2,
            borderWidthSelected = 4,
            font             = {"size": 12, "color": "white"},
        )

    # ── Add edges ─────────────────────────────────────────────────
    for u, v, data in G.edges(data=True):
        r     = data["correlation"]
        color = "#ff6b6b" if r > 0 else "#74b9ff"

        net.add_edge(
            u, v,
            title  = f"r = {r}",
            color  = {"color": color, "opacity": 0.7},
            width  = data["weight"] * 6,
            smooth = {"type": "curvedCW", "roundness": 0.2},
        )

    if show_buttons:
        net.show_buttons(filter_=["physics"])

    # ── Save ──────────────────────────────────────────────────────
    html_content = net.generate_html()
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(html_content)

    print(f"\n✅ Saved → {output_file}")
    print(f"   Open in your browser to explore the network!")

    return G


def plot_boxplots_all_or_multiple_features(
    df: pd.DataFrame,
    columns: list = None,
    n_cols: int = 6,
    row_height: int = 180,
    box_color: str = "#00838f",
    outlier_color: str = "#e05c5c",
    output_file: str = "boxplots_all_numeric.html",
    show: bool = True,
) -> go.Figure:
    """
    Create an interactive Plotly grid of boxplots for all numeric columns
    in a DataFrame, with one subplot per feature.

    Each boxplot displays the five-number summary of the distribution:

        Minimum  — lower whisker end (excluding outliers beyond 1.5×IQR)
        Q1       — 25th percentile (lower edge of the box)
        Median   — 50th percentile (middle line)
        Q3       — 75th percentile (upper edge of the box)
        Maximum  — upper whisker end (excluding outliers beyond 1.5×IQR)

    Additionally, ``boxmean="sd"`` overlays two extra markers:

        Mean     — dashed line inside the box
        ±1 SD    — dotted lines showing one standard deviation band

    Outliers (points beyond 1.5 × IQR from Q1 or Q3) are shown as
    individual red dots, making them easy to spot across all features
    in a single view.

    The grid layout is computed automatically:

        n_rows = ceil(n_features / n_cols)

    Vertical spacing is capped to satisfy Plotly's constraint:

        v_spacing < 1 / (n_rows - 1)

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame. All numeric columns are used by default.
        Non-numeric values within numeric columns are coerced to NaN
        and excluded from each individual boxplot.
    columns : list of str or None, optional
        Explicit list of column names to plot. If None, all numeric
        columns in ``df`` are used automatically. Default is None.
    n_cols : int, optional
        Number of subplot columns in the grid. Default is 6.
        Increase for wider screens, decrease for fewer features.
    row_height : int, optional
        Height in pixels allocated per grid row. Default is 180.
        Total figure height = max(600, n_rows × row_height).
    box_color : str, optional
        Hex color for box fill and border. Default is "#00838f" (teal).
    outlier_color : str, optional
        Hex color for outlier points. Default is "#e05c5c" (red).
    output_file : str, optional
        File path for the saved HTML output. Default is
        "boxplots_all_numeric.html".
    show : bool, optional
        If True, opens the figure in the default browser. Default is True.

    Returns
    -------
    fig : plotly.graph_objects.Figure
        The fully constructed Plotly figure. Can be used for further
        customisation or embedding in dashboards.

    Raises
    ------
    ValueError
        If no numeric columns are found or none of the requested
        ``columns`` exist in ``df``.

    Examples
    --------
    >>> # all numeric columns — default
    >>> fig = plot_boxplots_all_or_multiple_features(df)

    >>> # specific columns only
    >>> fig = plot_boxplots_all_or_multiple_features(df, columns=["KgCO2EQ", "SDCOM_WIDTH", "SI_BW"])

    >>> # wider grid + taller rows
    >>> fig = plot_boxplots_all_or_multiple_features(df, n_cols=8, row_height=220)

    >>> # custom colors + save path
    >>> fig = plot_boxplots_all_or_multiple_features(
    ...     df,
    ...     box_color     = "#9b59b6",
    ...     outlier_color = "#f0a500",
    ...     output_file   = "my_boxplots.html",
    ... )

    >>> # get figure back for further editing
    >>> fig = plot_boxplots_all_or_multiple_features(df, show=False)
    >>> fig.update_layout(title="My Custom Title")
    >>> fig.show(renderer="browser")
    """

    # ── Resolve columns ───────────────────────────────────────────
    if columns is not None:
        missing = [c for c in columns if c not in df.columns]
        if missing:
            print(f"⚠️  Columns not found and skipped: {missing}")
        numeric_cols = [c for c in columns if c in df.columns]
    else:
        numeric_cols = df.select_dtypes(include="number").columns.tolist()

    if not numeric_cols:
        raise ValueError("No numeric columns found to plot.")

    # ── Grid layout ───────────────────────────────────────────────
    n_features = len(numeric_cols)
    n_rows     = math.ceil(n_features / n_cols)
    v_spacing  = round(min(0.06, 0.8 / max(n_rows, 2)), 4)
    h_spacing  = 0.04

    print(f"{'─'*50}")
    print(f"  Boxplot Grid Report")
    print(f"{'─'*50}")
    print(f"  Numeric columns  : {n_features}")
    print(f"  Grid             : {n_rows} rows × {n_cols} cols")
    print(f"  Vertical spacing : {v_spacing}")
    print(f"  Output file      : {output_file}")
    print(f"{'─'*50}")

    # ── Pad subplot titles to fill the full grid ──────────────────
    titles = numeric_cols + [""] * (n_rows * n_cols - n_features)

    # ── Create subplot grid ───────────────────────────────────────
    fig = make_subplots(
        rows               = n_rows,
        cols               = n_cols,
        subplot_titles     = titles,
        vertical_spacing   = v_spacing,
        horizontal_spacing = h_spacing,
    )

    # ── Add one boxplot per column ────────────────────────────────
    fill_color = (
        f"rgba({int(box_color[1:3], 16)},"
        f"{int(box_color[3:5], 16)},"
        f"{int(box_color[5:7], 16)},0.2)"
    )

    for i, col in enumerate(numeric_cols):
        row_pos = (i // n_cols) + 1
        col_pos = (i %  n_cols) + 1

        data = pd.to_numeric(df[col], errors="coerce").dropna()

        fig.add_trace(
            go.Box(
                y         = data,
                name      = col,
                boxmean   = "sd",
                marker    = dict(
                    color        = box_color,
                    outliercolor = outlier_color,
                    size         = 3,
                    line         = dict(
                        outliercolor = outlier_color,
                        outlierwidth = 1
                    )
                ),
                line      = dict(color=box_color),
                fillcolor = fill_color,
                showlegend= False,
                hovertemplate=(
                    f"<b>{col}</b><br>"
                    "Max    : %{upperfence:.2f}<br>"
                    "Q3     : %{q3:.2f}<br>"
                    "Median : %{median:.2f}<br>"
                    "Q1     : %{q1:.2f}<br>"
                    "Min    : %{lowerfence:.2f}<br>"
                    "<extra></extra>"
                )
            ),
            row = row_pos,
            col = col_pos,
        )

    # ── Layout ────────────────────────────────────────────────────
    fig.update_layout(
        title=dict(
            text = "Boxplot distribution — all numeric features",
            font = dict(size=16)
        ),
        height       = max(600, n_rows * row_height),
        width        = 1600,
        plot_bgcolor = "white",
        paper_bgcolor= "white",
        showlegend   = False,
    )

    for annotation in fig.layout.annotations:
        annotation.font.size  = 8
        annotation.font.color = "#444"

    fig.update_xaxes(showticklabels=False, showgrid=False)
    fig.update_yaxes(
        showgrid  = True,
        gridcolor = "#eeeeee",
        tickfont  = dict(size=8)
    )

    # ── Save ──────────────────────────────────────────────────────
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(fig.to_html())
    print(f"✅ Saved → {output_file}")

    if show:
        fig.show(renderer="browser")

    return fig


def _get_categorical_cols(df: pd.DataFrame) -> list:
    """
    Safely detect ALL string-like columns regardless of pandas dtype.
    Covers: object, category, StringDtype, large_string (Arrow).
    """
    return [
        col for col in df.columns
        if pd.api.types.is_string_dtype(df[col])
        or pd.api.types.is_object_dtype(df[col])
        or isinstance(df[col].dtype, pd.CategoricalDtype)
        or str(df[col].dtype) in ("string", "large_string")
    ]


def encode_categorical_columns(
    df: pd.DataFrame,
    method: str = "auto",
    target_col: str = None,
    max_cardinality_onehot: int = 10,
    ordinal_mappings: dict = None,
    drop_original: bool = True,
    verbose: bool = True,
) -> pd.DataFrame:
    """
    Encode all categorical / string columns into numeric representations
    suitable for machine learning models.

    Correctly detects ALL string-like dtypes:
        object, category, string (pd.StringDtype), large_string (Arrow)

    Encoding methods
    ----------------
    auto
        Automatically selects the best encoding per column:
            n_unique == 2               → Label Encoding  (0 / 1)
            n_unique <= max_cardinality → One-Hot Encoding
            n_unique >  max_cardinality → Frequency Encoding
            col in ordinal_mappings     → Ordinal Encoding

    label
        Maps each unique value to an integer 0..N-1 (alphabetical).
        Best for binary columns or tree-based models.
        ⚠️  Implies ordinal order — avoid for linear models.

    onehot
        Creates one binary column per unique value.
        Best for linear models and neural networks.
        ⚠️  Column count = n_unique — avoid for high cardinality.

    frequency
        Replaces each value with its proportion of total rows:
            freq(v) = count(v) / n_rows
        Works well with tree models. No column explosion.

    ordinal
        Maps values to integers per a user-supplied ordered list.
        Unknown values → -1.

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame. All string-like columns are encoded.
        Numeric columns passed through unchanged.
    method : str, optional
        One of "auto", "label", "onehot", "frequency", "ordinal".
        Default is "auto".
    target_col : str or None, optional
        Target column to exclude from encoding. Default is None.
    max_cardinality_onehot : int, optional
        Max unique values allowed for One-Hot in "auto" mode.
        Above this threshold → Frequency Encoding. Default is 10.
    ordinal_mappings : dict or None, optional
        Required for method="ordinal". Maps column → ordered list:
            {"SI_GROESSE": ["KLEIN", "MITTEL", "GROSS"]}
        Default is None.
    drop_original : bool, optional
        Remove original string columns after encoding. Default is True.
    verbose : bool, optional
        Print encoding report. Default is True.

    Returns
    -------
    df_encoded : pd.DataFrame
        DataFrame with all string columns encoded as numeric.
        Original index is preserved.

    Raises
    ------
    ValueError
        If method is invalid.
    ValueError
        If method="ordinal" but ordinal_mappings is not provided.

    Examples
    --------
    >>> df_enc = encode_categorical_columns(df, target_col="KgCO2EQ")

    >>> df_enc = encode_categorical_columns(df, method="onehot",
    ...                                     target_col="KgCO2EQ")

    >>> df_enc = encode_categorical_columns(df, method="frequency",
    ...                                     target_col="KgCO2EQ")

    >>> df_enc = encode_categorical_columns(
    ...     df,
    ...     method           = "ordinal",
    ...     target_col       = "KgCO2EQ",
    ...     ordinal_mappings = {
    ...         "SI_GROESSE": ["KLEIN", "MITTEL", "GROSS"],
    ...         "SI_KLASSE" : ["K1", "K2", "K3"],
    ...     }
    ... )
    """

    # ── Validate ──────────────────────────────────────────────────
    valid_methods = {"auto", "label", "onehot", "frequency", "ordinal"}
    if method not in valid_methods:
        raise ValueError(
            f"Invalid method '{method}'. "
            f"Choose from: {sorted(valid_methods)}"
        )

    if method == "ordinal" and not ordinal_mappings:
        raise ValueError(
            "method='ordinal' requires ordinal_mappings dict.\n"
            "Example: {'SI_GROESSE': ['KLEIN','MITTEL','GROSS']}"
        )

    # ── Detect ALL string-like columns ────────────────────────────
    cat_cols = _get_categorical_cols(df)

    # exclude target column
    if target_col and target_col in cat_cols:
        cat_cols.remove(target_col)

    # exclude numeric target that might have slipped in
    if target_col and target_col in cat_cols:
        cat_cols.remove(target_col)

    if not cat_cols:
        if verbose:
            print("⚠️  No categorical columns found — returning original DataFrame.")
        return df.copy()

    df_out = df.copy()
    report = []

    # ── Encode each column ────────────────────────────────────────
    for col in cat_cols:

        # fill NaN with placeholder before encoding
        series   = df_out[col].fillna("__NaN__").astype(str)
        n_unique = series.nunique()

        # decide method per column in auto mode
        col_method = method
        if method == "auto":
            if ordinal_mappings and col in ordinal_mappings:
                col_method = "ordinal"
            elif n_unique == 2:
                col_method = "label"
            elif n_unique <= max_cardinality_onehot:
                col_method = "onehot"
            else:
                col_method = "frequency"

        # ── LABEL ─────────────────────────────────────────────────
        if col_method == "label":
            le           = LabelEncoder()
            encoded      = le.fit_transform(series)
            df_out[f"{col}_enc"] = encoded
            mapping      = dict(zip(le.classes_,
                                    le.transform(le.classes_)))
            sample       = str(dict(list(mapping.items())[:4]))
            report.append((col, "label", n_unique, 1, sample))

        # ── ONE-HOT ───────────────────────────────────────────────
        elif col_method == "onehot":
            dummies = pd.get_dummies(
                series,
                prefix     = col,
                drop_first = False,
                dtype      = int,
            )
            # drop the NaN placeholder column if created
            nan_col = f"{col}___NaN__"
            if nan_col in dummies.columns:
                dummies = dummies.drop(columns=[nan_col])

            df_out  = pd.concat([df_out, dummies], axis=1)
            report.append((col, "onehot", n_unique,
                           dummies.shape[1],
                           f"→ {dummies.shape[1]} new columns"))

        # ── FREQUENCY ─────────────────────────────────────────────
        elif col_method == "frequency":
            freq_map         = series.value_counts(normalize=True).to_dict()
            df_out[f"{col}_freq"] = series.map(freq_map).round(4)
            top3             = dict(list(
                sorted(freq_map.items(), key=lambda x: -x[1])[:3]
            ))
            report.append((col, "frequency", n_unique, 1,
                           f"top3: {top3}"))

        # ── ORDINAL ───────────────────────────────────────────────
        elif col_method == "ordinal":
            mapping = ordinal_mappings.get(col) if ordinal_mappings else None

            if not mapping:
                # fallback to label if no mapping provided
                le = LabelEncoder()
                df_out[f"{col}_enc"] = le.fit_transform(series)
                report.append((col, "label (fallback)", n_unique, 1,
                               "No ordinal mapping → used label"))
            else:
                ord_map = {str(v): i for i, v in enumerate(mapping)}
                df_out[f"{col}_ord"] = (
                    series.map(ord_map)
                    .fillna(-1)
                    .astype(int)
                )
                sample = str({v: i for i, v in enumerate(mapping)})
                report.append((col, "ordinal", n_unique, 1, sample))

        # drop original string column
        if drop_original:
            df_out = df_out.drop(columns=[col], errors="ignore")

    # ── Verbose report ────────────────────────────────────────────
    if verbose:
        sep = "─" * 72
        print(f"\n{sep}")
        print(f"  Categorical Encoding Report  (method='{method}')")
        print(f"{sep}")
        print(f"  String dtypes detected : object, category, string, large_string")
        print(f"  Columns encoded        : {len(cat_cols)}")
        print(f"  One-hot threshold      : ≤ {max_cardinality_onehot} unique values")
        print(f"  Drop originals         : {drop_original}")
        print(f"  Target excluded        : {target_col or 'None'}")
        print(f"{sep}")
        print(
            f"\n  {'Column':<35} {'Method':<18} "
            f"{'Unique':>7} {'New cols':>9}  Notes"
        )
        print(
            f"  {'─'*35} {'─'*18} "
            f"{'─'*7} {'─'*9}  {'─'*30}"
        )
        for col, enc, uniq, ncols, notes in report:
            print(
                f"  {col:<35} {enc:<18} "
                f"{uniq:>7} {ncols:>9}  {str(notes)[:50]}"
            )

        remaining_cat = _get_categorical_cols(df_out)
        if target_col and target_col in remaining_cat:
            remaining_cat.remove(target_col)

        print(f"\n  Input  shape : {df.shape}")
        print(f"  Output shape : {df_out.shape}")
        print(f"  Numeric cols : {df_out.select_dtypes(include='number').shape[1]}")

        if remaining_cat:
            print(f"  ⚠️  String cols still remaining: {remaining_cat}")
        else:
            print(f"  ✅ All categorical columns successfully encoded")
        print(f"{sep}\n")

    return df_out


def elbow_plot(data_table, k_range=(2, 12)):
    """
    Compute and plot the KMeans elbow curve.
    Replicates KElbowVisualizer from Yellowbrick.

    Parameters
    ----------
    data_table : pd.DataFrame or np.ndarray
    k_range    : tuple (min_k, max_k)

    Returns
    -------
    elbow_value : int  — the optimal number of clusters
    """

    # ── Preprocess — impute and scale ────────────────────────────
    imp = SimpleImputer(strategy="median")
    sca = StandardScaler()
    X   = sca.fit_transform(imp.fit_transform(data_table))

    # ── Fit KMeans for each k ─────────────────────────────────────
    k_values    = range(k_range[0], k_range[1] + 1)
    inertias    = []
    fit_times   = []

    print("Fitting KMeans for k =", list(k_values))
    for k in k_values:
        model = KMeans(n_clusters=k, random_state=4, n_init=10)
        model.fit(X)
        inertias.append(model.inertia_)
        print(f"  k={k:2d}  inertia={model.inertia_:,.1f}")

    # ── Detect elbow using maximum curvature ──────────────────────
    # draw line from first to last point, find max distance to it
    k_arr  = np.array(list(k_values), dtype=float)
    in_arr = np.array(inertias, dtype=float)

    # normalize both axes to [0,1] for fair distance comparison
    k_norm  = (k_arr  - k_arr.min())  / (k_arr.max()  - k_arr.min())
    in_norm = (in_arr - in_arr.min()) / (in_arr.max() - in_arr.min())

    # vector from first to last point
    vec    = np.array([k_norm[-1] - k_norm[0],
                       in_norm[-1] - in_norm[0]])
    vec    = vec / np.linalg.norm(vec)

    # perpendicular distance from each point to the line
    dists  = []
    for i in range(len(k_norm)):
        point = np.array([k_norm[i] - k_norm[0],
                          in_norm[i] - in_norm[0]])
        dist  = abs(vec[0] * point[1] - vec[1] * point[0])
        dists.append(dist)

    elbow_idx   = int(np.argmax(dists))
    elbow_value = list(k_values)[elbow_idx]
    print(f"\n✅ Elbow detected at k = {elbow_value}")

    # ── Plot ──────────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(9, 5))

    # main line
    ax.plot(
        k_values, inertias,
        color="#4c8eda", linewidth=2.5,
        marker="o", markersize=7,
        markerfacecolor="white",
        markeredgecolor="#4c8eda",
        markeredgewidth=2,
        label="distortion score"
    )

    # elbow vertical line
    ax.axvline(
        x=elbow_value,
        color="#e05c5c",
        linestyle="--",
        linewidth=1.8,
        label=f"elbow at k={elbow_value}"
    )

    # elbow point highlight
    ax.scatter(
        [elbow_value], [inertias[elbow_idx]],
        color="#e05c5c", s=120, zorder=5
    )

    # annotation
    ax.annotate(
        f"elbow k={elbow_value}",
        xy=(elbow_value, inertias[elbow_idx]),
        xytext=(elbow_value + 0.4, inertias[elbow_idx]),
        fontsize=10,
        color="#e05c5c",
        arrowprops=dict(arrowstyle="->",
                        color="#e05c5c",
                        lw=1.5)
    )

    ax.set_xlabel("k  (number of clusters)", fontsize=12)
    ax.set_ylabel("distortion score (inertia)", fontsize=12)
    ax.set_title(
        "KMeans Distortion Score Elbow\n"
        "for Optimal k",
        fontsize=13
    )
    ax.xaxis.set_major_locator(ticker.MultipleLocator(1))
    ax.yaxis.set_major_formatter(
        ticker.FuncFormatter(lambda x, _: f"{x:,.0f}")
    )
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig("kelbow.png", dpi=150, bbox_inches="tight")
    plt.show()
    print("✅ Saved → kelbow.png")

    return elbow_value


def stratified_split(
    df: pd.DataFrame,
    target_col: str,
    test_size: float = 0.2,
    n_bins: int = 5,
    random_state: int = 42,
    verbose: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.Series, pd.Series]:
    """
    Perform a stratified train-test split on a DataFrame with a
    continuous numeric target column.

    Why stratify a numeric target?
    --------------------------------
    Standard ``train_test_split`` with ``stratify`` only accepts
    categorical labels. For a continuous target like CO₂ emissions
    or price, a naive random split may place all high-value or
    low-value samples in one split, making the train and test sets
    unrepresentative of the full distribution.

    Solution — bin-based stratification:

        1. Discretize the target into ``n_bins`` equal-frequency bins
           using KBinsDiscretizer (strategy="quantile"):

               bin edges: Q(0), Q(1/n), Q(2/n), ..., Q(1)

           Each bin contains approximately the same number of samples,
           ensuring the full value range is represented in every bin.

        2. Use the bin labels as stratification keys for
           ``sklearn.model_selection.train_test_split``.

        3. Discard the bin labels after splitting — the original
           continuous target values are preserved in y_train / y_test.

    This guarantees that both train and test sets contain samples from
    every part of the target distribution (low, medium, high values).

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame containing features and the target column.
    target_col : str
        Name of the continuous numeric target column to split on.
        Must exist in ``df`` and contain numeric values.
    test_size : float, optional
        Fraction of samples allocated to the test set. Must be in
        range (0, 1). Default is 0.2 (80% train / 20% test).
    n_bins : int, optional
        Number of equal-frequency bins used for stratification.
        More bins = finer stratification but requires more samples.
        Rule of thumb: each bin should contain at least 2 samples.

            Minimum samples needed ≈ n_bins / test_size

        Default is 5.
    random_state : int, optional
        Random seed for reproducibility. Default is 42.
    verbose : bool, optional
        Print a detailed split report. Default is True.

    Returns
    -------
    X_train : pd.DataFrame — training features
    X_test  : pd.DataFrame — test features
    y_train : pd.Series    — training target values
    y_test  : pd.Series    — test target values

    Raises
    ------
    ValueError
        If ``target_col`` is not found in ``df``.
    ValueError
        If the target column contains too many NaN values to split.
    ValueError
        If ``n_bins`` is too large for the available sample count.

    Examples
    --------
    >>> X_train, X_test, y_train, y_test = stratified_split(
    ...     df, target_col="KgCO2EQ"
    ... )

    >>> # finer stratification, larger test set
    >>> X_train, X_test, y_train, y_test = stratified_split(
    ...     df,
    ...     target_col   = "KgCO2EQ",
    ...     test_size    = 0.25,
    ...     n_bins       = 10,
    ...     random_state = 0,
    ... )

    >>> # plug directly into sklearn pipeline
    >>> pipeline.fit(X_train, y_train)
    >>> y_pred = pipeline.predict(X_test)
    """

    # ── Validate ──────────────────────────────────────────────────
    if target_col not in df.columns:
        raise ValueError(
            f"Target column '{target_col}' not found in DataFrame.\n"
            f"Available columns: {list(df.columns[:10])} ..."
        )

    # ── Drop rows where target is NaN ────────────────────────────
    df_clean   = df.dropna(subset=[target_col]).copy()
    n_dropped  = len(df) - len(df_clean)

    if n_dropped > 0 and verbose:
        print(f"⚠️  Dropped {n_dropped} rows with NaN in '{target_col}'")

    if len(df_clean) < n_bins * 2:
        raise ValueError(
            f"Too few samples ({len(df_clean)}) for {n_bins} bins. "
            f"Reduce n_bins or provide more data."
        )

    # ── Separate features and target ──────────────────────────────
    X = df_clean.drop(columns=[target_col])
    y = pd.to_numeric(df_clean[target_col], errors="coerce")

    # ── Bin the target for stratification ────────────────────────
    binner  = KBinsDiscretizer(
        n_bins   = n_bins,
        encode   = "ordinal",
        strategy = "quantile",    # equal-frequency bins
    )
    y_bins  = binner.fit_transform(y.values.reshape(-1, 1)).ravel().astype(int)

    # ── Stratified split ──────────────────────────────────────────
    X_train, X_test, y_train, y_test, bins_train, bins_test = train_test_split(
        X, y, y_bins,
        test_size    = test_size,
        random_state = random_state,
        stratify     = y_bins,
    )

    # ── Verbose report ────────────────────────────────────────────
    if verbose:
        sep = "─" * 60
        print(f"\n{sep}")
        print(f"  Stratified Train-Test Split Report")
        print(f"{sep}")
        print(f"  Target column    : {target_col}")
        print(f"  Total samples    : {len(df_clean):,}")
        print(f"  Train samples    : {len(X_train):,}  ({(1-test_size)*100:.0f}%)")
        print(f"  Test  samples    : {len(X_test):,}  ({test_size*100:.0f}%)")
        print(f"  Features         : {X_train.shape[1]}")
        print(f"  Stratify bins    : {n_bins}")
        print(f"  Random state     : {random_state}")
        print(f"{sep}")

        # target distribution comparison
        print(f"\n  Target distribution comparison ({target_col}):\n")
        print(f"  {'Metric':<18} {'Full':>12} {'Train':>12} {'Test':>12}")
        print(f"  {'─'*18} {'─'*12} {'─'*12} {'─'*12}")

        for metric, fn in [
            ("mean",   np.mean),
            ("median", np.median),
            ("std",    np.std),
            ("min",    np.min),
            ("max",    np.max),
        ]:
            print(
                f"  {metric:<18} "
                f"{fn(y.values):>12.2f} "
                f"{fn(y_train.values):>12.2f} "
                f"{fn(y_test.values):>12.2f}"
            )

        # bin distribution
        print(f"\n  Bin distribution (stratification quality):\n")
        print(f"  {'Bin':>5} {'Range':>25} {'Full':>8} {'Train':>8} {'Test':>8}")
        print(f"  {'─'*5} {'─'*25} {'─'*8} {'─'*8} {'─'*8}")

        edges = binner.bin_edges_[0]
        for b in range(n_bins):
            rng        = f"[{edges[b]:.1f} – {edges[b+1]:.1f}]"
            full_pct   = (y_bins == b).mean() * 100
            train_pct  = (bins_train == b).mean() * 100
            test_pct   = (bins_test  == b).mean() * 100
            print(
                f"  {b+1:>5} {rng:>25} "
                f"{full_pct:>7.1f}% "
                f"{train_pct:>7.1f}% "
                f"{test_pct:>7.1f}%"
            )

        print(f"\n{sep}\n")

    return X_train, X_test, y_train, y_test


def cross_validate_random_forest(
    X_train: pd.DataFrame,
    y_train: pd.Series,
    n_folds: int = 5,
    n_estimators: int = 200,
    max_depth=None,
    min_samples_split: int = 5,
    min_samples_leaf: int = 2,
    max_features: str = "sqrt",
    random_state: int = 42,
    verbose: bool = True,
) -> dict:
    """
    Run K-Fold cross-validation for a Random Forest Regressor pipeline.

    Pipeline structure per fold
    ---------------------------
    Each fold runs the full preprocessing + model pipeline:

        X_train_fold
            ↓
        SimpleImputer (median)    — fills NaN with column median
            ↓                        fitted on train fold only
        StandardScaler            — zero mean, unit variance
            ↓                        fitted on train fold only
        RandomForestRegressor     — ensemble of decision trees
            ↓
        predictions on val fold

    Why fit preprocessing inside each fold?
    ----------------------------------------
    Fitting the imputer and scaler on the FULL training set before
    cross-validation causes data leakage — the validation fold
    indirectly influences the preprocessing statistics. By fitting
    inside each fold, each validation fold is truly unseen.

    K-Fold strategy
    ---------------
    The training set is split into K equal folds. For each fold k:

        Train on folds: {1, 2, ..., K} \ {k}
        Validate on   : fold k

    This gives K independent validation scores. The mean and standard
    deviation of these scores estimate generalisation performance:

        mean(scores)  → expected performance on new data
        std(scores)   → stability — low std = robust model

    Metrics computed per fold
    -------------------------
        R²    = 1 - SS_res / SS_tot          (1.0 = perfect)
        MAE   = mean |y - ŷ|                 (same units as target)
        RMSE  = sqrt(mean (y - ŷ)²)          (penalises large errors)
        MAPE  = mean |y - ŷ| / |y| × 100    (% error)

    Parameters
    ----------
    X_train : pd.DataFrame
        Training features. May contain NaN — handled by the pipeline.
        Must be fully numeric (encode categoricals first).
    y_train : pd.Series
        Training target values (continuous numeric).
    n_folds : int, optional
        Number of cross-validation folds (K). Default is 5.
        Rule of thumb:
            K = 5  → good default for most datasets
            K = 10 → lower bias, higher compute cost
            K = 3  → use for very small datasets
    n_estimators : int, optional
        Number of trees in the Random Forest. Default is 200.
    max_depth : int or None, optional
        Maximum depth of each tree. None = grow until pure leaves.
        Default is None.
    min_samples_split : int, optional
        Minimum samples required to split a node. Default is 5.
    min_samples_leaf : int, optional
        Minimum samples required at each leaf node. Default is 2.
    max_features : str or float, optional
        Features considered at each split. Default is "sqrt".
            "sqrt"  → sqrt(n_features) — standard for RF
            "log2"  → log2(n_features)
            0.5     → 50% of features
    random_state : int, optional
        Reproducibility seed. Default is 42.
    verbose : bool, optional
        Print fold-by-fold and summary results. Default is True.

    Returns
    -------
    results : dict
        Full cross-validation results with keys:
            pipeline       : fitted Pipeline object
            fold_scores    : pd.DataFrame — per-fold metrics
            mean_scores    : dict         — mean of each metric
            std_scores     : dict         — std of each metric
            cv_r2_scores   : np.ndarray  — raw R² per fold
    """

    # ── Validate ──────────────────────────────────────────────────
    if not isinstance(X_train, pd.DataFrame):
        X_train = pd.DataFrame(X_train)
    if not isinstance(y_train, (pd.Series, np.ndarray)):
        y_train = pd.Series(y_train)

    # ── Build pipeline ────────────────────────────────────────────
    pipeline = Pipeline([
        ("imputer", SimpleImputer(strategy="median")),
        ("scaler",  StandardScaler()),
        ("model",   RandomForestRegressor(
            n_estimators     = n_estimators,
            max_depth        = max_depth,
            min_samples_split= min_samples_split,
            min_samples_leaf = min_samples_leaf,
            max_features     = max_features,
            random_state     = random_state,
            n_jobs           = -1,
            oob_score        = False,   # disabled inside CV
        ))
    ])

    # ── K-Fold cross-validation ───────────────────────────────────
    kf = KFold(
        n_splits  = n_folds,
        shuffle   = True,
        random_state = random_state
    )

    # ── Per-fold scoring ──────────────────────────────────────────
    fold_records = []

    if verbose:
        sep = "─" * 65
        print(f"\n{sep}")
        print(f"  Random Forest — {n_folds}-Fold Cross-Validation")
        print(f"{sep}")
        print(f"  Features         : {X_train.shape[1]}")
        print(f"  Training samples : {X_train.shape[0]}")
        print(f"  Folds            : {n_folds}")
        print(f"  n_estimators     : {n_estimators}")
        print(f"  max_depth        : {max_depth or 'None (full)'}")
        print(f"  max_features     : {max_features}")
        print(f"{sep}")
        print(f"\n  {'Fold':>5} {'R²':>9} {'MAE':>12} {'RMSE':>12} {'MAPE %':>9}")
        print(f"  {'─'*5} {'─'*9} {'─'*12} {'─'*12} {'─'*9}")

    X_arr = X_train.values
    y_arr = y_train.values

    for fold_idx, (train_idx, val_idx) in enumerate(kf.split(X_arr), 1):

        X_fold_train = X_arr[train_idx]
        X_fold_val   = X_arr[val_idx]
        y_fold_train = y_arr[train_idx]
        y_fold_val   = y_arr[val_idx]

        # fit pipeline on train fold only
        pipeline.fit(X_fold_train, y_fold_train)

        # predict on validation fold
        y_pred = pipeline.predict(X_fold_val)

        # compute metrics
        r2   = r2_score(y_fold_val, y_pred)
        mae  = mean_absolute_error(y_fold_val, y_pred)
        rmse = np.sqrt(mean_squared_error(y_fold_val, y_pred))
        mape = mean_absolute_percentage_error(y_fold_val, y_pred) * 100

        fold_records.append({
            "fold" : fold_idx,
            "r2"   : round(r2,   4),
            "mae"  : round(mae,  4),
            "rmse" : round(rmse, 4),
            "mape" : round(mape, 4),
            "n_train": len(train_idx),
            "n_val"  : len(val_idx),
        })

        if verbose:
            print(
                f"  {fold_idx:>5} "
                f"{r2:>9.4f} "
                f"{mae:>12.2f} "
                f"{rmse:>12.2f} "
                f"{mape:>9.2f}%"
            )

    # ── Summary statistics ────────────────────────────────────────
    fold_df = pd.DataFrame(fold_records)

    mean_scores = {
        "r2"  : fold_df["r2"].mean(),
        "mae" : fold_df["mae"].mean(),
        "rmse": fold_df["rmse"].mean(),
        "mape": fold_df["mape"].mean(),
    }
    std_scores = {
        "r2"  : fold_df["r2"].std(),
        "mae" : fold_df["mae"].std(),
        "rmse": fold_df["rmse"].std(),
        "mape": fold_df["mape"].std(),
    }

    if verbose:
        print(f"\n  {'─'*5} {'─'*9} {'─'*12} {'─'*12} {'─'*9}")
        print(
            f"  {'mean':>5} "
            f"{mean_scores['r2']:>9.4f} "
            f"{mean_scores['mae']:>12.2f} "
            f"{mean_scores['rmse']:>12.2f} "
            f"{mean_scores['mape']:>9.2f}%"
        )
        print(
            f"  {'std':>5} "
            f"{std_scores['r2']:>9.4f} "
            f"{std_scores['mae']:>12.2f} "
            f"{std_scores['rmse']:>12.2f} "
            f"{std_scores['mape']:>9.2f}%"
        )
        print(f"\n{sep}")

        # interpretation
        r2_mean = mean_scores["r2"]
        r2_std  = std_scores["r2"]
        quality = (
            "✅ Excellent"  if r2_mean >= 0.85 else
            "✅ Good"       if r2_mean >= 0.70 else
            "⚠️  Moderate"  if r2_mean >= 0.50 else
            "❌ Poor"
        )
        stability = (
            "✅ Stable"     if r2_std  <= 0.05 else
            "⚠️  Variable"  if r2_std  <= 0.10 else
            "❌ Unstable"
        )
        print(f"\n  CV R² = {r2_mean:.4f} ± {r2_std:.4f}")
        print(f"  Model quality : {quality}")
        print(f"  Stability     : {stability}")
        print(f"{sep}\n")

    # ── Visualisation ─────────────────────────────────────────────
    _plot_cv_results(fold_df, mean_scores, std_scores, n_folds)

    # ── Refit on full training set ────────────────────────────────
    pipeline.fit(X_train.values, y_arr)

    return {
        "pipeline"    : pipeline,
        "fold_scores" : fold_df,
        "mean_scores" : mean_scores,
        "std_scores"  : std_scores,
        "cv_r2_scores": fold_df["r2"].values,
    }


def _plot_cv_results(fold_df, mean_scores, std_scores, n_folds):
    """Plot fold-by-fold metrics and distributions."""

    fig = plt.figure(figsize=(14, 10))
    gs  = gridspec.GridSpec(2, 2, hspace=0.4, wspace=0.35)

    folds   = fold_df["fold"].values
    metrics = [
        ("r2",   "R²",     "#4c8eda"),
        ("mae",  "MAE",    "#f0a500"),
        ("rmse", "RMSE",   "#6ab187"),
        ("mape", "MAPE %", "#e05c5c"),
    ]

    for idx, (key, label, color) in enumerate(metrics):
        ax  = fig.add_subplot(gs[idx // 2, idx % 2])
        vals = fold_df[key].values
        mean = mean_scores[key]
        std  = std_scores[key]

        # bar per fold
        bars = ax.bar(
            folds, vals,
            color=color, alpha=0.7,
            edgecolor="white", linewidth=0.8,
            zorder=3
        )

        # mean line
        ax.axhline(
            mean, color="#333", linewidth=1.5,
            linestyle="--", label=f"mean = {mean:.3f}", zorder=4
        )

        # ± 1 std band
        ax.axhspan(
            mean - std, mean + std,
            alpha=0.12, color=color, zorder=2,
            label=f"± std = {std:.3f}"
        )

        # value labels on bars
        for bar, val in zip(bars, vals):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + (vals.max() - vals.min()) * 0.02,
                f"{val:.3f}",
                ha="center", va="bottom", fontsize=8.5, color="#333"
            )

        ax.set_title(f"{label} per fold", fontsize=11, fontweight="bold")
        ax.set_xlabel("Fold",   fontsize=10)
        ax.set_ylabel(label,    fontsize=10)
        ax.set_xticks(folds)
        ax.set_xticklabels([f"Fold {f}" for f in folds], fontsize=9)
        ax.legend(fontsize=8, loc="upper right")
        ax.grid(True, axis="y", alpha=0.3)
        ax.set_facecolor("#fafafa")

    fig.suptitle(
        f"Random Forest — {n_folds}-Fold Cross-Validation Results\n"
        f"CV R² = {mean_scores['r2']:.4f} ± {std_scores['r2']:.4f}",
        fontsize=13, fontweight="bold", y=1.01
    )

    plt.savefig("cv_results.png", dpi=150, bbox_inches="tight")
    plt.show()
    print("✅ Saved → cv_results.png")


def _plot_tuning_results(cv_df, baseline_cv, best_cv, test_r2, n_iter):
    """Plot top combinations ranked by CV R²."""

    fig, axes = plt.subplots(1, 2, figsize=(14, 6))

    # ── Left: ranked bar chart of top 20 combinations ─────────────
    ax   = axes[0]
    top  = cv_df.head(20)
    ranks = top.index.values
    means = top["CV R² mean"].values
    stds  = top["CV R² std"].values

    colors = ["#f0a500" if i == 1 else "#4c8eda" for i in ranks]
    bars   = ax.barh(
        ranks[::-1], means[::-1],
        xerr=stds[::-1],
        color=colors[::-1],
        edgecolor="white", linewidth=0.5,
        capsize=3, ecolor="#aaa",
    )
    ax.axvline(baseline_cv, color="#e05c5c", linewidth=1.5,
               linestyle="--", label=f"Baseline CV R² = {baseline_cv:.3f}")
    ax.axvline(test_r2,     color="#6ab187", linewidth=1.5,
               linestyle=":",  label=f"Test R²        = {test_r2:.3f}")
    ax.set_xlabel("CV R²", fontsize=11)
    ax.set_ylabel("Rank", fontsize=11)
    ax.set_title(f"Top 20 / {n_iter} Parameter Combinations\n"
                 f"(error bars = CV std)", fontsize=11, fontweight="bold")
    ax.legend(fontsize=9)
    ax.grid(True, axis="x", alpha=0.3)
    ax.set_facecolor("#fafafa")

    # ── Right: comparison bar (baseline vs tuned CV vs test) ───────
    ax     = axes[1]
    labels = ["Baseline\nCV R²", "Tuned\nCV R²", "Test\nR²"]
    values = [baseline_cv, best_cv, test_r2]
    colors = ["#aaaaaa", "#f0a500", "#6ab187"]

    bars = ax.bar(labels, values, color=colors,
                  edgecolor="white", width=0.5)

    for bar, val in zip(bars, values):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.005,
            f"{val:.4f}",
            ha="center", va="bottom",
            fontsize=11, fontweight="bold", color="#333"
        )

    ax.set_ylim(0, min(1.0, max(values) * 1.15))
    ax.set_ylabel("R²", fontsize=11)
    ax.set_title("Baseline  vs  Tuned CV  vs  Test",
                 fontsize=11, fontweight="bold")
    ax.grid(True, axis="y", alpha=0.3)
    ax.set_facecolor("#fafafa")

    plt.suptitle("RandomizedSearchCV — Hyperparameter Tuning Results",
                 fontsize=13, fontweight="bold", y=1.02)
    plt.tight_layout()
    plt.savefig("tuning_results.png", dpi=150, bbox_inches="tight")
    plt.show()
    print("✅ Saved → tuning_results.png")


def tune_random_forest(
    pipeline,
    X_train: pd.DataFrame,
    y_train: pd.Series,
    X_test: pd.DataFrame,
    y_test: pd.Series,
    n_iter: int = 30,
    cv: int = 5,
    scoring: str = "r2",
    random_state: int = 42,
    param_dist: dict = None,
    verbose: bool = True,
) -> dict:
    """
    Tune a Random Forest regression pipeline using RandomizedSearchCV
    and evaluate the best model on the held-out test set.

    Process
    -------
    1. Compute baseline CV score  — default hyperparameters on X_train
    2. RandomizedSearchCV         — try ``n_iter`` random combinations,
                                    each evaluated with ``cv``-fold CV
                                    on X_train only
    3. Final evaluation           — best model on X_test (touched once)
    4. Comparison report          — baseline vs tuned vs test
    5. Visualisation              — top combinations ranked by CV R²

    Why RandomizedSearch and not GridSearch?
    ----------------------------------------
    GridSearchCV tries every combination in the grid:
        n_estimators=[150,200,300,400,500] × max_depth=[None,10,20,30]
        × min_samples_leaf=[1..8] × max_features=[sqrt,log2,0.5]
        = 5 × 4 × 8 × 3 = 480 combinations × cv folds = expensive

    RandomizedSearchCV samples ``n_iter`` combinations at random:
        n_iter=30 × cv=5 = 150 model fits — much cheaper
        Empirically finds near-optimal params in a fraction of the time.

    Parameters
    ----------
    pipeline : sklearn.pipeline.Pipeline
        A fitted or unfitted Pipeline containing an "imputer", "scaler",
        and "model" (RandomForestRegressor) step.
    X_train : pd.DataFrame
        Training features. NaN handled by the pipeline imputer.
    y_train : pd.Series
        Training target values.
    X_test : pd.DataFrame
        Test features. Touched ONLY for final evaluation.
    y_test : pd.Series
        Test target values.
    n_iter : int, optional
        Number of random parameter combinations to try. Default is 30.
        More = better coverage but slower.
    cv : int, optional
        Number of cross-validation folds per combination. Default is 5.
    scoring : str, optional
        Sklearn scoring metric. Default is "r2".
        Others: "neg_mean_absolute_error", "neg_root_mean_squared_error"
    random_state : int, optional
        Reproducibility seed. Default is 42.
    param_dist : dict or None, optional
        Custom hyperparameter search space. If None, uses a sensible
        default distribution for RandomForestRegressor:

            n_estimators     : randint(150, 500)
            max_depth        : [None, 10, 20, 30]
            min_samples_leaf : randint(1, 8)
            max_features     : ["sqrt", "log2", 0.5]

    verbose : bool, optional
        Print detailed report and plots. Default is True.

    Returns
    -------
    results : dict
        best_model      : fitted Pipeline with best hyperparameters
        best_params     : dict   — best hyperparameter values
        best_cv_score   : float  — best mean CV R² during tuning
        baseline_cv     : float  — CV R² before tuning
        test_metrics    : dict   — R², MAE, RMSE, MAPE on X_test
        cv_results_df   : pd.DataFrame — all combinations ranked
        search          : RandomizedSearchCV — full search object

    Examples
    --------
    >>> results = tune_random_forest(
    ...     pipeline, X_train, y_train, X_test, y_test
    ... )
    >>> best_model  = results["best_model"]
    >>> best_params = results["best_params"]
    >>> y_pred = best_model.predict(X_test)

    >>> # custom search space
    >>> results = tune_random_forest(
    ...     pipeline, X_train, y_train, X_test, y_test,
    ...     n_iter = 50,
    ...     param_dist = {
    ...         "model__n_estimators"    : randint(100, 800),
    ...         "model__max_depth"       : [None, 5, 10, 20, 30],
    ...         "model__min_samples_leaf": randint(1, 10),
    ...         "model__max_features"    : ["sqrt", "log2", 0.3, 0.5],
    ...     }
    ... )
    """

    # ── Default search space ──────────────────────────────────────
    if param_dist is None:
        param_dist = {
            "model__n_estimators"    : randint(150, 500),
            "model__max_depth"       : [None, 10, 20, 30],
            "model__min_samples_leaf": randint(1, 8),
            "model__max_features"    : ["sqrt", "log2", 0.5],
        }

    # ── Step 1: baseline CV (default params) ─────────────────────
    if verbose:
        print("─" * 62)
        print("  Step 1 — Baseline Cross-Validation (default params)")
        print("─" * 62)

    baseline_scores = cross_val_score(
        pipeline, X_train, y_train,
        cv=cv, scoring=scoring, n_jobs=-1
    )
    baseline_cv = baseline_scores.mean()
    baseline_std = baseline_scores.std()

    if verbose:
        print(f"  Baseline CV R²  : {baseline_cv:.4f} ± {baseline_std:.4f}\n")

    # ── Step 2: RandomizedSearchCV ────────────────────────────────
    if verbose:
        print("─" * 62)
        print(f"  Step 2 — RandomizedSearchCV  ({n_iter} combos × {cv} folds)")
        print("─" * 62)

    search = RandomizedSearchCV(
        estimator           = pipeline,
        param_distributions = param_dist,
        n_iter              = n_iter,
        cv                  = cv,
        scoring             = scoring,
        n_jobs              = -1,
        random_state        = random_state,
        verbose             = 1 if verbose else 0,
        refit               = True,
    )
    search.fit(X_train, y_train)

    best_model  = search.best_estimator_
    best_params = search.best_params_
    best_cv     = search.best_score_

    # ── Step 3: final evaluation on test set ─────────────────────
    y_pred = best_model.predict(X_test)

    test_metrics = {
        "r2"  : round(r2_score(y_test, y_pred), 4),
        "mae" : round(mean_absolute_error(y_test, y_pred), 4),
        "rmse": round(np.sqrt(mean_squared_error(y_test, y_pred)), 4),
        "mape": round(mean_absolute_percentage_error(y_test, y_pred) * 100, 4),
    }

    # ── Step 4: build cv results DataFrame ───────────────────────
    cv_df = (
        pd.DataFrame(search.cv_results_)
        .sort_values("mean_test_score", ascending=False)
        [[
            "mean_test_score", "std_test_score",
            "param_model__n_estimators",
            "param_model__max_depth",
            "param_model__min_samples_leaf",
            "param_model__max_features",
        ]]
        .rename(columns={
            "mean_test_score"               : "CV R² mean",
            "std_test_score"                : "CV R² std",
            "param_model__n_estimators"     : "n_estimators",
            "param_model__max_depth"        : "max_depth",
            "param_model__min_samples_leaf" : "min_samples_leaf",
            "param_model__max_features"     : "max_features",
        })
        .reset_index(drop=True)
    )
    cv_df.index += 1   # rank from 1

    # ── Step 5: verbose report ────────────────────────────────────
    if verbose:
        sep = "─" * 62
        print(f"\n{sep}")
        print(f"  Tuning Results")
        print(f"{sep}")
        print(f"  Baseline CV R²  : {baseline_cv:.4f} ± {baseline_std:.4f}")
        print(f"  Tuned    CV R²  : {best_cv:.4f}")
        improvement = best_cv - baseline_cv
        arrow = "↑" if improvement > 0 else "↓"
        print(f"  Improvement     : {arrow} {abs(improvement):.4f}")
        print(f"{sep}")
        print(f"\n  Best parameters:")
        for k, v in best_params.items():
            clean_k = k.replace("model__", "")
            print(f"    {clean_k:<25} : {v}")

        print(f"\n{sep}")
        print(f"  Final Test Set Performance")
        print(f"{sep}")
        print(f"  R²              : {test_metrics['r2']:.4f}")
        print(f"  MAE             : {test_metrics['mae']:.2f}")
        print(f"  RMSE            : {test_metrics['rmse']:.2f}")
        print(f"  MAPE            : {test_metrics['mape']:.2f}%")

        # gap check
        gap = abs(best_cv - test_metrics["r2"])
        print(f"\n  CV R²  vs Test R² gap : {gap:.4f}  "
              + ("✅ Good" if gap < 0.05 else "⚠️  Possible overfitting"))
        print(f"{sep}\n")

        print(f"  Top 10 parameter combinations:\n")
        print(cv_df.head(10).round(4).to_string())

        # ── Plot ──────────────────────────────────────────────────
        _plot_tuning_results(cv_df, baseline_cv, best_cv,
                             test_metrics["r2"], n_iter)

    return {
        "best_model"   : best_model,
        "best_params"  : best_params,
        "best_cv_score": best_cv,
        "baseline_cv"  : baseline_cv,
        "test_metrics" : test_metrics,
        "cv_results_df": cv_df,
        "search"       : search,
    }


def plot_feature_importance(
    best_model,
    X_train: pd.DataFrame,
    top_n: int = 25,
    figsize: tuple = (14, 10),
    colormap: str = "RdYlGn",
    save: bool = False,
    save_dir: str = ".",
    plot_plotly: bool = True,
) -> pd.DataFrame:
    """
    Extract and visualise feature importances from a fitted Random Forest
    pipeline using three complementary views:

    View 1 — Horizontal bar chart (matplotlib)
        Top N features ranked by mean decrease in impurity (MDI).
        Color encodes importance magnitude from low (red) to high (green).
        Error bars show standard deviation across trees — wide bars mean
        the importance estimate is noisy.

    View 2 — Cumulative importance curve
        Shows how many features are needed to explain X% of the total
        importance. A steep early rise means a few features dominate.
        A flat curve means importance is spread across many features.
        The 80% and 95% thresholds are marked for reference.

    View 3 — Interactive Plotly bar chart
        Hover to see exact importance values. Saved as HTML.
        Color encodes importance from red (low) to green (high).

    What is Mean Decrease in Impurity (MDI)?
    -----------------------------------------
    At each split in each tree, the chosen feature reduces impurity
    (variance for regression) by some amount. MDI for feature j is:

        MDI(j) = (1/n_trees) * sum_t sum_nodes_using_j
                     [ p(node) * impurity_decrease(node) ]

    where p(node) = fraction of samples reaching that node.

    MDI is fast to compute (free from the fitted forest) but can be
    biased toward high-cardinality features. For unbiased importance
    use SHAP (which we also built separately).

    Parameters
    ----------
    best_model : fitted sklearn Pipeline
        Must contain a "model" step with a fitted RandomForestRegressor.
        Typically the output of ``tune_random_forest()["best_model"]``.
    X_train : pd.DataFrame
        Training features — used to recover column names.
        Must have the same columns as when the model was trained.
    top_n : int, optional
        Number of top features to show. Default is 25.
    figsize : tuple, optional
        Figure size in inches for the matplotlib plot. Default (14, 10).
    colormap : str, optional
        Matplotlib colormap for bar colors. Default "RdYlGn".
        Others: "Blues", "viridis", "coolwarm", "plasma".
    save : bool, optional
        Save matplotlib figure as PNG and Plotly as HTML. Default False.
    save_dir : str, optional
        Directory for saved files. Default is current directory.
    plot_plotly : bool, optional
        Also produce an interactive Plotly chart. Default is True.

    Returns
    -------
    importance_df : pd.DataFrame
        Full importance table sorted descending. Columns:
            feature     : str   — feature name
            importance  : float — mean MDI importance
            std         : float — std of importance across trees
            rank        : int   — rank (1 = most important)
            cumulative  : float — cumulative importance up to this feature
            pct_of_max  : float — importance as % of top feature

    Raises
    ------
    ValueError
        If "model" step not found in pipeline or model has no
        ``feature_importances_`` attribute.

    Examples
    --------
    >>> imp_df = plot_feature_importance(best_model, X_train)

    >>> # top 30 features + save outputs
    >>> imp_df = plot_feature_importance(
    ...     best_model, X_train,
    ...     top_n    = 30,
    ...     save     = True,
    ...     save_dir = "outputs"
    ... )

    >>> # access results after plotting
    >>> imp_df.head(10)
    >>> top5 = imp_df["feature"].head(5).tolist()
    """

    # ── Extract RF model from pipeline ───────────────────────────
    if hasattr(best_model, "named_steps"):
        if "model" not in best_model.named_steps:
            raise ValueError(
                "Pipeline must contain a step named 'model'.\n"
                f"Found steps: {list(best_model.named_steps.keys())}"
            )
        rf = best_model.named_steps["model"]
    else:
        rf = best_model

    if not hasattr(rf, "feature_importances_"):
        raise ValueError(
            "Model does not have feature_importances_. "
            "Ensure it is a fitted RandomForestRegressor."
        )

    # ── Match feature names to importances ───────────────────────
    importances = rf.feature_importances_
    stds        = np.std(
        [tree.feature_importances_ for tree in rf.estimators_], axis=0
    )

    feat_names = X_train.columns.tolist()
    min_len    = min(len(importances), len(feat_names))
    importances = importances[:min_len]
    stds        = stds[:min_len]
    feat_names  = feat_names[:min_len]

    # ── Build full importance DataFrame ──────────────────────────
    importance_df = (
        pd.DataFrame({
            "feature"   : feat_names,
            "importance": importances,
            "std"       : stds,
        })
        .sort_values("importance", ascending=False)
        .reset_index(drop=True)
    )
    importance_df["rank"]       = importance_df.index + 1
    importance_df["cumulative"] = importance_df["importance"].cumsum().round(4)
    importance_df["pct_of_max"] = (
        importance_df["importance"] / importance_df["importance"].max() * 100
    ).round(2)

    top_df = importance_df.head(top_n).copy()

    # ── Console report ────────────────────────────────────────────
    n_trees      = len(rf.estimators_)
    n80          = (importance_df["cumulative"] <= 0.80).sum() + 1
    n95          = (importance_df["cumulative"] <= 0.95).sum() + 1

    sep = "─" * 62
    print(f"\n{sep}")
    print(f"  Feature Importance Report")
    print(f"{sep}")
    print(f"  Total features       : {len(feat_names)}")
    print(f"  Trees in forest      : {n_trees}")
    print(f"  Features shown       : top {top_n}")
    print(f"  Features for 80% imp : {n80}")
    print(f"  Features for 95% imp : {n95}")
    print(f"{sep}")
    print(f"\n  {'Rank':>4}  {'Feature':<40} {'Importance':>11}  {'± Std':>8}  {'% of Max':>9}")
    print(f"  {'─'*4}  {'─'*40} {'─'*11}  {'─'*8}  {'─'*9}")
    for _, row in top_df.head(15).iterrows():
        bar = "█" * int(row["pct_of_max"] / 5)
        print(
            f"  {int(row['rank']):>4}  "
            f"{row['feature']:<40} "
            f"{row['importance']:>11.4f}  "
            f"{row['std']:>8.4f}  "
            f"{row['pct_of_max']:>8.1f}%  {bar}"
        )
    print(f"\n{sep}\n")

    # ── Matplotlib figure — 2 panels ──────────────────────────────
    fig = plt.figure(figsize=figsize)
    gs  = gridspec.GridSpec(1, 2, width_ratios=[3, 1.2], wspace=0.35)

    ax_bar  = fig.add_subplot(gs[0])
    ax_cum  = fig.add_subplot(gs[1])

    # ── Panel 1: horizontal bar chart ────────────────────────────
    norm   = Normalize(
        vmin=top_df["importance"].min(),
        vmax=top_df["importance"].max()
    )
    cmap   = cm.get_cmap(colormap)
    colors = [cmap(norm(v)) for v in top_df["importance"].values]

    # plot bars in ascending order so most important is at the top
    plot_df = top_df.iloc[::-1]
    plot_colors = colors[::-1]

    bars = ax_bar.barh(
        plot_df["feature"],
        plot_df["importance"],
        xerr   = plot_df["std"],
        color  = plot_colors,
        edgecolor = "white",
        linewidth = 0.5,
        capsize   = 3,
        ecolor    = "#aaaaaa",
        height    = 0.7,
    )

    # value labels on bars
    for bar, val, std in zip(
        bars, plot_df["importance"], plot_df["std"]
    ):
        ax_bar.text(
            bar.get_width() + std + 0.001,
            bar.get_y() + bar.get_height() / 2,
            f"{val:.4f}",
            va="center", ha="left",
            fontsize=8, color="#333"
        )

    ax_bar.set_xlabel(
        "Mean Decrease in Impurity (MDI)",
        fontsize=11
    )
    ax_bar.set_title(
        f"Top {top_n} Feature Importances\n"
        f"(error bars = std across {n_trees} trees)",
        fontsize=12, fontweight="bold", pad=12
    )
    ax_bar.grid(True, axis="x", alpha=0.3)
    ax_bar.set_facecolor("#fafafa")
    ax_bar.tick_params(axis="y", labelsize=9)

    # colorbar
    sm = cm.ScalarMappable(cmap=cmap, norm=norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, ax=ax_bar, pad=0.01, shrink=0.6)
    cbar.set_label("Importance", fontsize=9)
    cbar.ax.tick_params(labelsize=8)

    # ── Panel 2: cumulative importance curve ──────────────────────
    cum_vals = importance_df["cumulative"].values
    n_feats  = np.arange(1, len(cum_vals) + 1)

    ax_cum.plot(
        cum_vals, n_feats,
        color="#4c8eda", linewidth=2
    )
    ax_cum.fill_betweenx(
        n_feats, 0, cum_vals,
        alpha=0.12, color="#4c8eda"
    )

    # 80% and 95% threshold lines
    for thresh, color, label in [
        (0.80, "#f0a500", "80%"),
        (0.95, "#e05c5c", "95%"),
    ]:
        n_at = (importance_df["cumulative"] <= thresh).sum() + 1
        ax_cum.axvline(thresh, color=color,
                       linestyle="--", linewidth=1.2, alpha=0.8)
        ax_cum.axhline(n_at,   color=color,
                       linestyle=":",  linewidth=1.0, alpha=0.8)
        ax_cum.annotate(
            f"{label}\n({n_at} feats)",
            xy=(thresh, n_at),
            xytext=(thresh - 0.25, n_at + len(cum_vals) * 0.05),
            fontsize=8, color=color,
            arrowprops=dict(arrowstyle="->", color=color, lw=1)
        )

    ax_cum.set_xlabel("Cumulative Importance", fontsize=10)
    ax_cum.set_ylabel("Number of Features",    fontsize=10)
    ax_cum.set_title(
        "Cumulative Importance\nCurve",
        fontsize=11, fontweight="bold", pad=12
    )
    ax_cum.set_xlim(0, 1.02)
    ax_cum.set_ylim(0, len(cum_vals) + 1)
    ax_cum.grid(True, alpha=0.3)
    ax_cum.set_facecolor("#fafafa")

    fig.suptitle(
        "Random Forest — Feature Importance Analysis",
        fontsize=14, fontweight="bold", y=1.01
    )

    plt.tight_layout()

    if save:
        import os
        os.makedirs(save_dir, exist_ok=True)
        path = os.path.join(save_dir, "feature_importance.png")
        plt.savefig(path, dpi=150, bbox_inches="tight")
        print(f"✅ Saved → {path}")

    plt.show()

    # ── Plotly interactive chart ──────────────────────────────────
    if plot_plotly:
        fig_pl = go.Figure(go.Bar(
            x            = top_df["importance"],
            y            = top_df["feature"],
            orientation  = "h",
            error_x      = dict(
                type     = "data",
                array    = top_df["std"],
                visible  = True,
                color    = "#aaaaaa",
            ),
            marker       = dict(
                color    = top_df["importance"],
                colorscale = colormap,
                showscale  = True,
                colorbar   = dict(title="Importance", thickness=14),
            ),
            hovertemplate = (
                "<b>%{y}</b><br>"
                "Importance : %{x:.4f}<br>"
                "Rank       : %{customdata[0]}<br>"
                "% of Max   : %{customdata[1]:.1f}%<br>"
                "Cumulative : %{customdata[2]:.4f}"
                "<extra></extra>"
            ),
            customdata = top_df[["rank", "pct_of_max", "cumulative"]].values,
        ))

        fig_pl.update_layout(
            title = dict(
                text = (
                    f"Top {top_n} Feature Importances — Random Forest<br>"
                    f"<sup>MDI importance  |  "
                    f"{n_trees} trees  |  "
                    f"{n80} features explain 80% of importance</sup>"
                ),
                font = dict(size=15)
            ),
            xaxis_title  = "Mean Decrease in Impurity (MDI)",
            yaxis        = dict(autorange="reversed"),
            height       = max(500, top_n * 22),
            width        = 1000,
            plot_bgcolor = "white",
            paper_bgcolor= "white",
            margin       = dict(l=250, r=30),
        )

        if save:
            path_html = os.path.join(save_dir, "feature_importance.html")
            with open(path_html, "w", encoding="utf-8") as f:
                f.write(fig_pl.to_html())
            print(f"✅ Saved → {path_html}")

        fig_pl.show(renderer="browser")

    return importance_df


def compute_metrics(
    y_true,
    y_pred,
    residuals=None,
    n_features: int | None = None,
    eps: float = 1e-9,
) -> dict:
    """
    Compute a comprehensive set of regression metrics + residual diagnostics.

    Parameters
    ----------
    y_true : array-like
        Ground-truth target values.
    y_pred : array-like
        Model predictions.
    residuals : array-like, optional
        Pre-computed residuals (y_true - y_pred). If None, computed internally.
    n_features : int, optional
        Number of predictors. If provided, adjusted R² is reported.
    eps : float
        Small constant to avoid division by zero in relative-error metrics.

    Returns
    -------
    dict
        Grouped metrics:

        Accuracy
            r2, adj_r2, explained_variance
        Error magnitude
            mae, rmse, max_error
        Relative error (%)
            mape, smape, median_ape, p90_ape
        Residual diagnostics
            mean_res, std_res, skew, kurtosis,
            shapiro_p     (normality test, p<0.05 → non-normal)
            heterosced_p  (Spearman |res| vs y_pred, p<0.05 → heteroscedastic)
        Sample
            n_samples
    """
    # ── coerce to 1-D float arrays ────────────────────────────────
    y_true = np.asarray(y_true, dtype=float).ravel()
    y_pred = np.asarray(y_pred, dtype=float).ravel()

    if y_true.shape != y_pred.shape:
        raise ValueError(
            f"Shape mismatch: y_true {y_true.shape} vs y_pred {y_pred.shape}"
        )
    if y_true.size == 0:
        raise ValueError("Empty input arrays.")

    if residuals is None:
        residuals = y_true - y_pred
    residuals = np.asarray(residuals, dtype=float).ravel()

    n = y_true.size

    # ── relative-error metrics (robust to zeros) ──────────────────
    denom_mape  = np.maximum(np.abs(y_true), eps)
    ape         = np.abs(residuals) / denom_mape                     # absolute % err
    smape_arr   = 2 * np.abs(residuals) / np.maximum(
                      np.abs(y_true) + np.abs(y_pred), eps)          # symmetric MAPE

    # ── adjusted R² (only if n_features known and valid) ──────────
    r2 = r2_score(y_true, y_pred)
    if n_features is not None and n - n_features - 1 > 0:
        adj_r2 = 1 - (1 - r2) * (n - 1) / (n - n_features - 1)
    else:
        adj_r2 = None

    # ── residual diagnostics ──────────────────────────────────────
    # Shapiro-Wilk: tests if residuals are normally distributed
    if 3 <= n <= 5000:
        shapiro_p = float(stats.shapiro(residuals).pvalue)
    else:
        shapiro_p = None  # Shapiro is unreliable / undefined outside this range

    # Heteroscedasticity proxy: |residual| should NOT correlate with prediction
    if n >= 3 and np.std(y_pred) > 0:
        heterosced_p = float(
            stats.spearmanr(np.abs(residuals), y_pred).pvalue
        )
    else:
        heterosced_p = None

    return {
        # accuracy
        "r2"                : round(r2, 4),
        "adj_r2"            : None if adj_r2 is None else round(adj_r2, 4),
        "explained_variance": round(explained_variance_score(y_true, y_pred), 4),

        # error magnitude (in target units)
        "mae"               : round(mean_absolute_error(y_true, y_pred), 4),
        "rmse"              : round(float(np.sqrt(mean_squared_error(y_true, y_pred))), 4),
        "max_error"         : round(float(max_error(y_true, y_pred)), 4),

        # relative error (%)
        "mape_pct"          : round(mean_absolute_percentage_error(y_true, y_pred) * 100, 4),
        "smape_pct"         : round(float(np.mean(smape_arr)) * 100, 4),
        "median_ape_pct"    : round(float(np.median(ape)) * 100, 4),
        "p90_ape_pct"       : round(float(np.percentile(ape, 90)) * 100, 4),

        # residual diagnostics
        "mean_res"          : round(float(np.mean(residuals)), 4),
        "std_res"           : round(float(np.std(residuals, ddof=1)), 4),
        "skew"              : round(float(stats.skew(residuals)), 4),
        "kurtosis"          : round(float(stats.kurtosis(residuals)), 4),  # excess
        "shapiro_p"         : None if shapiro_p   is None else round(shapiro_p,   4),
        "heterosced_p"      : None if heterosced_p is None else round(heterosced_p, 4),

        # bookkeeping
        "n_samples"         : int(n),
    }


def plot_residuals(
    model,
    X_train: pd.DataFrame,
    y_train: pd.Series,
    X_test: pd.DataFrame,
    y_test: pd.Series,
    model_name: str = "Random Forest",
    train_color: str = "#4878d0",
    test_color: str = "#a6d854",
    figsize: tuple = (20, 16),
    save: bool = False,
    save_dir: str = ".",
    file_name_to_save: str = "residuals_diagnostic.png"
) -> dict:
    """
    Comprehensive residual analysis using compute_metrics() for all
    metric calculations.

    Panels
    ------
    1. Residuals vs Predicted   — bias and heteroscedasticity
    2. Q-Q Plot                 — normality of standardised residuals
    3. Residual Distribution    — shape, centring, KDE
    4. Predicted vs Actual      — overall accuracy
    5. Residuals vs Row Order   — temporal / ordering bias
    6. Scale-Location           — homoscedasticity check

    Parameters
    ----------
    model      : fitted sklearn Pipeline
    X_train    : pd.DataFrame — training features
    y_train    : pd.Series    — training target
    X_test     : pd.DataFrame — test features
    y_test     : pd.Series    — test target
    model_name : str          — shown in plot titles
    train_color: str          — color for train points
    test_color : str          — color for test points
    figsize    : tuple        — figure size in inches
    save       : bool         — save PNG to save_dir
    save_dir   : str          — output directory

    Returns
    -------
    dict with keys:
        train_metrics : dict from compute_metrics()
        test_metrics  : dict from compute_metrics()
        residuals     : {"train": array, "test": array}
        predictions   : {"train": array, "test": array}

    Examples
    --------
    >>> results = plot_residuals(final_model, X_train, y_train,
    ...                          X_test, y_test)
    >>> results["test_metrics"]["r2"]
    >>> results["test_metrics"]["rmse"]
    >>> results["test_metrics"]["shapiro_p"]
    """

    # ── Predictions & residuals ───────────────────────────────────
    y_pred_train = model.predict(X_train)
    y_pred_test  = model.predict(X_test)

    res_train = np.array(y_train) - y_pred_train
    res_test  = np.array(y_test)  - y_pred_test

    # ── Metrics via compute_metrics() ─────────────────────────────
    train_metrics = compute_metrics(y_train, y_pred_train, residuals=res_train)
    test_metrics  = compute_metrics(y_test,  y_pred_test,  residuals=res_test)

    # ── Standardise residuals ─────────────────────────────────────
    def standardize(x):
        return (x - np.mean(x)) / (np.std(x) + 1e-9)

    res_train_std = standardize(res_train)
    res_test_std  = standardize(res_test)

    # ── Console report ────────────────────────────────────────────
    sep = "─" * 65
    print(f"\n{sep}")
    print(f"  Residual Analysis Report — {model_name}")
    print(f"{sep}")
    print(f"  {'Metric':<22} {'Train':>12} {'Test':>12}  Note")
    print(f"  {'─'*22} {'─'*12} {'─'*12}  {'─'*20}")

    report_rows = [
        ("r2",            "R²"),
        ("adj_r2",        "Adj R²"),
        ("explained_variance", "Expl. variance"),
        ("mae",           "MAE"),
        ("rmse",          "RMSE"),
        ("max_error",     "Max error"),
        ("mape_pct",      "MAPE %"),
        ("smape_pct",     "sMAPE %"),
        ("median_ape_pct","Median APE %"),
        ("p90_ape_pct",   "P90 APE %"),
        ("mean_res",      "Mean residual"),
        ("std_res",       "Std residual"),
        ("skew",          "Skewness"),
        ("kurtosis",      "Kurtosis"),
        ("shapiro_p",     "Shapiro-Wilk p"),
        ("heterosced_p",  "Heterosced. p"),
    ]

    for key, label in report_rows:
        tr_val = train_metrics.get(key)
        te_val = test_metrics.get(key)

        tr_str = "N/A" if tr_val is None else str(tr_val)
        te_str = "N/A" if te_val is None else str(te_val)

        # interpretation flags
        flag = ""
        if key == "r2":
            flag = "✅" if te_val and te_val >= 0.8 else ("⚠️" if te_val and te_val >= 0.6 else "❌")
        elif key == "mean_res":
            flag = "✅" if te_val and abs(te_val) < (test_metrics["std_res"] * 0.1) else "⚠️ bias"
        elif key == "skew":
            flag = "✅" if te_val and abs(te_val) < 0.5 else "⚠️ skewed"
        elif key == "shapiro_p":
            if te_val is not None:
                flag = "✅ normal" if te_val > 0.05 else "⚠️ non-normal"
        elif key == "heterosced_p":
            if te_val is not None:
                flag = "✅ homosced" if te_val > 0.05 else "⚠️ heterosced"

        print(f"  {label:<22} {tr_str:>12} {te_str:>12}  {flag}")

    r2_gap = abs(train_metrics["r2"] - test_metrics["r2"])
    print(f"\n  R² gap (train - test) : {r2_gap:.4f}  "
          + ("✅ Good" if r2_gap < 0.05 else "⚠️  Possible overfitting"))
    print(f"  n_samples train       : {train_metrics['n_samples']:,}")
    print(f"  n_samples test        : {test_metrics['n_samples']:,}")
    print(f"{sep}\n")

    # ── Figure layout ─────────────────────────────────────────────
    fig = plt.figure(figsize=figsize)
    gs  = gridspec.GridSpec(
        3, 2,
        hspace=0.42, wspace=0.28,
        height_ratios=[1, 1, 1]
    )
    ax1 = fig.add_subplot(gs[0, 0])
    ax2 = fig.add_subplot(gs[0, 1])
    ax3 = fig.add_subplot(gs[1, 0])
    ax4 = fig.add_subplot(gs[1, 1])
    ax5 = fig.add_subplot(gs[2, 0])
    ax6 = fig.add_subplot(gs[2, 1])

    legend_handles = [
        Line2D([0], [0], marker="o", color="w",
               markerfacecolor=train_color, markersize=9,
               label=f"Train  R²={train_metrics['r2']:.3f}  "
                     f"n={train_metrics['n_samples']:,}"),
        Line2D([0], [0], marker="o", color="w",
               markerfacecolor=test_color, markersize=9,
               label=f"Test   R²={test_metrics['r2']:.3f}  "
                     f"n={test_metrics['n_samples']:,}"),
    ]

    kw_train = dict(color=train_color, alpha=0.55, s=30, edgecolors="none")
    kw_test  = dict(color=test_color,  alpha=0.80, s=35, edgecolors="none")

    # ── Panel 1: Residuals vs Predicted ──────────────────────────
    ax1.scatter(y_pred_train, res_train, **kw_train)
    ax1.scatter(y_pred_test,  res_test,  **kw_test)
    ax1.axhline(0, color="black", linewidth=1.5, linestyle="--")
    try:
        from statsmodels.nonparametric.smoothers_lowess import lowess
        smooth = lowess(res_test, y_pred_test, frac=0.4)
        ax1.plot(smooth[:, 0], smooth[:, 1],
                 color="#e05c5c", linewidth=2, label="LOWESS")
    except ImportError:
        pass
    ax1.set_xlabel("Predicted Value", fontsize=10)
    ax1.set_ylabel("Residuals",       fontsize=10)
    ax1.set_title("Residuals vs Predicted", fontsize=11, fontweight="bold")
    ax1.legend(handles=legend_handles, fontsize=8, framealpha=0.8)
    ax1.grid(True, alpha=0.25)
    ax1.set_facecolor("#fafafa")

    # ── Panel 2: Q-Q plot ─────────────────────────────────────────
    (osm_tr, osr_tr), (sl_tr, ic_tr, _) = probplot(res_train_std, dist="norm")
    (osm_te, osr_te), (sl_te, ic_te, _) = probplot(res_test_std,  dist="norm")

    ax2.scatter(osm_tr, osr_tr, **kw_train, label="Train")
    ax2.scatter(osm_te, osr_te, **kw_test,  label="Test")
    x_line = np.linspace(osm_tr.min(), osm_tr.max(), 200)
    ax2.plot(x_line, sl_tr * x_line + ic_tr,
             color="black", linewidth=1.5, label="Reference")

    shapiro_note = ""
    if test_metrics["shapiro_p"] is not None:
        shapiro_note = (
            f"  Shapiro p={test_metrics['shapiro_p']:.3f} "
            + ("✅" if test_metrics["shapiro_p"] > 0.05 else "⚠️")
        )

    ax2.set_xlabel("Theoretical quantiles", fontsize=10)
    ax2.set_ylabel("Observed quantiles",    fontsize=10)
    ax2.set_title(
        f"Q-Q Plot (standardised residuals)\n{shapiro_note}",
        fontsize=11, fontweight="bold"
    )
    ax2.yaxis.set_label_position("right")
    ax2.yaxis.tick_right()
    ax2.legend(fontsize=8)
    ax2.grid(True, alpha=0.25)
    ax2.set_facecolor("#fafafa")

    # ── Panel 3: Residual Distribution ───────────────────────────
    bins = min(30, max(10, len(res_test) // 3))
    for res, color, label in [
        (res_train, train_color, "Train"),
        (res_test,  test_color,  "Test"),
    ]:
        ax3.hist(res, bins=bins, density=True,
                 color=color, alpha=0.4, label=label, edgecolor="none")
        kde = stats.gaussian_kde(res)
        xr  = np.linspace(res.min(), res.max(), 300)
        ax3.plot(xr, kde(xr), color=color, linewidth=2)

    ax3.axvline(0, color="black", linewidth=1.5, linestyle="--")
    ax3.axvline(
        test_metrics["mean_res"], color=test_color,
        linewidth=1.2, linestyle=":",
        label=f"Test mean={test_metrics['mean_res']:.1f}"
    )
    ax3.set_xlabel("Residual", fontsize=10)
    ax3.set_ylabel("Density",  fontsize=10)
    ax3.set_title(
        f"Residual Distribution\n"
        f"Test skew={test_metrics['skew']:.3f}  "
        f"kurt={test_metrics['kurtosis']:.3f}",
        fontsize=11, fontweight="bold"
    )
    ax3.legend(fontsize=8)
    ax3.grid(True, alpha=0.25)
    ax3.set_facecolor("#fafafa")

    # ── Panel 4: Predicted vs Actual ──────────────────────────────
    ax4.scatter(y_train, y_pred_train, **kw_train)
    ax4.scatter(y_test,  y_pred_test,  **kw_test)
    all_vals = np.concatenate([
        np.array(y_train), np.array(y_test),
        y_pred_train, y_pred_test
    ])
    mn, mx = all_vals.min(), all_vals.max()
    ax4.plot([mn, mx], [mn, mx], color="black",
             linewidth=1.5, linestyle="--", label="Perfect fit")
    ax4.set_xlabel("Actual Value",    fontsize=10)
    ax4.set_ylabel("Predicted Value", fontsize=10)
    ax4.set_title(
        f"Predicted vs Actual\n"
        f"Test R²={test_metrics['r2']:.4f}  "
        f"MAE={test_metrics['mae']:.2f}  "
        f"RMSE={test_metrics['rmse']:.2f}",
        fontsize=11, fontweight="bold"
    )
    ax4.legend(handles=legend_handles + [
        Line2D([0], [0], color="black", linestyle="--", label="Perfect fit")
    ], fontsize=8, framealpha=0.8)
    ax4.grid(True, alpha=0.25)
    ax4.set_facecolor("#fafafa")

    # ── Panel 5: Residuals vs Row Order ──────────────────────────
    ax5.scatter(range(len(res_train)), res_train, **kw_train, label="Train")
    ax5.scatter(
        range(len(res_train), len(res_train) + len(res_test)),
        res_test, **kw_test, label="Test"
    )
    ax5.axhline(0, color="black", linewidth=1.5, linestyle="--")
    ax5.axvline(len(res_train), color="#e05c5c",
                linewidth=1, linestyle=":", alpha=0.7,
                label="Train/Test boundary")
    ax5.set_xlabel("Row index", fontsize=10)
    ax5.set_ylabel("Residuals", fontsize=10)
    ax5.set_title("Residuals vs Row Order\n(checks ordering bias)",
                  fontsize=11, fontweight="bold")
    ax5.legend(fontsize=8)
    ax5.grid(True, alpha=0.25)
    ax5.set_facecolor("#fafafa")

    # ── Panel 6: Scale-Location ───────────────────────────────────
    ax6.scatter(y_pred_train, np.sqrt(np.abs(res_train_std)), **kw_train)
    ax6.scatter(y_pred_test,  np.sqrt(np.abs(res_test_std)),  **kw_test)
    try:
        from statsmodels.nonparametric.smoothers_lowess import lowess
        smooth = lowess(np.sqrt(np.abs(res_test_std)), y_pred_test, frac=0.4)
        ax6.plot(smooth[:, 0], smooth[:, 1],
                 color="#e05c5c", linewidth=2, label="LOWESS")
    except ImportError:
        pass

    heterosced_note = ""
    if test_metrics["heterosced_p"] is not None:
        heterosced_note = (
            f"  Heterosced p={test_metrics['heterosced_p']:.3f} "
            + ("✅" if test_metrics["heterosced_p"] > 0.05 else "⚠️")
        )

    ax6.set_xlabel("Predicted Value",           fontsize=10)
    ax6.set_ylabel("√|Standardised Residuals|", fontsize=10)
    ax6.set_title(
        f"Scale-Location (homoscedasticity)\n{heterosced_note}",
        fontsize=11, fontweight="bold"
    )
    ax6.legend(handles=legend_handles, fontsize=8)
    ax6.grid(True, alpha=0.25)
    ax6.set_facecolor("#fafafa")

    # ── Super title ───────────────────────────────────────────────
    fig.suptitle(
        f"Residual Diagnostic Plots — {model_name}\n"
        f"Train R²={train_metrics['r2']:.4f}  │  "
        f"Test R²={test_metrics['r2']:.4f}  │  "
        f"Test MAE={test_metrics['mae']:.2f}  │  "
        f"Test RMSE={test_metrics['rmse']:.2f}  │  "
        f"Test MAPE={test_metrics['mape_pct']:.1f}%",
        fontsize=13, fontweight="bold", y=1.01
    )

    plt.tight_layout()

    if save:
        import os
        os.makedirs(save_dir, exist_ok=True)
        path = os.path.join(save_dir, file_name_to_save)
        plt.savefig(path, dpi=150, bbox_inches="tight")
        print(f"✅ Saved → {path}")

    plt.show()

    return {
        "train_metrics": train_metrics,
        "test_metrics" : test_metrics,
        "residuals"    : {"train": res_train, "test": res_test},
        "predictions"  : {"train": y_pred_train, "test": y_pred_test},
    }


def retrain_final_model(
    best_model,
    X_train: pd.DataFrame,
    y_train: pd.Series,
    X_test: pd.DataFrame,
    y_test: pd.Series,
    save: bool = True,
    save_dir: str = ".",
    model_name: str = "final_model.pkl",
    verbose: bool = True,
) -> Pipeline:
    """
    Retrain the best model on the full dataset (train + test combined)
    using the exact same hyperparameters found during tuning.

    Why retrain on all data?
    ------------------------
    During the ML pipeline we deliberately held out X_test to get an
    honest, unbiased estimate of model performance. Now that we trust
    the model (Test R² is acceptable), we retrain on ALL available data:

        More data → more patterns learned → slightly better generalisation
        on truly new, unseen configurations in production.

    The Test R² score we measured earlier remains valid because:
        1. We evaluated BEFORE this retraining step
        2. The test set was never used to make any tuning decisions
        3. This retraining is purely for production deployment

    What stays the same
    -------------------
        ✅ Same hyperparameters  (extracted from best_model)
        ✅ Same pipeline steps   (imputer → scaler → RF)
        ✅ Same feature columns  (X_train + X_test columns)
        ✅ Same random_state     (reproducibility)

    What changes
    ------------
        📈 Training data = X_train + X_test (all available rows)
        📈 More trees see more data → slightly lower variance

    Parameters
    ----------
    best_model : fitted sklearn Pipeline
        Output of ``tune_random_forest()["best_model"]``.
        Must contain "imputer", "scaler", "model" steps.
    X_train : pd.DataFrame — training features
    y_train : pd.Series    — training target
    X_test  : pd.DataFrame — test features
    y_test  : pd.Series    — test target
    save : bool, optional
        Save final model to disk as .pkl. Default True.
    save_dir : str, optional
        Directory for saved model. Default is current directory.
    model_name : str, optional
        Filename for saved model. Default "final_model.pkl".
    verbose : bool, optional
        Print report. Default True.

    Returns
    -------
    final_model : fitted sklearn Pipeline
        Retrained pipeline on full data. Ready for production.

    Examples
    --------
    >>> final_model = retrain_final_model(
    ...     best_model, X_train, y_train, X_test, y_test
    ... )

    >>> # predict on new door configurations
    >>> y_new = final_model.predict(X_new)

    >>> # load saved model later
    >>> import joblib
    >>> model = joblib.load("final_model.pkl")
    >>> y_new = model.predict(X_new)
    """

    # ── Step 1: extract best hyperparameters ─────────────────────
    rf          = best_model.named_steps["model"]
    best_params = rf.get_params()

    if verbose:
        sep = "─" * 60
        print(f"\n{sep}")
        print(f"  Final Model Retraining")
        print(f"{sep}")
        print(f"  Hyperparameters from tuning:")
        important_params = [
            "n_estimators", "max_depth", "min_samples_split",
            "min_samples_leaf", "max_features", "random_state"
        ]
        for p in important_params:
            print(f"    {p:<25} : {best_params.get(p, 'N/A')}")

    # ── Step 2: combine train + test ─────────────────────────────
    X_all = pd.concat([X_train, X_test], axis=0).reset_index(drop=True)
    y_all = pd.concat([y_train, y_test], axis=0).reset_index(drop=True)

    if verbose:
        print(f"\n  Data:")
        print(f"    X_train rows     : {len(X_train):,}")
        print(f"    X_test  rows     : {len(X_test):,}")
        print(f"    X_all   rows     : {len(X_all):,}  ← full dataset")
        print(f"    Features         : {X_all.shape[1]}")

    # ── Step 3: build fresh pipeline with same params ─────────────
    final_model = Pipeline([
        ("imputer", SimpleImputer(strategy="median")),
        ("scaler",  StandardScaler()),
        ("model",   RandomForestRegressor(**best_params)),
    ])

    # ── Step 4: fit on all data ───────────────────────────────────
    if verbose:
        print(f"\n  Training on full dataset ⏳ ...")

    final_model.fit(X_all, y_all)

    if verbose:
        print(f"  Training done ✅")

    # ── Step 5: sanity check ──────────────────────────────────────
    # predict on train and test separately to compare
    y_pred_train = final_model.predict(X_train)
    y_pred_test  = final_model.predict(X_test)

    train_r2   = r2_score(y_train, y_pred_train)
    test_r2    = r2_score(y_test,  y_pred_test)
    train_mae  = mean_absolute_error(y_train, y_pred_train)
    test_mae   = mean_absolute_error(y_test,  y_pred_test)
    train_rmse = np.sqrt(mean_squared_error(y_train, y_pred_train))
    test_rmse  = np.sqrt(mean_squared_error(y_test,  y_pred_test))

    if verbose:
        print(f"\n{sep}")
        print(f"  Sanity Check — Final Model vs Original Splits")
        print(f"{sep}")
        print(f"  {'Metric':<10} {'Train':>12} {'Test':>12}")
        print(f"  {'─'*10} {'─'*12} {'─'*12}")
        print(f"  {'R²':<10} {train_r2:>12.4f} {test_r2:>12.4f}")
        print(f"  {'MAE':<10} {train_mae:>12.2f} {test_mae:>12.2f}")
        print(f"  {'RMSE':<10} {train_rmse:>12.2f} {test_rmse:>12.2f}")
        print(f"{sep}")
        print(f"\n  ℹ️  Note: these scores use data the model was trained")
        print(f"     on — they are NOT unbiased estimates.")
        print(f"     The unbiased Test R² from BEFORE retraining")
        print(f"     remains the honest performance estimate.")
        print(f"{sep}")

    # ── Step 6: save to disk ──────────────────────────────────────
    if save:
        os.makedirs(save_dir, exist_ok=True)
        save_path = os.path.join(save_dir, model_name)
        joblib.dump(final_model, save_path)
        data_all = X_all.copy()
        data_all["KgCO2EQ"] = y_all
        data_all.to_csv(os.path.join(save_dir, "final_training_data.csv"), index=False)

        if verbose:
            print(f"\n  ✅ Final model saved → {save_path}")
            print(f"     Load later with:")
            print(f"       import joblib")
            print(f"       model = joblib.load('{save_path}')")
            print(f"       y_new = model.predict(X_new)\n")
            print(f"  ✅ Final training data saved → {os.path.join(save_dir, 'final_training_data.csv')}")
            print(f"     Load later with:")
            print(f"       import pandas as pd")
            print(f"       data = pd.read_csv('{os.path.join(save_dir, 'final_training_data.csv')}')\n")


    return final_model, data_all


def preprocess_new_data(
    df_new: pd.DataFrame,
    train_columns: list,
    numeric_threshold: float = 1.0,
    encoding_method: str = "auto",
    target_col: str = "KgCO2EQ",
    verbose: bool = True,
) -> pd.DataFrame:
    """
    Apply the exact same preprocessing pipeline used during training
    to new raw data before prediction.

    Steps applied (must match training exactly)
    -------------------------------------------
    1. detect_column_types     — convert string numbers to float
    2. encode_categorical_cols — same method used during training
    3. align columns           — add missing cols as 0, drop extras
                                 ensures exact same shape as X_train

    Note: imputation and scaling are INSIDE final_model.predict()
    so they are applied automatically — do not apply them here.

    Parameters
    ----------
    df_new         : raw new DataFrame (same format as original df)
    train_columns  : X_train.columns.tolist() — the exact columns
                     the model was trained on
    numeric_threshold : same value used in detect_column_types()
    encoding_method   : same method used in encode_categorical_columns()
    target_col        : column to drop if present in new data
    verbose           : print report

    Returns
    -------
    X_new : pd.DataFrame — ready for final_model.predict()
    """

    if verbose:
        print(f"{'─'*55}")
        print(f"  Preprocessing New Data")
        print(f"{'─'*55}")
        print(f"  Input shape      : {df_new.shape}")

    df = df_new.copy()

    # ── Step 1: detect and convert types ─────────────────────────
    df = detect_column_types(df, numeric_threshold=numeric_threshold,
                             verbose=False)

    # ── Step 2: drop target if present in new data ───────────────
    if target_col in df.columns:
        df = df.drop(columns=[target_col])
        if verbose:
            print(f"  Target dropped   : {target_col}")

    # ── Step 3: encode categoricals — same method as training ─────
    df = encode_categorical_columns(
        df,
        method     = encoding_method,
        target_col = None,
        verbose    = False,
    )

    # ── Step 4: align columns to match training exactly ──────────
    # add any columns that are in train but missing in new data
    missing_cols = set(train_columns) - set(df.columns)
    for col in missing_cols:
        df[col] = 0     # fill with 0 — imputer will handle NaN later

    # drop any extra columns not seen during training
    extra_cols = set(df.columns) - set(train_columns)
    if extra_cols:
        df = df.drop(columns=list(extra_cols))

    # reorder columns to exactly match training order
    df = df[train_columns]

    if verbose:
        print(f"  Missing cols added  : {len(missing_cols)}")
        print(f"  Extra  cols dropped : {len(extra_cols)}")
        print(f"  Output shape        : {df.shape}")
        print(f"  ✅ Ready for prediction")
        print(f"{'─'*55}")

    return df


def safe_kde(values: np.ndarray, x_range: np.ndarray) -> np.ndarray:
    """
    Compute KDE safely — handles zero-variance and constant columns.

    Falls back to a histogram density estimate if gaussian_kde fails
    due to a singular covariance matrix (all values identical).

    Parameters
    ----------
    values  : array of data points
    x_range : array of x positions to evaluate the density on

    Returns
    -------
    density : np.ndarray same shape as x_range
    """
    vals = np.array(values, dtype=float)
    vals = vals[np.isfinite(vals)]

    if len(vals) < 2:
        return np.zeros_like(x_range, dtype=float)

    if np.std(vals) < 1e-10:
        # constant column — spike at the single value
        density      = np.zeros_like(x_range, dtype=float)
        idx          = np.argmin(np.abs(x_range - vals[0]))
        density[idx] = 1.0
        return density

    try:
        return stats.gaussian_kde(vals)(x_range)
    except Exception:
        counts, bin_edges = np.histogram(vals, bins=max(4, len(x_range)//4),
                                         density=True)
        bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
        return np.interp(x_range, bin_centers, counts, left=0, right=0)


def compare_data_distributions(
    X_train: pd.DataFrame,
    y_train: pd.Series,
    X_new: pd.DataFrame,
    y_new: pd.Series,
    train_columns: list,
    target_col: str = "KgCO2EQ",
    top_n_features: int = 20,
    train_color: str = "#4878d0",
    new_color: str = "#e05c5c",
    save: bool = False,
    save_dir: str = ".",
) -> dict:
    """
    Compare feature and target distributions between training data
    and new incoming data BEFORE making predictions.

    Uses the Kolmogorov-Smirnov (KS) test to detect drift per feature:

        KS statistic = max |F_train(x) - F_new(x)|

        KS ≈ 0   → same distribution     ✅
        KS ≈ 1   → very different         ⚠️
        p < 0.05 → statistically drifted  ⚠️

    Panels
    ------
    1. KS drift bar chart      — all features ranked by drift score
    2. Target distribution     — train vs new CO₂ values
    3. Mean comparison scatter — feature means train vs new
    4. Top 6 drifted features  — overlaid KDE curves

    Parameters
    ----------
    X_train        : training features (reference distribution)
    y_train        : training target values
    X_new          : new data features (preprocessed)
    y_new          : new data target values
    train_columns  : X_train.columns.tolist()
    target_col     : name of target column
    top_n_features : features to show in bar chart
    train_color    : color for train data
    new_color      : color for new data
    save           : save PNG to save_dir
    save_dir       : output directory

    Returns
    -------
    dict with keys:
        drift_df      : pd.DataFrame — KS stats per feature sorted desc
        drifted_cols  : list         — features with p < 0.05
        stable_cols   : list         — features with no drift
        target_ks     : float        — KS stat on target column
        target_pvalue : float        — p-value on target KS test

    Examples
    --------
    >>> drift = compare_data_distributions(
    ...     X_train, y_train, X_new_proc, y_new,
    ...     train_columns = train_columns,
    ...     save          = True,
    ...     save_dir      = "outputs",
    ... )
    >>> drift["drift_df"].head(10)
    >>> drift["drifted_cols"]
    """

    # ── Align to common columns ───────────────────────────────────
    common_cols = [c for c in train_columns if c in X_new.columns]
    X_tr = X_train[common_cols].copy()
    X_nw = X_new[common_cols].copy()
    y_tr = pd.to_numeric(y_train, errors="coerce").dropna()
    y_nw = pd.to_numeric(y_new,   errors="coerce").dropna()

    # ── KS test per feature ───────────────────────────────────────
    records = []
    for col in common_cols:
        tr_vals = pd.to_numeric(X_tr[col], errors="coerce").dropna().values
        nw_vals = pd.to_numeric(X_nw[col], errors="coerce").dropna().values

        if len(tr_vals) < 2 or len(nw_vals) < 2:
            continue

        ks_stat, ks_pval = stats.ks_2samp(tr_vals, nw_vals)

        records.append({
            "feature"      : col,
            "ks_stat"      : round(ks_stat, 4),
            "ks_pval"      : round(ks_pval, 4),
            "train_mean"   : round(float(tr_vals.mean()), 3),
            "new_mean"     : round(float(nw_vals.mean()), 3),
            "train_std"    : round(float(tr_vals.std()),  3),
            "new_std"      : round(float(nw_vals.std()),  3),
            "train_min"    : round(float(tr_vals.min()),  3),
            "new_min"      : round(float(nw_vals.min()),  3),
            "train_max"    : round(float(tr_vals.max()),  3),
            "new_max"      : round(float(nw_vals.max()),  3),
            "mean_diff_pct": round(
                abs(float(tr_vals.mean()) - float(nw_vals.mean())) /
                (abs(float(tr_vals.mean())) + 1e-9) * 100, 2
            ),
            "drifted"      : ks_pval < 0.05,
        })

    drift_df          = (
        pd.DataFrame(records)
        .sort_values("ks_stat", ascending=False)
        .reset_index(drop=True)
    )
    drift_df["rank"]  = drift_df.index + 1
    drifted_cols      = drift_df[drift_df["drifted"]]["feature"].tolist()
    stable_cols       = drift_df[~drift_df["drifted"]]["feature"].tolist()
    target_ks, target_pval = stats.ks_2samp(y_tr, y_nw)

    # ── Console report ────────────────────────────────────────────
    sep = "─" * 65
    print(f"\n{sep}")
    print(f"  Data Distribution Comparison Report")
    print(f"{sep}")
    print(f"  Train samples    : {len(X_tr):,}")
    print(f"  New   samples    : {len(X_nw):,}")
    print(f"  Features checked : {len(drift_df)}")
    print(f"  Drifted (p<0.05) : {len(drifted_cols)}"
          + ("  ⚠️" if drifted_cols else "  ✅"))
    print(f"  Stable           : {len(stable_cols)}  ✅")
    print(f"\n  Target ({target_col}):")
    print(f"    Train mean/std  : {y_tr.mean():.2f} / {y_tr.std():.2f}")
    print(f"    New   mean/std  : {y_nw.mean():.2f} / {y_nw.std():.2f}")
    print(f"    KS stat         : {target_ks:.4f}  "
          + ("⚠️  drift" if target_pval < 0.05 else "✅ stable"))
    print(f"\n  Top 10 drifted features:")
    print(f"  {'Rank':>4}  {'Feature':<38} {'KS':>7}  {'p-val':>8}  Status")
    print(f"  {'─'*4}  {'─'*38} {'─'*7}  {'─'*8}  {'─'*10}")
    for _, row in drift_df.head(10).iterrows():
        status = "⚠️  DRIFT" if row["drifted"] else "✅ stable"
        print(f"  {int(row['rank']):>4}  {row['feature']:<38} "
              f"{row['ks_stat']:>7.4f}  {row['ks_pval']:>8.4f}  {status}")
    print(f"{sep}\n")

    # ── Figure ────────────────────────────────────────────────────
    fig = plt.figure(figsize=(20, 18))
    gs  = gridspec.GridSpec(
        3, 2,
        hspace=0.45, wspace=0.30,
        height_ratios=[1.2, 1, 1]
    )
    ax1 = fig.add_subplot(gs[0, :])
    ax2 = fig.add_subplot(gs[1, 0])
    ax3 = fig.add_subplot(gs[1, 1])

    # ── Panel 1: KS drift bar chart ───────────────────────────────
    top_plot   = drift_df.head(top_n_features).iloc[::-1].copy()
    bar_colors = ["#e05c5c" if d else "#4c8eda"
                  for d in top_plot["drifted"]]

    bars = ax1.barh(
        top_plot["feature"], top_plot["ks_stat"],
        color=bar_colors, edgecolor="white",
        linewidth=0.5, height=0.7,
    )
    for bar, val in zip(bars, top_plot["ks_stat"]):
        ax1.text(
            bar.get_width() + 0.003,
            bar.get_y() + bar.get_height() / 2,
            f"{val:.3f}", va="center", ha="left", fontsize=8
        )

    ax1.axvline(0.1, color="#f0a500", linewidth=1.5,
                linestyle="--", label="Moderate drift (0.1)")
    ax1.axvline(0.2, color="#e05c5c", linewidth=1.5,
                linestyle="--", label="High drift (0.2)")
    ax1.set_xlabel(
        "KS Statistic  (0 = identical · 1 = completely different)",
        fontsize=10
    )
    ax1.set_title(
        f"Feature Drift — Train vs New  "
        f"({len(drifted_cols)}/{len(drift_df)} drifted  p<0.05)  "
        f"🔴 Drifted   🔵 Stable",
        fontsize=12, fontweight="bold"
    )
    ax1.legend(fontsize=9, loc="lower right")
    ax1.grid(True, axis="x", alpha=0.3)
    ax1.set_facecolor("#fafafa")
    ax1.tick_params(axis="y", labelsize=8)

    # ── Panel 2: Target distribution ─────────────────────────────
    xr_target = np.linspace(
        min(float(y_tr.min()), float(y_nw.min())),
        max(float(y_tr.max()), float(y_nw.max())),
        300
    )
    bins = min(40, max(10, len(y_nw) // 3))

    for vals, color, label, ls, alpha in [
        (y_tr.values, train_color, f"Train  μ={y_tr.mean():.1f}", "-",   0.3),
        (y_nw.values, new_color,   f"New    μ={y_nw.mean():.1f}", "--",  0.4),
    ]:
        ax2.hist(vals, bins=bins, density=True,
                 color=color, alpha=alpha, edgecolor="none")
        ax2.plot(xr_target, safe_kde(vals, xr_target),   # ← safe_kde
                 color=color, linewidth=2.5, linestyle=ls, label=label)

    ax2.axvline(float(y_tr.mean()), color=train_color,
                linewidth=1, linestyle=":", alpha=0.7)
    ax2.axvline(float(y_nw.mean()), color=new_color,
                linewidth=1, linestyle=":", alpha=0.7)
    ax2.set_xlabel(target_col, fontsize=10)
    ax2.set_ylabel("Density",   fontsize=10)
    ax2.set_title(
        f"Target Distribution — {target_col}\n"
        f"KS = {target_ks:.4f}  "
        + ("⚠️  Drift detected" if target_pval < 0.05
           else "✅ No significant drift"),
        fontsize=11, fontweight="bold"
    )
    ax2.legend(fontsize=9)
    ax2.grid(True, alpha=0.25)
    ax2.set_facecolor("#fafafa")

    # ── Panel 3: Mean comparison scatter ──────────────────────────
    sc = ax3.scatter(
        drift_df["train_mean"], drift_df["new_mean"],
        c=drift_df["ks_stat"], cmap="RdYlGn_r",
        s=50, alpha=0.7, edgecolors="white", linewidths=0.3, zorder=3
    )
    mn = min(drift_df["train_mean"].min(), drift_df["new_mean"].min())
    mx = max(drift_df["train_mean"].max(), drift_df["new_mean"].max())
    ax3.plot([mn, mx], [mn, mx], color="black",
             linewidth=1.5, linestyle="--", label="Perfect match")
    plt.colorbar(sc, ax=ax3, label="KS statistic", shrink=0.8)
    ax3.set_xlabel("Train mean", fontsize=10)
    ax3.set_ylabel("New mean",   fontsize=10)
    ax3.set_title(
        "Feature Means — Train vs New\n"
        "(diagonal = no change in mean)",
        fontsize=11, fontweight="bold"
    )
    ax3.legend(fontsize=9)
    ax3.grid(True, alpha=0.25)
    ax3.set_facecolor("#fafafa")

    # ── Panel 4: Top 6 drifted features ───────────────────────────
    top6      = drift_df.head(6)["feature"].tolist()
    inner_gs  = gridspec.GridSpecFromSubplotSpec(
        1, 6, subplot_spec=gs[2, :], wspace=0.4
    )

    for i, col in enumerate(top6):
        ax_sub  = fig.add_subplot(inner_gs[i])
        tr_vals = pd.to_numeric(
            X_tr[col], errors="coerce"
        ).dropna().values
        nw_vals = pd.to_numeric(
            X_nw[col], errors="coerce"
        ).dropna().values

        if len(tr_vals) < 2 or len(nw_vals) < 2:
            ax_sub.set_title(f"{col[:18]}\n(no data)", fontsize=8)
            continue

        # add tiny jitter to avoid zero-range linspace
        lo = min(tr_vals.min(), nw_vals.min())
        hi = max(tr_vals.max(), nw_vals.max())
        if hi - lo < 1e-9:
            hi = lo + 1.0

        xr = np.linspace(lo, hi, 200)

        kde_tr = safe_kde(tr_vals, xr)      # ← safe_kde
        kde_nw = safe_kde(nw_vals, xr)      # ← safe_kde

        ax_sub.plot(xr, kde_tr, color=train_color,
                    linewidth=2, label="Train")
        ax_sub.plot(xr, kde_nw, color=new_color,
                    linewidth=2, linestyle="--", label="New")
        ax_sub.fill_between(xr, kde_tr, alpha=0.15, color=train_color)
        ax_sub.fill_between(xr, kde_nw, alpha=0.15, color=new_color)

        ks_row  = drift_df[drift_df["feature"] == col].iloc[0]
        ax_sub.set_title(
            f"{col[:18]}\nKS={ks_row['ks_stat']:.3f} "
            + ("⚠️" if ks_row["drifted"] else "✅"),
            fontsize=8, fontweight="bold"
        )
        ax_sub.tick_params(labelsize=7)
        ax_sub.set_ylabel("Density" if i == 0 else "", fontsize=8)
        ax_sub.grid(True, alpha=0.2)
        ax_sub.set_facecolor("#fafafa")
        if i == 0:
            ax_sub.legend(fontsize=7)

    # ── Super title ───────────────────────────────────────────────
    fig.suptitle(
        f"Distribution Comparison — Train vs New Data\n"
        f"Train n={len(X_tr):,}  │  New n={len(X_nw):,}  │  "
        f"Drifted: {len(drifted_cols)}/{len(drift_df)}  │  "
        f"Target KS={target_ks:.4f}",
        fontsize=13, fontweight="bold", y=1.01
    )

    plt.tight_layout()

    if save:
        import os
        os.makedirs(save_dir, exist_ok=True)
        path = os.path.join(save_dir, "distribution_comparison.png")
        plt.savefig(path, dpi=150, bbox_inches="tight")
        print(f"✅ Saved → {path}")

    plt.show()

    return {
        "drift_df"     : drift_df,
        "drifted_cols" : drifted_cols,
        "stable_cols"  : stable_cols,
        "target_ks"    : target_ks,
        "target_pvalue": target_pval,
    }


def predict_new_data(
    df_new: pd.DataFrame,
    model_path: str,
    train_columns: list,
    target_col: str = "KgCO2EQ",
    encoding_method: str = "auto",
    verbose: bool = True,
) -> pd.DataFrame:
    """
    Load the saved model, preprocess new raw data, and return predictions.

    Parameters
    ----------
    df_new         : raw new DataFrame (same columns as original df)
    model_path     : path to saved .pkl model file
    train_columns  : X_train.columns.tolist()
    target_col     : name of the target column
    encoding_method: encoding method used during training
    verbose        : print report

    Returns
    -------
    results_df : pd.DataFrame
        Original df_new with a new column "predicted_{target_col}"
    """

    # ── Load model ────────────────────────────────────────────────
    model = joblib.load(model_path)
    if verbose:
        print(f"✅ Model loaded from → {model_path}")

    # ── Preprocess ────────────────────────────────────────────────
    X_new = preprocess_new_data(
        df_new,
        train_columns   = train_columns,
        target_col      = target_col,
        encoding_method = encoding_method,
        verbose         = verbose,
    )

    # ── Predict ───────────────────────────────────────────────────
    predictions = model.predict(X_new)

    # ── Attach predictions to original data ───────────────────────
    results_df = df_new.copy()
    results_df[f"predicted_{target_col}"] = predictions.round(2)

    if verbose:
        print(f"\n  Predictions summary:")
        print(f"  n predictions : {len(predictions):,}")
        print(f"  min           : {predictions.min():.2f}")
        print(f"  max           : {predictions.max():.2f}")
        print(f"  mean          : {predictions.mean():.2f}")
        print(f"  median        : {np.median(predictions):.2f}")

    return results_df


def shap_analysis(
    model,
    X_train: pd.DataFrame,
    X_test: pd.DataFrame,
    model_name: str = "Random Forest",
    top_n: int = 20,
    max_display: int = 20,
    sample_background: int = 100,
    figsize_summary: tuple = (12, 8),
    figsize_bar: tuple = (10, 8),
    figsize_dependence: tuple = (14, 10),
    figsize_waterfall: tuple = (12, 8),
    colormap: str = "RdBu_r",
    save: bool = False,
    save_dir: str = ".",
) -> dict:
    """
    Comprehensive SHAP (SHapley Additive exPlanations) analysis for a
    fitted Random Forest regression pipeline.

    SHAP theory
    -----------
    SHAP values are grounded in cooperative game theory. For a prediction
    f(x), the SHAP value phi_j for feature j is:

        phi_j = sum_{S in F\\{j}} |S|!(|F|-|S|-1)!/|F|! * [f(S+{j}) - f(S)]

    where F is the full feature set and S is every possible subset that
    does NOT contain j. This computes the weighted average marginal
    contribution of feature j across all possible feature orderings.

    For tree-based models, TreeExplainer computes exact SHAP values in
    O(TLD^2) time where T = trees, L = max leaves, D = max depth.

    Plots produced
    --------------
    Plot 1 — Beeswarm
        Each dot = one test sample. X = SHAP value (impact on prediction).
        Color = feature value (red=high, blue=low). Y = feature sorted by
        mean |SHAP|.

    Plot 2 — Bar chart (global importance)
        Mean |SHAP| per feature. More robust than MDI feature_importances_
        because it is not biased toward high-cardinality features.

    Plot 3 — Dependence plots (top 6 features)
        X = feature value. Y = SHAP value. Color = interaction feature.
        Reveals linear, non-linear, and interaction effects.

    Plot 4 — Waterfall (single prediction)
        Shows how each feature pushes ONE prediction from E[f(x)] to f(x).
        Red = pushes up. Blue = pushes down. Good for explaining one decision.

    Plot 5 — Heatmap (top N features)
        Rows = features. Columns = samples sorted by predicted value.
        Color = SHAP value. Reveals patterns and subgroups across all samples.

    Parameters
    ----------
    model : fitted sklearn Pipeline
        Must contain "imputer", "scaler", "model" steps where "model"
        is a tree-based estimator (RandomForestRegressor).
    X_train : pd.DataFrame
        Training features used as background for TreeExplainer.
    X_test : pd.DataFrame
        Test features — SHAP values computed on this set.
    model_name : str, optional
        Shown in plot titles. Default "Random Forest".
    top_n : int, optional
        Top features shown in bar and dependence plots. Default 20.
    max_display : int, optional
        Max features in beeswarm and bar plots. Default 20.
    sample_background : int, optional
        Background samples for TreeExplainer. Default 100.
        Larger = more accurate but slower. Use 50-200.
    figsize_summary : tuple, optional
        Figure size for beeswarm. Default (12, 8).
    figsize_bar : tuple, optional
        Figure size for bar chart. Default (10, 8).
    figsize_dependence : tuple, optional
        Figure size for dependence grid. Default (14, 10).
    figsize_waterfall : tuple, optional
        Figure size for waterfall. Default (12, 8).
    colormap : str, optional
        Colormap for heatmap. Default "RdBu_r".
    save : bool, optional
        Save all figures as PNG. Default False.
    save_dir : str, optional
        Output directory. Default current directory.

    Returns
    -------
    results : dict
        shap_values   : np.ndarray   — SHAP matrix (n_test x n_features)
        base_value    : float        — E[f(x)] model baseline
        importance_df : pd.DataFrame — features ranked by mean |SHAP|
        explainer     : shap.TreeExplainer — fitted explainer object
        X_test_proc   : pd.DataFrame — preprocessed test features

    Examples
    --------
    >>> results = shap_analysis(final_model, X_train, X_test)

    >>> results = shap_analysis(
    ...     final_model, X_train, X_test,
    ...     model_name        = "Tuned RF",
    ...     top_n             = 25,
    ...     sample_background = 150,
    ...     save              = True,
    ...     save_dir          = "outputs",
    ... )

    >>> results["importance_df"].head(10)
    >>> results["shap_values"]       # shape (n_test, n_features)
    >>> results["base_value"]        # E[f(x)]
    """

    if not HAS_SHAP:
        raise ImportError("shap is required. Run:  pip install shap")

    if not hasattr(model, "named_steps") or "model" not in model.named_steps:
        raise ValueError(
            f"Pipeline must contain a step named 'model'.\n"
            f"Found: {list(model.named_steps.keys()) if hasattr(model, 'named_steps') else 'not a pipeline'}"
        )

    import os
    if save:
        os.makedirs(save_dir, exist_ok=True)

    # ── Step 1: preprocess through pipeline (all steps except model) ─
    preprocessor_steps = list(model.named_steps.keys())[:-1]
    rf_model           = model.named_steps["model"]
    feature_names      = X_train.columns.tolist()

    X_train_proc = X_train.values.copy()
    X_test_proc  = X_test.values.copy()

    for step_name in preprocessor_steps:
        step         = model.named_steps[step_name]
        X_train_proc = step.transform(X_train_proc)
        X_test_proc  = step.transform(X_test_proc)

    X_train_proc = pd.DataFrame(X_train_proc, columns=feature_names)
    X_test_proc  = pd.DataFrame(X_test_proc,  columns=feature_names)

    # ── Step 2: background sample for TreeExplainer ───────────────
    n_bg = min(sample_background, len(X_train_proc))
    bg   = shap.sample(X_train_proc, n_bg, random_state=42)

    sep = "─" * 62
    print(f"\n{sep}")
    print(f"  SHAP Analysis — {model_name}")
    print(f"{sep}")
    print(f"  Test samples       : {len(X_test_proc):,}")
    print(f"  Features           : {len(feature_names)}")
    print(f"  Background samples : {n_bg}")
    print(f"  Explainer          : TreeExplainer (exact, fast)")
    print(f"  Computing SHAP values ⏳  ...")

    # ── Step 3: TreeExplainer ─────────────────────────────────────
    explainer   = shap.TreeExplainer(rf_model, data=bg)
    shap_values = explainer.shap_values(X_test_proc)
    base_value  = float(explainer.expected_value)

    print(f"  Done ✅")
    print(f"  Base value E[f(x)] : {base_value:.2f}")
    print(f"  SHAP matrix shape  : {shap_values.shape}")

    # ── Step 4: importance DataFrame ─────────────────────────────
    mean_abs_shap = np.abs(shap_values).mean(axis=0)

    importance_df = (
        pd.DataFrame({
            "feature"      : feature_names,
            "mean_abs_shap": mean_abs_shap,
            "mean_shap"    : shap_values.mean(axis=0),
            "max_shap"     : shap_values.max(axis=0),
            "min_shap"     : shap_values.min(axis=0),
            "std_shap"     : shap_values.std(axis=0),
        })
        .sort_values("mean_abs_shap", ascending=False)
        .reset_index(drop=True)
    )
    importance_df["rank"]           = importance_df.index + 1
    importance_df["cumulative"]     = importance_df["mean_abs_shap"].cumsum()
    importance_df["cumulative_pct"] = (
        importance_df["cumulative"] /
        importance_df["mean_abs_shap"].sum() * 100
    ).round(2)

    # ── Console importance table ──────────────────────────────────
    print(f"\n  Top 15 features by mean |SHAP|:\n")
    print(f"  {'Rank':>4}  {'Feature':<38} {'|SHAP|':>8}  Direction")
    print(f"  {'─'*4}  {'─'*38} {'─'*8}  {'─'*12}")
    for _, row in importance_df.head(15).iterrows():
        direction = "↑ positive" if row["mean_shap"] > 0 else "↓ negative"
        bar       = "█" * int(row["mean_abs_shap"] / mean_abs_shap.max() * 18)
        print(
            f"  {int(row['rank']):>4}  {row['feature']:<38} "
            f"{row['mean_abs_shap']:>8.4f}  {direction}  {bar}"
        )

    n_for_80 = (importance_df["cumulative_pct"] <= 80).sum() + 1
    print(f"\n  Features for 80% SHAP importance : {n_for_80}")
    print(f"{sep}\n")

    # ── Build shap.Explanation with base_values as array ─────────
    # base_values MUST be np.ndarray — plain float causes AttributeError
    shap_explanation = shap.Explanation(
        values        = shap_values,
        base_values   = np.full(len(shap_values), base_value),  # ← array
        data          = X_test_proc.values,
        feature_names = feature_names,
    )

    # ── Plot 1: Beeswarm ──────────────────────────────────────────
    print("  Plotting beeswarm ...")
    plt.figure(figsize=figsize_summary)
    shap.plots.beeswarm(
        shap_explanation,
        max_display = max_display,
        show        = False,
    )
    plt.title(
        f"SHAP Beeswarm — {model_name}\n"
        f"Each dot = one test sample  ·  X = impact on prediction  ·  "
        f"Color = feature value (red=high · blue=low)",
        fontsize=12, fontweight="bold", pad=14
    )
    plt.tight_layout()
    if save:
        path = os.path.join(save_dir, "shap_beeswarm.png")
        plt.savefig(path, dpi=150, bbox_inches="tight")
        print(f"  ✅ Saved → {path}")
    plt.show()

    # ── Plot 2: Bar chart ─────────────────────────────────────────
    print("  Plotting bar chart ...")
    plt.figure(figsize=figsize_bar)
    shap.plots.bar(
        shap_explanation,
        max_display = max_display,
        show        = False,
    )
    plt.title(
        f"SHAP Global Importance — {model_name}\n"
        f"Mean |SHAP value| per feature  ·  "
        f"Unbiased alternative to MDI feature_importances_",
        fontsize=12, fontweight="bold", pad=14
    )
    plt.tight_layout()
    if save:
        path = os.path.join(save_dir, "shap_bar.png")
        plt.savefig(path, dpi=150, bbox_inches="tight")
        print(f"  ✅ Saved → {path}")
    plt.show()

    # ── Plot 3: Dependence plots (top 6) ─────────────────────────
    print("  Plotting dependence plots ...")
    top6 = importance_df["feature"].head(6).tolist()

    fig, axes = plt.subplots(2, 3, figsize=figsize_dependence)
    axes      = axes.flatten()

    for i, feat in enumerate(top6):
        feat_idx  = feature_names.index(feat)
        ax        = axes[i]
        feat_vals = X_test_proc[feat].values
        shap_vals = shap_values[:, feat_idx]

        # auto-select interaction feature (highest mean |SHAP|, not feat itself)
        interaction_idx = np.abs(shap_values).mean(axis=0).argsort()[::-1]
        color_idx   = next(j for j in interaction_idx if j != feat_idx)
        color_vals  = X_test_proc.iloc[:, color_idx].values
        color_feat  = feature_names[color_idx]

        sc = ax.scatter(
            feat_vals, shap_vals,
            c=color_vals, cmap="coolwarm",
            s=20, alpha=0.6, edgecolors="none"
        )
        ax.axhline(0, color="black", linewidth=1,
                   linestyle="--", alpha=0.5)
        ax.set_xlabel(feat,         fontsize=9)
        ax.set_ylabel("SHAP value", fontsize=9)
        ax.set_title(
            f"{feat[:30]}\n(color = {color_feat[:25]})",
            fontsize=9, fontweight="bold"
        )
        plt.colorbar(sc, ax=ax, label=color_feat[:20], shrink=0.8)
        ax.grid(True, alpha=0.2)
        ax.set_facecolor("#fafafa")

    fig.suptitle(
        f"SHAP Dependence Plots — Top 6 Features — {model_name}\n"
        f"X = feature value  ·  Y = SHAP impact on prediction  ·  "
        f"Color = interaction feature",
        fontsize=12, fontweight="bold", y=1.02
    )
    plt.tight_layout()
    if save:
        path = os.path.join(save_dir, "shap_dependence.png")
        plt.savefig(path, dpi=150, bbox_inches="tight")
        print(f"  ✅ Saved → {path}")
    plt.show()

    # ── Plot 4: Waterfall — most impactful sample ─────────────────
    print("  Plotting waterfall ...")
    total_shap = np.abs(shap_values).sum(axis=1)
    sample_idx = int(np.argmax(total_shap))

    # base_values for single sample — use plain float
    sample_explanation = shap.Explanation(
        values        = shap_values[sample_idx],
        base_values   = base_value,                    # ← plain float for single
        data          = X_test_proc.iloc[sample_idx].values,
        feature_names = feature_names,
    )

    plt.figure(figsize=figsize_waterfall)
    shap.plots.waterfall(
        sample_explanation,
        max_display = max_display,
        show        = False,
    )
    pred_value = base_value + shap_values[sample_idx].sum()
    plt.title(
        f"SHAP Waterfall — Most Impactful Prediction (sample #{sample_idx})\n"
        f"Base E[f(x)] = {base_value:.2f}  →  f(x) = {pred_value:.2f}",
        fontsize=11, fontweight="bold", pad=14
    )
    plt.tight_layout()
    if save:
        path = os.path.join(save_dir, "shap_waterfall.png")
        plt.savefig(path, dpi=150, bbox_inches="tight")
        print(f"  ✅ Saved → {path}")
    plt.show()

    # ── Plot 5: Heatmap ───────────────────────────────────────────
    print("  Plotting heatmap ...")
    top_n_actual = min(top_n, len(feature_names))
    top_feats    = importance_df["feature"].head(top_n_actual).tolist()
    top_n_idx    = [feature_names.index(f) for f in top_feats]
    shap_top     = shap_values[:, top_n_idx]

    # sort samples by total predicted value (ascending)
    pred_order  = np.argsort(shap_values.sum(axis=1))
    shap_sorted = shap_top[pred_order, :]
    vmax        = np.abs(shap_top).max()

    fig_h = max(6,  top_n_actual // 2)
    fig_w = max(14, len(X_test_proc) // 20)
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))

    im = ax.imshow(
        shap_sorted.T,
        aspect        = "auto",
        cmap          = colormap,
        interpolation = "nearest",
        vmin          = -vmax,
        vmax          =  vmax,
    )
    ax.set_yticks(range(top_n_actual))
    ax.set_yticklabels(top_feats, fontsize=8)
    ax.set_xlabel(
        "Samples sorted by predicted value (low → high)",
        fontsize=10
    )
    ax.set_ylabel("Feature", fontsize=10)
    ax.set_title(
        f"SHAP Heatmap — Top {top_n_actual} Features — {model_name}\n"
        f"Red = pushes prediction UP  ·  Blue = pushes prediction DOWN  ·  "
        f"Columns sorted by predicted value",
        fontsize=12, fontweight="bold", pad=14
    )
    plt.colorbar(im, ax=ax, label="SHAP value", shrink=0.6)
    plt.tight_layout()
    if save:
        path = os.path.join(save_dir, "shap_heatmap.png")
        plt.savefig(path, dpi=150, bbox_inches="tight")
        print(f"  ✅ Saved → {path}")
    plt.show()

    # ── Final summary ─────────────────────────────────────────────
    print(f"{sep}")
    print(f"  SHAP analysis complete ✅")
    print(f"  Plots: beeswarm · bar · dependence · waterfall · heatmap")
    if save:
        print(f"  All saved to: {save_dir}/")
    print(f"{sep}\n")

    return {
        "shap_values"  : shap_values,
        "base_value"   : base_value,
        "importance_df": importance_df,
        "explainer"    : explainer,
        "X_test_proc"  : X_test_proc,
    }


